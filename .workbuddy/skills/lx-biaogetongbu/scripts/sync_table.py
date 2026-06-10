#!/usr/bin/env python3
"""Sync rows between local Excel files or Tencent Docs online sheets."""

from __future__ import annotations

import argparse
import copy
import json
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from online_backends import OnlineBackendError, build_online_backend


SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
MAX_ONLINE_ROWS = 1000
MAX_ONLINE_COLUMNS = 200
MAX_ONLINE_CELLS = 10000
MAX_BATCH_REQUESTS = 5


class SyncError(Exception):
    """Expected user-facing sync failure."""


@dataclass
class RowRecord:
    values: dict[str, Any]
    row_number: int


@dataclass
class TableData:
    label: str
    sheet_name: str
    header_row: int
    headers: dict[str, int]
    rows: list[RowRecord]


@dataclass
class OnlineSheet:
    file_id: str
    sheet_id: str
    title: str
    row_count: int
    column_count: int


@dataclass
class UpdateCell:
    key_text: str
    source_row_number: int
    target_row_number: int
    target_column: str
    target_column_number: int
    old_value: Any
    new_value: Any


def find_project_root(start: Path | None = None) -> Path:
    current = (start or Path.cwd()).resolve()
    for candidate in [current, *current.parents]:
        if (candidate / ".workbuddy" / "skills").is_dir() and (candidate / "workspace").is_dir():
            return candidate
    return Path.cwd().resolve()


PROJECT_ROOT = find_project_root()


def resolve_path(value: str | None, *, required: bool = False) -> Path | None:
    if not value:
        if required:
            raise SyncError("缺少必填路径参数")
        return None
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path.resolve()


def load_profile(path: Path | None) -> dict[str, Any]:
    if not path:
        return {}
    if not path.exists():
        raise SyncError(f"profile 不存在: {path}")
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, dict):
        raise SyncError(f"profile 必须是 JSON object: {path}")
    return data


def first_value(cli_value: Any, profile: dict[str, Any], key: str, default: Any = None) -> Any:
    if cli_value not in (None, "", []):
        return cli_value
    return profile.get(key, default)


def parse_pairs(values: list[str] | None, label: str) -> dict[str, str]:
    pairs: dict[str, str] = {}
    for raw in values or []:
        if "=" not in raw:
            raise SyncError(f"{label} 参数格式错误，应为 左=右: {raw}")
        left, right = raw.split("=", 1)
        left = left.strip()
        right = right.strip()
        if not left or not right:
            raise SyncError(f"{label} 参数不能为空: {raw}")
        pairs[left] = right
    return pairs


def profile_mapping(value: Any) -> dict[str, str]:
    if not value:
        return {}
    if isinstance(value, dict):
        return {str(k): str(v) for k, v in value.items()}
    if isinstance(value, list):
        return parse_pairs([str(item) for item in value], "profile.mapping")
    raise SyncError("profile.mapping 必须是 object 或 list")


def split_values(values: list[str] | str | None) -> list[str]:
    if not values:
        return []
    raw_values = [values] if isinstance(values, str) else values
    result: list[str] = []
    for raw in raw_values:
        for part in str(raw).split(","):
            value = part.strip()
            if value:
                result.append(value)
    return result


def normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def normalize_key_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def values_equal(left: Any, right: Any) -> bool:
    return normalize_key_value(left) == normalize_key_value(right)


def is_blank(value: Any) -> bool:
    return value is None or value == ""


def row_is_blank(values: dict[str, Any]) -> bool:
    return all(is_blank(value) for value in values.values())


def require_excel(path: Path, label: str) -> None:
    if not path.exists():
        raise SyncError(f"{label} 不存在: {path}")
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise SyncError(f"{label} 只支持 .xlsx / .xlsm: {path}")


def select_sheet(workbook: Any, sheet_name: str | None, label: str) -> Worksheet:
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise SyncError(f"{label} sheet 不存在: {sheet_name}")
        return workbook[sheet_name]
    return workbook.active


def read_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    duplicates: list[str] = []
    for cell in ws[header_row]:
        name = normalize_header(cell.value)
        if not name:
            continue
        if name in headers:
            duplicates.append(name)
            continue
        headers[name] = cell.column
    if duplicates:
        raise SyncError(f"{ws.title} 表头存在重复列: {', '.join(sorted(set(duplicates)))}")
    if not headers:
        raise SyncError(f"{ws.title} 第 {header_row} 行没有可用表头")
    return headers


def read_row_records(ws: Worksheet, headers: dict[str, int], header_row: int) -> list[RowRecord]:
    rows: list[RowRecord] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = {header: ws.cell(row=row_idx, column=col_idx).value for header, col_idx in headers.items()}
        if row_is_blank(row):
            continue
        rows.append(RowRecord(row, row_idx))
    return rows


def add_missing_target_columns(
    ws: Worksheet,
    target_headers: dict[str, int],
    needed_columns: list[str],
    header_row: int,
) -> dict[str, int]:
    next_col = max(target_headers.values(), default=0) + 1
    for column_name in needed_columns:
        if column_name in target_headers:
            continue
        ws.cell(row=header_row, column=next_col).value = column_name
        target_headers[column_name] = next_col
        next_col += 1
    return target_headers


def build_mapping(
    source_headers: dict[str, int],
    target_headers: dict[str, int],
    explicit_mapping: dict[str, str],
) -> dict[str, str]:
    if explicit_mapping:
        missing_source = [name for name in explicit_mapping if name not in source_headers]
        if missing_source:
            raise SyncError(f"A 表缺少映射源列: {', '.join(missing_source)}")
        missing_target = [name for name in explicit_mapping.values() if name not in target_headers]
        if missing_target:
            raise SyncError(f"B 表缺少映射目标列: {', '.join(missing_target)}")
        return dict(explicit_mapping)

    common = sorted(set(source_headers) & set(target_headers))
    if not common:
        raise SyncError("A 表和 B 表没有同名列，请使用 --map 显式指定字段映射")
    return {name: name for name in common}


def ensure_key_mapping(
    mapping: dict[str, str],
    keys: list[str],
    source_headers: dict[str, int],
    target_headers: dict[str, int],
) -> dict[str, str]:
    resolved = dict(mapping)
    for key in keys:
        if key in source_headers and key in target_headers and key not in resolved.values():
            resolved[key] = key
    return resolved


def resolve_key_columns(keys: list[str], mapping: dict[str, str], target_headers: dict[str, int]) -> list[str]:
    resolved: list[str] = []
    for key in keys:
        if key in target_headers:
            resolved.append(key)
        elif key in mapping:
            target_key = mapping[key]
            if target_key not in target_headers:
                raise SyncError(f"去重键映射到不存在的 B 表列: {key} -> {target_key}")
            resolved.append(target_key)
        else:
            raise SyncError(f"去重键不在 B 表列或映射源列中: {key}")
    return resolved


def target_key(row: dict[str, Any], key_columns: list[str]) -> tuple[str, ...]:
    return tuple(normalize_key_value(row.get(column)) for column in key_columns)


def key_text(key: tuple[str, ...]) -> str:
    return " | ".join(key)


def make_target_records(
    source_rows: list[RowRecord],
    mapping: dict[str, str],
    literals: dict[str, str],
) -> list[RowRecord]:
    result: list[RowRecord] = []
    for source_row in source_rows:
        target_row = {target: source_row.values.get(source) for source, target in mapping.items()}
        target_row.update(literals)
        result.append(RowRecord(target_row, source_row.row_number))
    return result


def existing_keys(
    target_rows: list[RowRecord],
    key_columns: list[str],
) -> set[tuple[str, ...]]:
    keys: set[tuple[str, ...]] = set()
    for row in target_rows:
        key = target_key(row.values, key_columns)
        if any(key):
            keys.add(key)
    return keys


def filter_append_rows(
    rows: list[RowRecord],
    key_columns: list[str],
    seen: set[tuple[str, ...]],
) -> tuple[list[dict[str, Any]], list[tuple[str, str]]]:
    append_rows: list[dict[str, Any]] = []
    skipped: list[tuple[str, str]] = []
    batch_seen: set[tuple[str, ...]] = set()

    for index, row in enumerate(rows, 1):
        if not key_columns:
            append_rows.append(row.values)
            continue

        key = target_key(row.values, key_columns)
        text = key_text(key)
        if not any(key):
            skipped.append((str(index), "去重键为空"))
            continue
        if key in seen:
            skipped.append((text, "B 表已存在"))
            continue
        if key in batch_seen:
            skipped.append((text, "A 表本次重复"))
            continue
        batch_seen.add(key)
        append_rows.append(row.values)

    return append_rows, skipped


def resolve_update_columns(
    values: list[str],
    profile: dict[str, Any],
    key_columns: list[str],
    target_headers: dict[str, int],
) -> list[str]:
    columns = split_values(values or profile.get("update_columns", []))
    if not columns:
        raise SyncError("update-by-key 模式必须指定 --update-column 或 profile.update_columns")
    missing = [column for column in columns if column not in target_headers]
    if missing:
        raise SyncError(f"B 表缺少更新目标列: {', '.join(missing)}")
    key_overlap = [column for column in columns if column in key_columns]
    if key_overlap:
        raise SyncError(f"不允许更新定位键列: {', '.join(key_overlap)}")
    return columns


def build_update_cells(
    source_rows: list[RowRecord],
    target_rows: list[RowRecord],
    key_columns: list[str],
    update_columns: list[str],
    target_headers: dict[str, int],
    *,
    allow_blank_updates: bool = False,
) -> tuple[list[UpdateCell], list[tuple[str, str]], int]:
    if not key_columns:
        raise SyncError("update-by-key 模式必须配置 --key")

    skipped: list[tuple[str, str]] = []
    updates: list[UpdateCell] = []
    unchanged = 0

    source_index: dict[tuple[str, ...], list[RowRecord]] = {}
    for row in source_rows:
        key = target_key(row.values, key_columns)
        if not any(key):
            skipped.append((str(row.row_number), "A 表定位键为空"))
            continue
        source_index.setdefault(key, []).append(row)

    target_index: dict[tuple[str, ...], list[RowRecord]] = {}
    for row in target_rows:
        key = target_key(row.values, key_columns)
        if not any(key):
            continue
        target_index.setdefault(key, []).append(row)

    for key, matched_source_rows in source_index.items():
        text = key_text(key)
        if len(matched_source_rows) > 1:
            skipped.append((text, "A 表重复定位键"))
            continue

        matched_target_rows = target_index.get(key, [])
        if not matched_target_rows:
            skipped.append((text, "B 表无匹配行"))
            continue
        if len(matched_target_rows) > 1:
            skipped.append((text, "B 表重复定位键"))
            continue

        source_row = matched_source_rows[0]
        target_row_record = matched_target_rows[0]
        for column in update_columns:
            new_value = source_row.values.get(column)
            if is_blank(new_value) and not allow_blank_updates:
                skipped.append((f"{text} / {column}", "新值为空，默认不覆盖"))
                continue
            old_value = target_row_record.values.get(column)
            if values_equal(old_value, new_value):
                unchanged += 1
                continue
            updates.append(
                UpdateCell(
                    key_text=text,
                    source_row_number=source_row.row_number,
                    target_row_number=target_row_record.row_number,
                    target_column=column,
                    target_column_number=target_headers[column],
                    old_value=old_value,
                    new_value=new_value,
                )
            )

    return updates, skipped, unchanged


def copy_row_style(ws: Worksheet, from_row: int, to_row: int, max_col: int) -> None:
    if from_row < 1:
        return
    for col in range(1, max_col + 1):
        source = ws.cell(row=from_row, column=col)
        target = ws.cell(row=to_row, column=col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)


def append_to_sheet(ws: Worksheet, rows: list[dict[str, Any]], target_headers: dict[str, int]) -> None:
    header_order = sorted(target_headers.items(), key=lambda item: item[1])
    start_row = ws.max_row + 1
    style_row = ws.max_row if ws.max_row >= 1 else 0
    max_col = max(target_headers.values(), default=0)

    for offset, row_data in enumerate(rows):
        row_idx = start_row + offset
        copy_row_style(ws, style_row, row_idx, max_col)
        for header, col_idx in header_order:
            if header in row_data:
                ws.cell(row=row_idx, column=col_idx).value = row_data[header]


def apply_updates_to_sheet(ws: Worksheet, updates: list[UpdateCell]) -> None:
    for update in updates:
        ws.cell(row=update.target_row_number, column=update.target_column_number).value = update.new_value


def col_to_a1(column: int) -> str:
    if column < 1:
        raise SyncError(f"列号必须从 1 开始: {column}")
    result = ""
    current = column
    while current:
        current, rem = divmod(current - 1, 26)
        result = chr(65 + rem) + result
    return result


def a1_range(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    return f"{col_to_a1(start_col)}{start_row}:{col_to_a1(end_col)}{end_row}"


def extract_tencent_doc_id(value: str) -> str:
    raw = value.strip()
    if not raw:
        raise SyncError("腾讯文档 URL / file_id 为空")
    if not raw.startswith(("http://", "https://")):
        return raw

    parsed = urlparse(raw)
    parts = [part for part in parsed.path.split("/") if part]
    for marker in ("sheet", "doc", "smartsheet", "slide", "mind", "flowchart"):
        if marker in parts:
            index = parts.index(marker)
            if index + 1 < len(parts):
                return parts[index + 1]
    if parts:
        return parts[-1]
    raise SyncError(f"无法从腾讯文档 URL 提取 file_id: {value}")


def find_nested_key(value: Any, names: set[str]) -> Any:
    if isinstance(value, dict):
        for key, item in value.items():
            if key in names and item:
                return item
        for item in value.values():
            found = find_nested_key(item, names)
            if found:
                return found
    elif isinstance(value, list):
        for item in value:
            found = find_nested_key(item, names)
            if found:
                return found
    return None


def load_saas_client(args: argparse.Namespace) -> Any:
    try:
        return build_online_backend(args)
    except OnlineBackendError as exc:
        raise SyncError(str(exc)) from exc


def select_online_sheet(properties: list[dict[str, Any]], tab: str | None, label: str) -> OnlineSheet:
    if not properties:
        raise SyncError(f"{label} 没有可用 sheet")

    selected: dict[str, Any] | None = None
    if tab:
        for item in properties:
            if str(item.get("sheet_id", "")) == str(tab) or str(item.get("title", "")) == str(tab):
                selected = item
                break
        if selected is None:
            available = ", ".join(str(item.get("title", item.get("sheet_id", ""))) for item in properties)
            raise SyncError(f"{label} sheet 不存在: {tab}；可用 sheet: {available}")
    else:
        selected = properties[0]

    sheet_id = str(selected.get("sheet_id") or "").strip()
    if not sheet_id:
        raise SyncError(f"{label} sheet 缺少 sheet_id")
    return OnlineSheet(
        file_id="",
        sheet_id=sheet_id,
        title=str(selected.get("title") or sheet_id),
        row_count=int(selected.get("row_count") or 1),
        column_count=int(selected.get("column_count") or selected.get("column_total") or 1),
    )


def cell_to_python(cell: Any) -> Any:
    if not isinstance(cell, dict):
        return None
    cell_value = cell.get("cell_value")
    if not isinstance(cell_value, dict):
        return None
    if "text" in cell_value:
        return cell_value.get("text")
    if "number" in cell_value:
        return cell_value.get("number")
    if "link" in cell_value:
        link = cell_value.get("link")
        if isinstance(link, dict):
            return link.get("text") or link.get("url")
    if "time" in cell_value:
        return cell_value.get("time")
    if "location" in cell_value:
        location = cell_value.get("location")
        if isinstance(location, dict):
            return location.get("name")
    if "select" in cell_value:
        selected = cell_value.get("select")
        if isinstance(selected, dict):
            value = selected.get("value")
            return ",".join(str(item) for item in value) if isinstance(value, list) else value
    return None


def python_to_cell(value: Any) -> dict[str, Any]:
    if value is None:
        return {"cell_value": {"text": ""}}
    if isinstance(value, bool):
        return {"cell_value": {"text": str(value)}}
    if isinstance(value, (int, float)):
        return {"cell_value": {"number": value}}
    return {"cell_value": {"text": str(value)}}


def grid_rows_to_values(result: dict[str, Any], width: int) -> list[list[Any]]:
    grid = result.get("grid_data", {})
    rows = grid.get("rows", []) if isinstance(grid, dict) else []
    values: list[list[Any]] = []
    for row in rows:
        raw_values = row.get("values", []) if isinstance(row, dict) else []
        parsed = [cell_to_python(cell) for cell in raw_values[:width]]
        if len(parsed) < width:
            parsed.extend([None] * (width - len(parsed)))
        values.append(parsed)
    return values


def read_online_range(client: Any, file_id: str, sheet_id: str, range_text: str, width: int) -> list[list[Any]]:
    result = client.call_tool(
        "sheet.get_range",
        {"file_id": file_id, "sheet_id": sheet_id, "range": range_text},
    )
    return grid_rows_to_values(result, width)


def read_online_table(
    client: Any,
    url_or_id: str,
    tab: str | None,
    header_row: int,
    label: str,
) -> tuple[TableData, OnlineSheet]:
    short_id = extract_tencent_doc_id(url_or_id)
    info = client.query_file_info(short_id)
    file_id = str(find_nested_key(info, {"file_id", "fileID", "id"}) or short_id)

    sheet_info = client.sheet_get_info(file_id, concise=False)
    properties = sheet_info.get("properties", [])
    if not isinstance(properties, list):
        raise SyncError(f"{label} sheet.get_info 返回 properties 不是列表")

    online_sheet = select_online_sheet(properties, tab, label)
    online_sheet.file_id = file_id
    width = min(max(online_sheet.column_count, 1), MAX_ONLINE_COLUMNS)
    max_row = max(online_sheet.row_count, header_row)
    if max_row < header_row:
        raise SyncError(f"{label} 无法读取表头行: {header_row}")

    header_values = read_online_range(
        client,
        file_id,
        online_sheet.sheet_id,
        a1_range(header_row, 1, header_row, width),
        width,
    )
    if not header_values:
        raise SyncError(f"{label} 第 {header_row} 行没有可用表头")

    headers: dict[str, int] = {}
    duplicates: list[str] = []
    for offset, value in enumerate(header_values[0], 1):
        name = normalize_header(value)
        if not name:
            continue
        if name in headers:
            duplicates.append(name)
            continue
        headers[name] = offset
    if duplicates:
        raise SyncError(f"{label} 表头存在重复列: {', '.join(sorted(set(duplicates)))}")
    if not headers:
        raise SyncError(f"{label} 第 {header_row} 行没有可用表头")

    data_rows: list[RowRecord] = []
    if max_row > header_row:
        rows_per_read = max(1, min(MAX_ONLINE_ROWS, MAX_ONLINE_CELLS // width))
        current = header_row + 1
        while current <= max_row:
            end_row = min(max_row, current + rows_per_read - 1)
            values = read_online_range(
                client,
                file_id,
                online_sheet.sheet_id,
                a1_range(current, 1, end_row, width),
                width,
            )
            for offset, row_values in enumerate(values):
                row_number = current + offset
                row = {header: row_values[col_idx - 1] for header, col_idx in headers.items()}
                if not row_is_blank(row):
                    data_rows.append(RowRecord(row, row_number))
            current = end_row + 1

    table = TableData(label=label, sheet_name=online_sheet.title, header_row=header_row, headers=headers, rows=data_rows)
    return table, online_sheet


def add_missing_online_columns(
    table: TableData,
    sheet: OnlineSheet,
    needed_columns: list[str],
) -> list[UpdateCell]:
    header_updates: list[UpdateCell] = []
    next_col = max(table.headers.values(), default=0) + 1
    for column in needed_columns:
        if column in table.headers:
            continue
        table.headers[column] = next_col
        sheet.column_count = max(sheet.column_count, next_col)
        header_updates.append(
            UpdateCell(
                key_text="表头",
                source_row_number=table.header_row,
                target_row_number=table.header_row,
                target_column=column,
                target_column_number=next_col,
                old_value=None,
                new_value=column,
            )
        )
        next_col += 1
    return header_updates


def online_update_requests(sheet_id: str, updates: list[UpdateCell]) -> list[dict[str, Any]]:
    requests: list[dict[str, Any]] = []
    by_row: dict[int, list[UpdateCell]] = {}
    for update in updates:
        by_row.setdefault(update.target_row_number, []).append(update)

    for row_number in sorted(by_row):
        row_updates = sorted(by_row[row_number], key=lambda item: item.target_column_number)
        group: list[UpdateCell] = []
        for update in row_updates:
            if not group or update.target_column_number == group[-1].target_column_number + 1:
                group.append(update)
                continue
            requests.append(make_online_update_request(sheet_id, group))
            group = [update]
        if group:
            requests.append(make_online_update_request(sheet_id, group))
    return requests


def make_online_update_request(sheet_id: str, updates: list[UpdateCell]) -> dict[str, Any]:
    start_col = updates[0].target_column_number
    return {
        "update_range": {
            "sheet_id": sheet_id,
            "grid_data": {
                "start_row": updates[0].target_row_number - 1,
                "start_column": start_col - 1,
                "rows": [
                    {
                        "values": [python_to_cell(update.new_value) for update in updates],
                    }
                ],
            },
        }
    }


def online_append_requests(
    sheet_id: str,
    start_row: int,
    rows: list[dict[str, Any]],
    target_headers: dict[str, int],
) -> list[dict[str, Any]]:
    if not rows:
        return []
    header_order = sorted(target_headers.items(), key=lambda item: item[1])
    width = len(header_order)
    rows_per_request = max(1, min(MAX_ONLINE_ROWS, MAX_ONLINE_CELLS // max(width, 1)))
    requests: list[dict[str, Any]] = []
    for offset in range(0, len(rows), rows_per_request):
        chunk = rows[offset : offset + rows_per_request]
        requests.append(
            {
                "update_range": {
                    "sheet_id": sheet_id,
                    "grid_data": {
                        "start_row": start_row - 1 + offset,
                        "start_column": 0,
                        "rows": [
                            {
                                "values": [python_to_cell(row.get(header)) for header, _ in header_order],
                            }
                            for row in chunk
                        ],
                    },
                }
            }
        )
    return requests


def apply_online_requests(client: Any, file_id: str, requests: list[dict[str, Any]]) -> None:
    for offset in range(0, len(requests), MAX_BATCH_REQUESTS):
        client.sheet_batch_update(file_id, requests[offset : offset + MAX_BATCH_REQUESTS])


def verify_online_cells(client: Any, file_id: str, sheet_id: str, updates: list[UpdateCell]) -> None:
    for update in updates[:20]:
        range_text = a1_range(
            update.target_row_number,
            update.target_column_number,
            update.target_row_number,
            update.target_column_number,
        )
        values = read_online_range(client, file_id, sheet_id, range_text, 1)
        actual = values[0][0] if values and values[0] else None
        if not values_equal(actual, update.new_value):
            raise SyncError(
                f"线上写入后验证失败: {range_text} 预期 {update.new_value!r}，实际 {actual!r}"
            )


def write_report(
    report_dir: Path,
    *,
    mode: str,
    backend: str,
    source_label: str,
    target_label: str,
    output: Path | None,
    backup: Path | None,
    source_sheet: str,
    target_sheet: str,
    mapping: dict[str, str],
    literals: dict[str, str],
    key_columns: list[str],
    source_count: int,
    target_count: int,
    append_count: int,
    update_count: int,
    unchanged_count: int,
    skipped: list[tuple[str, str]],
    updates: list[UpdateCell],
) -> Path:
    report_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = report_dir / f"{timestamp}_lx-biaogetongbu_处理日志.md"

    lines = [
        "# lx-biaogetongbu 表格同步处理日志",
        "",
        f"**处理时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**后端**: {backend}",
        f"**模式**: {mode}",
        f"**A 表**: {source_label}",
        f"**B 表**: {target_label}",
        f"**A sheet**: {source_sheet}",
        f"**B sheet**: {target_sheet}",
    ]
    if output:
        lines.append(f"**输出文件**: {output}")
    if backup:
        lines.append(f"**B 表备份**: {backup}")

    lines.extend(
        [
            "",
            "## 统计",
            "",
            "| 指标 | 数值 |",
            "|---|---:|",
            f"| A 表有效行 | {source_count} |",
            f"| B 表已有行 | {target_count} |",
            f"| 本次追加行 | {append_count} |",
            f"| 本次更新单元格 | {update_count} |",
            f"| 值未变化单元格 | {unchanged_count} |",
            f"| 跳过/冲突 | {len(skipped)} |",
            "",
            "## 字段映射",
            "",
            "| A 表列 | B 表列 |",
            "|---|---|",
        ]
    )
    for source_col, target_col in mapping.items():
        lines.append(f"| {source_col} | {target_col} |")

    if literals:
        lines.extend(["", "## 固定写入", "", "| B 表列 | 固定值 |", "|---|---|"])
        for target_col, value in literals.items():
            lines.append(f"| {target_col} | {value} |")

    lines.extend(["", "## 定位键", "", ", ".join(key_columns) if key_columns else "未设置，append 模式全部追加"])

    if updates:
        lines.extend(["", "## 更新预览", "", "| 定位键 | A行 | B行 | B列 | 原值 | 新值 |", "|---|---:|---:|---|---|---|"])
        for update in updates[:200]:
            lines.append(
                f"| {update.key_text} | {update.source_row_number} | {update.target_row_number} | "
                f"{update.target_column} | {update.old_value if update.old_value is not None else ''} | "
                f"{update.new_value if update.new_value is not None else ''} |"
            )
        if len(updates) > 200:
            lines.append(f"| ... |  |  |  |  | 还有 {len(updates) - 200} 个单元格未展示 |")

    if skipped:
        lines.extend(["", "## 跳过记录", "", "| 标识 | 原因 |", "|---|---|"])
        for identifier, reason in skipped[:200]:
            lines.append(f"| {identifier} | {reason} |")
        if len(skipped) > 200:
            lines.append(f"| ... | 还有 {len(skipped) - 200} 条未展示 |")

    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="从 A 表同步记录到 B 表。")
    parser.add_argument("--profile", help="JSON profile 路径，可预置 sheet、mapping、key、literal")
    parser.add_argument("--mode", choices=["append", "update-by-key"], help="同步模式，默认 append")
    parser.add_argument("--source", help="A 表本地 Excel 路径")
    parser.add_argument("--target", help="B 表本地 Excel 路径")
    parser.add_argument("--output", help="另存输出路径；不传则 confirmed 时写回 B 表")
    parser.add_argument("--source-sheet", help="A 表 sheet 名")
    parser.add_argument("--target-sheet", help="B 表 sheet 名")
    parser.add_argument("--source-header-row", type=int, help="A 表表头行，默认 1")
    parser.add_argument("--target-header-row", type=int, help="B 表表头行，默认 1")
    parser.add_argument("--map", action="append", default=[], help="字段映射，格式 源列=目标列，可重复")
    parser.add_argument("--key", action="append", default=[], help="定位/去重键，可重复，也可逗号分隔")
    parser.add_argument("--update-column", action="append", default=[], help="update-by-key 允许更新的 B 表列，可重复或逗号分隔")
    parser.add_argument("--literal", action="append", default=[], help="固定写入，格式 目标列=固定值，可重复")
    parser.add_argument("--add-missing-target-columns", action="store_true", help="B 表缺少目标列时自动追加表头")
    parser.add_argument("--allow-blank-updates", action="store_true", help="允许 update-by-key 用空值覆盖 B 表")
    parser.add_argument("--online", action="store_true", help="使用腾讯文档在线表格后端")
    parser.add_argument(
        "--online-backend",
        choices=["auto", "mcp", "saas-api"],
        help="线上后端：auto 先走 MCP，限流/不可用时在写入前切到 lx-txsaasdocs API；默认 auto",
    )
    parser.add_argument("--source-url", help="A 表腾讯文档 URL 或 file_id")
    parser.add_argument("--target-url", help="B 表腾讯文档 URL 或 file_id")
    parser.add_argument("--source-tab", help="A 表在线 sheet 标题或 sheet_id")
    parser.add_argument("--target-tab", help="B 表在线 sheet 标题或 sheet_id")
    parser.add_argument("--mcp-config", help="腾讯文档 SaaS MCP 配置路径，默认 ~/.workbuddy/mcp.json")
    parser.add_argument("--mcp-server-name", help="MCP server 名称，默认 tencent-docs")
    parser.add_argument("--saas-config-path", help="lx-txsaasdocs API 配置路径；默认读取 config/fog_config.yaml")
    parser.add_argument("--timeout", type=int, default=60, help="线上 API 请求超时秒数")
    parser.add_argument("--min-interval", type=float, default=0.0, help="线上 API 调用最小间隔秒数")
    parser.add_argument("--retries", type=int, default=0, help="线上 API 重试次数")
    parser.add_argument("--rate-limit-sleep", type=int, default=300, help="线上 API 限流重试等待秒数")
    parser.add_argument("--skip-online-verify", action="store_true", help="线上 confirmed 写入后跳过读回验证")
    parser.add_argument("--report-dir", default="workspace/10表格同步/处理日志", help="处理日志输出目录")
    parser.add_argument("--dry-run", action="store_true", help="只预览，不写入")
    parser.add_argument("--confirmed", action="store_true", help="确认写入")
    return parser


def prepare_sync(args: argparse.Namespace, profile: dict[str, Any], source_table: TableData, target_table: TableData) -> tuple[
    str,
    dict[str, str],
    dict[str, str],
    list[str],
    list[str],
    bool,
]:
    mode = str(first_value(args.mode, profile, "mode", "append"))
    if mode not in {"append", "update-by-key"}:
        raise SyncError(f"不支持的同步模式: {mode}")

    explicit_mapping = profile_mapping(profile.get("mapping"))
    explicit_mapping.update(parse_pairs(args.map, "--map"))
    literals = profile_mapping(profile.get("literals"))
    literals.update(parse_pairs(args.literal, "--literal"))
    key_columns_raw = split_values(first_value(args.key, profile, "keys", []))
    add_missing = bool(args.add_missing_target_columns or profile.get("add_missing_target_columns"))

    if not add_missing:
        missing_literal_columns = [name for name in literals if name not in target_table.headers]
        if missing_literal_columns:
            raise SyncError(f"B 表缺少固定写入目标列: {', '.join(missing_literal_columns)}")

    mapping = build_mapping(source_table.headers, target_table.headers, explicit_mapping)
    mapping = ensure_key_mapping(mapping, key_columns_raw, source_table.headers, target_table.headers)
    key_columns = resolve_key_columns(key_columns_raw, mapping, target_table.headers)
    allow_blank_updates = bool(args.allow_blank_updates or profile.get("allow_blank_updates"))
    update_columns = (
        resolve_update_columns(args.update_column, profile, key_columns, target_table.headers)
        if mode == "update-by-key"
        else []
    )
    if mode == "update-by-key" and not key_columns:
        raise SyncError("update-by-key 模式必须配置 --key")

    return mode, mapping, literals, key_columns, update_columns, allow_blank_updates


def run_excel(args: argparse.Namespace, profile: dict[str, Any]) -> int:
    source = resolve_path(first_value(args.source, profile, "source"), required=True)
    target = resolve_path(first_value(args.target, profile, "target"), required=True)
    assert source is not None and target is not None
    output = resolve_path(first_value(args.output, profile, "output")) if first_value(args.output, profile, "output") else None
    report_dir = resolve_path(args.report_dir, required=True)
    assert report_dir is not None

    source_sheet_name = first_value(args.source_sheet, profile, "source_sheet")
    target_sheet_name = first_value(args.target_sheet, profile, "target_sheet")
    source_header_row = int(first_value(args.source_header_row, profile, "source_header_row", 1))
    target_header_row = int(first_value(args.target_header_row, profile, "target_header_row", 1))
    add_missing = bool(args.add_missing_target_columns or profile.get("add_missing_target_columns"))

    require_excel(source, "A 表")
    require_excel(target, "B 表")
    if output and output.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise SyncError(f"输出文件只支持 .xlsx / .xlsm: {output}")

    source_wb = load_workbook(source, data_only=True)
    target_wb = load_workbook(target, keep_vba=target.suffix.lower() == ".xlsm")
    source_ws = select_sheet(source_wb, source_sheet_name, "A 表")
    target_ws = select_sheet(target_wb, target_sheet_name, "B 表")

    source_headers = read_headers(source_ws, source_header_row)
    target_headers = read_headers(target_ws, target_header_row)
    explicit_mapping = profile_mapping(profile.get("mapping"))
    explicit_mapping.update(parse_pairs(args.map, "--map"))
    literals = profile_mapping(profile.get("literals"))
    literals.update(parse_pairs(args.literal, "--literal"))
    update_column_candidates = split_values(args.update_column or profile.get("update_columns", []))
    if add_missing:
        target_headers = add_missing_target_columns(
            target_ws,
            target_headers,
            list(explicit_mapping.values()) + list(literals) + update_column_candidates,
            target_header_row,
        )

    source_table = TableData("A 表", source_ws.title, source_header_row, source_headers, read_row_records(source_ws, source_headers, source_header_row))
    target_table = TableData("B 表", target_ws.title, target_header_row, target_headers, read_row_records(target_ws, target_headers, target_header_row))
    mode, mapping, literals, key_columns, update_columns, allow_blank_updates = prepare_sync(args, profile, source_table, target_table)

    mapped_rows = make_target_records(source_table.rows, mapping, literals)
    append_rows: list[dict[str, Any]] = []
    updates: list[UpdateCell] = []
    skipped: list[tuple[str, str]] = []
    unchanged = 0
    if mode == "append":
        append_rows, skipped = filter_append_rows(mapped_rows, key_columns, existing_keys(target_table.rows, key_columns))
    else:
        updates, skipped, unchanged = build_update_cells(
            mapped_rows,
            target_table.rows,
            key_columns,
            update_columns,
            target_table.headers,
            allow_blank_updates=allow_blank_updates,
        )

    run_mode = "dry-run" if args.dry_run else "confirmed"
    backup: Path | None = None
    if args.confirmed and (append_rows or updates):
        save_path = output or target
        save_path.parent.mkdir(parents=True, exist_ok=True)
        if output is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup = target.with_name(f"{target.name}.bak.{timestamp}")
            shutil.copy2(target, backup)
        if append_rows:
            append_to_sheet(target_ws, append_rows, target_table.headers)
        if updates:
            apply_updates_to_sheet(target_ws, updates)
        target_wb.save(save_path)

    report_path = write_report(
        report_dir,
        mode=f"{run_mode}/{mode}",
        backend="excel",
        source_label=str(source),
        target_label=str(target),
        output=output,
        backup=backup,
        source_sheet=source_table.sheet_name,
        target_sheet=target_table.sheet_name,
        mapping=mapping,
        literals=literals,
        key_columns=key_columns,
        source_count=len(source_table.rows),
        target_count=len(target_table.rows),
        append_count=len(append_rows),
        update_count=len(updates),
        unchanged_count=unchanged,
        skipped=skipped,
        updates=updates,
    )

    print_summary(run_mode, mode, "excel", len(source_table.rows), len(target_table.rows), len(append_rows), len(updates), unchanged, skipped)
    if backup:
        print(f"B 表备份: {backup}")
    if output:
        print(f"输出文件: {output}")
    print(f"处理日志: {report_path}")
    return 0


def run_online(args: argparse.Namespace, profile: dict[str, Any]) -> int:
    source_url = first_value(args.source_url, profile, "source_url")
    target_url = first_value(args.target_url, profile, "target_url")
    if not source_url or not target_url:
        raise SyncError("online 模式必须提供 --source-url 和 --target-url")
    report_dir = resolve_path(args.report_dir, required=True)
    assert report_dir is not None

    source_header_row = int(first_value(args.source_header_row, profile, "source_header_row", 1))
    target_header_row = int(first_value(args.target_header_row, profile, "target_header_row", 1))
    source_tab = first_value(args.source_tab, profile, "source_tab")
    target_tab = first_value(args.target_tab, profile, "target_tab")
    add_missing = bool(args.add_missing_target_columns or profile.get("add_missing_target_columns"))
    args.online_backend = str(first_value(args.online_backend, profile, "online_backend", "auto"))

    client = load_saas_client(args)
    source_table, source_sheet = read_online_table(client, str(source_url), source_tab, source_header_row, "A 表")
    target_table, target_sheet = read_online_table(client, str(target_url), target_tab, target_header_row, "B 表")

    explicit_mapping = profile_mapping(profile.get("mapping"))
    explicit_mapping.update(parse_pairs(args.map, "--map"))
    literals = profile_mapping(profile.get("literals"))
    literals.update(parse_pairs(args.literal, "--literal"))
    header_updates: list[UpdateCell] = []
    update_column_candidates = split_values(args.update_column or profile.get("update_columns", []))
    if add_missing:
        header_updates = add_missing_online_columns(
            target_table,
            target_sheet,
            list(explicit_mapping.values()) + list(literals) + update_column_candidates,
        )

    mode, mapping, literals, key_columns, update_columns, allow_blank_updates = prepare_sync(args, profile, source_table, target_table)
    mapped_rows = make_target_records(source_table.rows, mapping, literals)

    append_rows: list[dict[str, Any]] = []
    updates: list[UpdateCell] = []
    skipped: list[tuple[str, str]] = []
    unchanged = 0
    if mode == "append":
        append_rows, skipped = filter_append_rows(mapped_rows, key_columns, existing_keys(target_table.rows, key_columns))
    else:
        updates, skipped, unchanged = build_update_cells(
            mapped_rows,
            target_table.rows,
            key_columns,
            update_columns,
            target_table.headers,
            allow_blank_updates=allow_blank_updates,
        )

    run_mode = "dry-run" if args.dry_run else "confirmed"
    if args.confirmed:
        requests = online_update_requests(target_sheet.sheet_id, header_updates)
        if append_rows:
            start_row = max(target_sheet.row_count, target_header_row) + 1
            requests.extend(online_append_requests(target_sheet.sheet_id, start_row, append_rows, target_table.headers))
        if updates:
            requests.extend(online_update_requests(target_sheet.sheet_id, updates))
        if requests:
            apply_online_requests(client, target_sheet.file_id, requests)
            if not args.skip_online_verify:
                verify_online_cells(client, target_sheet.file_id, target_sheet.sheet_id, header_updates + updates)

    backend_label = str(getattr(client, "backend_label", "tencent-docs"))
    report_path = write_report(
        report_dir,
        mode=f"{run_mode}/{mode}",
        backend=backend_label,
        source_label=str(source_url),
        target_label=str(target_url),
        output=None,
        backup=None,
        source_sheet=source_table.sheet_name,
        target_sheet=target_table.sheet_name,
        mapping=mapping,
        literals=literals,
        key_columns=key_columns,
        source_count=len(source_table.rows),
        target_count=len(target_table.rows),
        append_count=len(append_rows),
        update_count=len(updates),
        unchanged_count=unchanged,
        skipped=skipped,
        updates=header_updates + updates,
    )

    print_summary(run_mode, mode, backend_label, len(source_table.rows), len(target_table.rows), len(append_rows), len(updates), unchanged, skipped)
    if header_updates:
        print(f"将新增/已新增 B 表表头列: {len(header_updates)}")
    print(f"处理日志: {report_path}")
    return 0


def print_summary(
    run_mode: str,
    mode: str,
    backend: str,
    source_count: int,
    target_count: int,
    append_count: int,
    update_count: int,
    unchanged_count: int,
    skipped: list[tuple[str, str]],
) -> None:
    print(f"后端: {backend}")
    print(f"模式: {run_mode}/{mode}")
    print(f"A 表有效行: {source_count}")
    print(f"B 表已有行: {target_count}")
    print(f"本次追加行: {append_count}")
    print(f"本次更新单元格: {update_count}")
    print(f"值未变化单元格: {unchanged_count}")
    print(f"跳过/冲突: {len(skipped)}")


def run(args: argparse.Namespace) -> int:
    profile_path = resolve_path(args.profile) if args.profile else None
    profile = load_profile(profile_path)

    if args.dry_run == args.confirmed:
        raise SyncError("必须且只能选择一个模式：--dry-run 或 --confirmed")

    online = bool(args.online or first_value(args.source_url, profile, "source_url") or first_value(args.target_url, profile, "target_url"))
    if online:
        return run_online(args, profile)
    return run_excel(args, profile)


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return run(args)
    except (SyncError, OnlineBackendError) as exc:
        print(f"[FAIL] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
