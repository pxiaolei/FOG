"""
免佣卡策略生成脚本

功能：根据共补时段数据，生成免佣卡策略
- 支持从配置文件读取城市卡券配置
- 从Excel码表读取价格
- 改进的公式复制（精确匹配单元格引用）
- 支持批量执行和 --all-brands 参数

使用示例：
    # 单城市单品牌
    python3 create_mianyongka.py --file "城市策略活动表.xlsm" --city "福州市" --brand "哈啰轻快" --start 2026-04-20 --end 2026-04-23

    # 单城市所有品牌
    python3 create_mianyongka.py --file "城市策略活动表.xlsm" --city "福州市" --all-brands --start 2026-04-20 --end 2026-04-23

    # 批量执行所有城市-品牌组合
    python3 create_mianyongka.py --batch --file "城市策略活动表.xlsm" --start 2026-04-20 --end 2026-04-23
"""

import re
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.formula import ArrayFormula
from datetime import datetime, timedelta
import yaml
import argparse

# 添加 lxx_share 到路径
_skills_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, _skills_dir)

from config_loader import load_celuehuodong_config
from lxx_share import DatabaseConnector


# 配置文件路径（默认在skill根目录）
DEFAULT_CONFIG_PATH = os.path.join(os.path.dirname(__file__), '..', 'config.yaml')


def get_all_city_brands(config):
    """获取配置中所有城市-品牌组合"""
    results = []
    cities_config = config.get('cities', {})
    for city_name, city_config in cities_config.items():
        brands = city_config.get('brands', {})
        for brand_name in brands.keys():
            results.append((city_name, brand_name))
    return results


def get_city_brands(config, city_name):
    """获取指定城市的所有品牌"""
    city_config = config.get('cities', {}).get(city_name, {})
    return list(city_config.get('brands', {}).keys())


def load_config(config_path=None):
    """加载配置文件"""
    if config_path is None:
        return load_celuehuodong_config()
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def get_brand_card_config(config, city_name, brand_name):
    """获取城市-品牌卡券配置

    配置层级：城市 → 品牌 → 卡券
    如果未找到具体配置，使用默认配置
    """
    cities = config.get('cities', {})

    if city_name in cities:
        city_config = cities[city_name]
        brands = city_config.get('brands', {})
        if brand_name in brands:
            return brands[brand_name].get('cards', {})
        # 城市有配置但品牌无配置，使用默认
        return config.get('default', {}).get('cards', {})

    # 城市无配置，使用默认
    return config.get('default', {}).get('cards', {})


def parse_non_b_periods_to_hours(value):
    if not value:
        return None

    text = str(value).replace('，', ',').replace('；', '、').replace(';', '、')
    has_strategy_word = '免佣' in text or '流水' in text
    periods = []

    for segment in re.split(r'[、\n]+', text):
        segment = segment.strip()
        if not segment:
            continue
        if '免佣' in segment:
            segment = segment.split('免佣', 1)[0]
        elif has_strategy_word:
            continue

        for start, end in re.findall(r'(\d{1,2}):\d{2}(?::\d{2})?-(\d{1,2}):\d{2}(?::\d{2})?', segment):
            start_h = int(start)
            end_h = int(end)
            if end_h == 0:
                end_h = 24
            if 0 <= start_h < end_h <= 24:
                periods.append((start_h, end_h))

    return periods or None


def get_style_config(config):
    """获取样式配置"""
    styles = config.get('styles', {})
    font_cfg = styles.get('font', {})
    align_cfg = styles.get('alignment', {})

    return {
        'font': Font(
            name=font_cfg.get('name', '思源黑体'),
            size=font_cfg.get('size', 10)
        ),
        'alignment': Alignment(
            horizontal=align_cfg.get('horizontal', 'center'),
            vertical=align_cfg.get('vertical', 'center'),
            wrap_text=align_cfg.get('wrap_text', False)
        )
    }


def get_gongbu_periods(conn, city_name, date):
    """从数据库获取指定城市日期的共补时段

    当天没有数据时，默认无共补活动。
    返回: (gongbu_periods, is_isolated, quick_gongbu_periods)
    """
    cur = conn.cursor()
    cur.execute('''
        SELECT strategy_type,
               start_time_1, end_time_1,
               start_time_2, end_time_2,
               start_time_3, end_time_3,
               start_time_4, end_time_4,
               start_time_5, end_time_5,
               is_isolated_city,
               non_b_isolated_time_range
        FROM hhdata.fact_gongbu_strategy
        WHERE city_name = %s AND date = %s AND strategy_type = '免佣'
    ''', (city_name, date))

    rows = cur.fetchall()
    cur.close()

    if not rows:
        return [], False, None  # 无共补数据

    gongbu_periods = []
    is_isolated = False
    quick_gongbu_periods = None

    for row in rows:
        is_isolated = row[11] if row[11] is not None else False
        non_b_periods = parse_non_b_periods_to_hours(row[12])
        if non_b_periods:
            quick_gongbu_periods = non_b_periods

        for i in range(1, 6):
            start = row[i*2 - 1]
            end = row[i*2]

            if start and end:
                start_hour = start.hour
                # end_time +1秒转小时数：23:00:00→23, 23:59:59→24
                # 00:00:00 表示午夜24点
                if end.hour == 0 and end.minute == 0 and end.second == 0:
                    end_hour = 24
                else:
                    end_total_seconds = (end.hour * 3600 + end.minute * 60 + end.second) + 1
                    end_hour = end_total_seconds // 3600
                    if end_hour > 24:
                        end_hour = 24
                gongbu_periods.append((start_hour, end_hour))

    return gongbu_periods, is_isolated, quick_gongbu_periods


def subtract_periods(base_start, base_end, subtract_list):
    """从基础时段扣除指定时段"""
    covered_hours = set(range(base_start, base_end))

    for s_start, s_end in subtract_list:
        for h in range(s_start, min(s_end, base_end)):
            if h >= base_start:
                covered_hours.discard(h)

    if not covered_hours:
        return []

    sorted_hours = sorted(covered_hours)
    periods = []
    current_start = sorted_hours[0]
    prev_hour = sorted_hours[0]

    for h in sorted_hours[1:]:
        if h != prev_hour + 1:
            periods.append((current_start, prev_hour + 1))
            current_start = h
        prev_hour = h

    periods.append((current_start, prev_hour + 1))
    return periods


def calculate_card_periods(card_config, gongbu_periods, is_isolated):
    """计算卡券时段"""
    base_hours = card_config.get('base_hours', [0, 24])
    base_start, base_end = base_hours[0], base_hours[1]

    subtract_gongbu = card_config.get('subtract_gongbu', True)
    k_only = card_config.get('k_only', False)

    if not subtract_gongbu:
        return [(base_start, base_end)]

    if k_only:
        # 快车单卡
        # B补隔离城市(is_isolated=True)：共补不含快车单 → 返回全天(0-24)
        # 非B补隔离城市(is_isolated=False)：共补含快车单 → 扣除共补时段
        if is_isolated:
            return [(base_start, base_end)]
        else:
            return subtract_periods(base_start, base_end, gongbu_periods)

    return subtract_periods(base_start, base_end, gongbu_periods)


def round_to_point_nine(price):
    """将价格调整到距离最近的.9结尾数字

    规则：取距离最近的.9值，等距时取高的值
    例如：
    - 15.3：距离14.9是0.4，距离15.9是0.6 → 取14.9
    - 15.5：距离14.9是0.6，距离15.9是0.4 → 取15.9
    - 15.4：距离14.9是0.5，距离15.9是0.5 → 取15.9（等距取高）
    """
    # 边界处理：最低价格0.9
    if price < 0.9:
        return 0.9

    int_part = int(price)
    dec_part = price - int_part

    # 计算相邻的两个.9值
    if dec_part < 0.9:
        # price在整数部分和整数部分+0.9之间
        # 比如15.3在15和15.9之间，比较14.9和15.9
        low_nine = int_part - 0.1  # 14.9
        high_nine = int_part + 0.9  # 15.9
        # 但如果int_part=0，low_nine会是-0.1，需要处理
        if int_part == 0:
            low_nine = 0.0  # 0作为下限
    else:
        # price在整数部分+0.9和整数部分+1之间
        # 比如15.95，比较15.9和16.9
        low_nine = int_part + 0.9
        high_nine = int_part + 1.9

    # 比较距离
    dist_low = price - low_nine
    dist_high = high_nine - price

    if dist_low < dist_high:
        return low_nine
    elif dist_low > dist_high:
        return high_nine
    else:
        return high_nine  # 等距取高


def get_price_from_excel(wb, sheet_name, card_code, city_name, brand, hours):
    """从Excel码表读取单价，计算价格并调整到.9结尾

    读取逻辑：
    1. 从码表sheet读取单价（匹配品牌+城市+卡券类型）
    2. 计算基础价格 = 小时数 × 单价
    3. 调整到距离最近的.9结尾数字
    """
    try:
        if sheet_name not in wb.sheetnames:
            # 码表不存在，用默认单价计算
            default_unit_price = 1.5
            base_price = hours * default_unit_price
            return round_to_point_nine(base_price)

        sheet = wb[sheet_name]

        # 码表结构：A列品牌，B列城市，D列卡券类型，F列单价
        unit_price = None
        for row in range(2, sheet.max_row + 1):
            row_brand = sheet.cell(row=row, column=1).value
            row_city = sheet.cell(row=row, column=2).value
            row_card_type = sheet.cell(row=row, column=4).value
            row_price = sheet.cell(row=row, column=6).value

            if row_brand == brand and row_city == city_name and row_card_type == card_code:
                unit_price = row_price
                break

        if unit_price is None:
            # 未找到匹配，使用默认单价
            unit_price = 1.5

        # 计算基础价格 = 小时数 × 单价
        base_price = hours * unit_price
        return round_to_point_nine(base_price)

    except Exception:
        # 备用：基于默认单价计算
        default_unit_price = 1.5
        base_price = hours * default_unit_price
        return round_to_point_nine(base_price)


def replace_cell_references(formula_text, old_row, new_row):
    """替换公式中的单元格引用行号

    只匹配Excel单元格引用格式：字母+行号（如A2723、B2723）
    不替换纯数字（避免误改数值）
    """
    # 匹配单元格引用：字母(1-3个) + 行号
    # 如 A2723、AB2723、ABC2723
    pattern = r'([A-Z]{1,3})' + str(old_row)

    def replace_func(match):
        col_letter = match.group(1)
        return col_letter + str(new_row)

    return re.sub(pattern, replace_func, formula_text)


def copy_cell_with_formula(source_cell, target_cell, old_row, new_row):
    """复制单元格，包括公式、格式和ArrayFormula

    改进：使用精确的单元格引用匹配，避免误替换纯数字
    """
    if source_cell.value is None:
        return

    # 复制number_format
    target_cell.number_format = source_cell.number_format

    if isinstance(source_cell.value, ArrayFormula):
        # ArrayFormula：使用text属性获取公式，精确替换行号
        formula_text = source_cell.value.text if hasattr(source_cell.value, 'text') else ''
        if formula_text:
            new_formula = replace_cell_references(formula_text, old_row, new_row)
            target_cell.value = ArrayFormula(ref=f"{target_cell.coordinate}", text=new_formula)
    elif str(source_cell.value).startswith('='):
        # 普通公式：精确替换行号
        new_formula = replace_cell_references(str(source_cell.value), old_row, new_row)
        target_cell.value = new_formula
    else:
        # 非公式，直接复制
        target_cell.value = source_cell.value


def create_mianyongka_records(file_path, city_name, brand, start_date, end_date,
                              output_path=None, config=None):
    """更新免佣卡sheet"""
    # 加载配置
    if config is None:
        config = load_config()

    # 获取城市-品牌卡券配置
    brand_cards = get_brand_card_config(config, city_name, brand)
    style_config = get_style_config(config)

    # 使用 DatabaseConnector 获取数据库连接
    db = DatabaseConnector()

    # 加载xlsm文件
    wb = load_workbook(file_path, keep_vba=True)
    sheet = wb['免佣卡']

    # 确定模板行
    if config.get('template_row'):
        template_row = config['template_row']
    elif config.get('use_last_row_as_template', True):
        template_row = sheet.max_row
    else:
        template_row = sheet.max_row

    last_row = template_row

    # 价格sheet名称
    price_sheet_name = config.get('prices', {}).get('sheet_name', '码表-各类型卡单价')

    card_records = []

    with db.connect() as conn:
        current_date = start_date
        while current_date <= end_date:
            gongbu_periods, is_isolated, quick_gongbu_periods = get_gongbu_periods(conn, city_name, current_date)

            for card_code, card_config in brand_cards.items():
                source_periods = quick_gongbu_periods if card_config.get('k_only') and quick_gongbu_periods else gongbu_periods
                periods = calculate_card_periods(card_config, source_periods, is_isolated)

                if not periods:
                    continue

                new_row = last_row + 1
                last_row += 1

                # 计算小时数
                total_hours = sum(e - s for s, e in periods)

                # 卡券名称
                card_name = f"{current_date.day}号{card_config.get('name', '免佣卡')}{card_code}"

                # 写入数据列
                # A列: 城市名称
                cell_a = sheet.cell(row=new_row, column=1)
                cell_a.value = city_name
                cell_a.font = style_config['font']
                cell_a.alignment = style_config['alignment']

                # B列: 品牌
                cell_b = sheet.cell(row=new_row, column=2)
                cell_b.value = brand
                cell_b.font = style_config['font']
                cell_b.alignment = style_config['alignment']

                # E列: 免佣卡名称
                cell_e = sheet.cell(row=new_row, column=5)
                cell_e.value = card_name
                cell_e.font = style_config['font']
                cell_e.alignment = style_config['alignment']

                # F列: 日期（格式 yyyy/m/d;@）
                cell_f = sheet.cell(row=new_row, column=6)
                cell_f.value = current_date
                cell_f.font = style_config['font']
                cell_f.alignment = style_config['alignment']
                cell_f.number_format = 'yyyy/m/d;@'

                # G列: 日期类型（周判断）
                weekday = current_date.weekday()  # 0=周一, 4=周五, 5=周六, 6=周日
                if weekday <= 3:  # 周一~周四
                    date_type = '周一周四'
                else:  # 周五~周日
                    date_type = '周五周末'
                cell_g = sheet.cell(row=new_row, column=7)
                cell_g.value = date_type
                cell_g.font = style_config['font']
                cell_g.alignment = style_config['alignment']

                # H列: 卡券天数
                cell_h = sheet.cell(row=new_row, column=8)
                cell_h.value = 1
                cell_h.font = style_config['font']
                cell_h.alignment = style_config['alignment']

                # I-R列: 时段1-5开始/结束
                for i in range(1, 6):
                    start_col = 9 + (i-1) * 2
                    end_col = 10 + (i-1) * 2

                    cell_start = sheet.cell(row=new_row, column=start_col)
                    cell_end = sheet.cell(row=new_row, column=end_col)

                    if i <= len(periods):
                        s, e = periods[i-1]
                        cell_start.value = s
                        cell_end.value = e - 1
                    else:
                        cell_start.value = None
                        cell_end.value = None

                    cell_start.font = style_config['font']
                    cell_start.alignment = style_config['alignment']
                    cell_end.font = style_config['font']
                    cell_end.alignment = style_config['alignment']

                # V列: 修正价格（从Excel码表读取，调整到.9）
                price = get_price_from_excel(wb, price_sheet_name, card_code, city_name, brand, total_hours)
                cell_v = sheet.cell(row=new_row, column=22)
                cell_v.value = price
                cell_v.font = style_config['font']
                cell_v.alignment = style_config['alignment']

                # X列: 补贴类型
                subsidy_types = "快车单" if card_config.get('k_only', False) else "稳收单,普通单,特赚单,快车单"
                cell_x = sheet.cell(row=new_row, column=24)
                cell_x.value = subsidy_types
                cell_x.font = style_config['font']
                cell_x.alignment = style_config['alignment']

                # 复制公式列
                formula_cols = [3, 4, 19, 20, 21, 23, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50]

                for col in formula_cols:
                    source_cell = sheet.cell(row=template_row, column=col)
                    target_cell = sheet.cell(row=new_row, column=col)
                    copy_cell_with_formula(source_cell, target_cell, template_row, new_row)
                    target_cell.font = style_config['font']

                    # C列(列3)特殊处理：需要换行
                    if col == 3:
                        target_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        target_cell.alignment = style_config['alignment']

                # AY列(列51): 活动要求 - 从模板行复制
                source_ay = sheet.cell(row=template_row, column=51)
                ay_cell = sheet.cell(row=new_row, column=51)
                if source_ay.value:
                    ay_cell.value = source_ay.value
                ay_cell.font = style_config['font']
                ay_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

                card_records.append({
                    '城市': city_name,
                    '品牌': brand,
                    '日期': current_date.strftime('%Y-%m-%d'),
                    '卡类型': card_code,
                    '卡券名称': card_name,
                    '时段': periods,
                    '小时数': total_hours,
                    '价格': price,
                    'B补隔离': is_isolated
                })

            current_date += timedelta(days=1)

    # 保存文件
    save_path = output_path or file_path

    # 调整C列列宽（自动换行需要足够宽度）
    from openpyxl.utils import get_column_letter
    sheet.column_dimensions[get_column_letter(3)].width = 25

    wb.save(save_path)
    wb.close()

    return card_records


def main():
    parser = argparse.ArgumentParser(description='生成免佣卡策略')
    parser.add_argument('--file', required=True, help='xlsm文件路径')
    parser.add_argument('--city', help='城市名称（单城市模式）')
    parser.add_argument('--brand', help='品牌名称（单品牌模式）')
    parser.add_argument('--all-brands', action='store_true', help='处理指定城市的所有品牌')
    parser.add_argument('--batch', action='store_true', help='批量模式：处理配置中所有城市-品牌组合')
    parser.add_argument('--start', required=True, help='开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end', required=True, help='结束日期 (YYYY-MM-DD)')
    parser.add_argument('--output', help='输出文件路径（可选）')
    parser.add_argument('--config', help='配置文件路径（可选，默认使用skill目录下的config.yaml）')

    args = parser.parse_args()

    # 加载配置
    config = load_config(args.config) if args.config else load_config()

    start_date = datetime.strptime(args.start, '%Y-%m-%d')
    end_date = datetime.strptime(args.end, '%Y-%m-%d')

    # 确定要处理的城市-品牌组合
    city_brands = []

    if args.batch:
        # 批量模式：处理所有城市-品牌组合
        city_brands = get_all_city_brands(config)
    elif args.city:
        if args.brand:
            # 单城市单品牌
            city_brands = [(args.city, args.brand)]
        elif args.all_brands:
            # 单城市所有品牌
            brands = get_city_brands(config, args.city)
            for brand in brands:
                city_brands.append((args.city, brand))
        else:
            # 默认：处理该城市所有品牌
            brands = get_city_brands(config, args.city)
            for brand in brands:
                city_brands.append((args.city, brand))
    else:
        # 未指定城市，使用批量模式
        city_brands = get_all_city_brands(config)
        print(f"未指定城市，批量处理所有配置的城市-品牌组合")

    if not city_brands:
        print("未找到城市-品牌配置")
        return

    print(f"处理 {len(city_brands)} 个城市-品牌组合")

    total_cards = 0
    all_records = []

    for city_name, brand_name in city_brands:
        print(f"\n处理：{city_name} - {brand_name}")

        records = create_mianyongka_records(
            file_path=args.file,
            city_name=city_name,
            brand=brand_name,
            start_date=start_date,
            end_date=end_date,
            output_path=args.output,
            config=config
        )

        total_cards += len(records)
        all_records.extend(records)

        for r in records:
            period_str = ', '.join([f"{s}-{e}点" for s, e in r['时段']])
            print(f"  {r['日期']} {r['卡券名称']}: {period_str} ({r['小时数']}小时, ¥{r['价格']})")

    print(f"\n===== 汇总 =====")
    print(f"共生成 {total_cards} 张免佣卡")
    for city_name, brand_name in city_brands:
        count = sum(1 for r in all_records if r['城市'] == city_name and r['品牌'] == brand_name)
        print(f"  {city_name} ({brand_name}): {count}张")


if __name__ == '__main__':
    main()
