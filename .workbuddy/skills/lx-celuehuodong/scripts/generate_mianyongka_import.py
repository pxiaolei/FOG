#!/usr/bin/env python3
"""
生成免佣卡导入文件

从"城市策略活动表.xlsm"的"免佣卡"sheet读取数据，按品牌分组生成免佣卡导入文件。

用法:
    python generate_mianyongka_import.py --date-range 0420-0423
    python generate_mianyongka_import.py --date-range 0420-0423 --file "城市策略活动表2604版_v2.xlsm"
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config_loader import load_celuehuodong_config, parse_date_range_token

# 路径配置
SKILL_DIR = Path(__file__).parent.parent
ASSETS_DIR = SKILL_DIR / "assets"
TEMPLATE_FILE = ASSETS_DIR / "免佣卡导入模版.xlsx"

# 固定值
FIXED_TAG = "普通卡"
FIXED_STOCK = 999
FIXED_GET_TYPE = "购买"
FIXED_CARD_RULE = "每天前x单免佣"
FIXED_ORDER_LIMIT = 99
FIXED_ACTIVITY_RULE = """【活动规则】
在活动指定范围内，订单全部收入归为司机收入，平台不抽佣

【奖励限制条件】
1.若奖励活动有区域要求，则订单起始位置在指定活动区域外订单不计入奖励
2.抢单大厅的订单不计入奖励，指派订单按车主接单时间计算

【免佣卡与其他活动或权益的兼容性】
1.无限次免佣卡与免佣活动并存时，优先核销免佣卡活动
2.有限次免佣卡与免佣活动并存时，优先核销免佣活动
3.免佣卡与"X免X"免佣活动并存（如2免1）时，优先核销免佣卡活动
4.同一笔订单，有多张免佣卡可生效时，每笔订单至多生效一个免佣卡权益
5.本活动不能与部分完单奖等活动同时享受
6.如有疑问，可以咨询客服工作人员

【免佣卡可退款规则】
1.卡券时段重复：如周卡与日卡/峰期卡时段重复，可退日卡/峰期卡；如日卡与日卡/峰期卡时段重复，可退日卡/峰期卡
2.因封禁、交通事故、出车城市不符等原因需退款，需符合免佣卡过期未核销条件，提供封禁凭证/交通事故认定书/近期出车城市订单等依据登记退款
3.如有疑问，可以咨询客服工作人员

【注意事项】
1.如车主有恶意取消订单，或诱导乘客取消订单行为，将无法得到当日奖励

【反作弊规则】
如果用户以任何不正当手段或舞弊方式参与本活动，一经发现，平台有权对该用户的账号进行异常标记。对于账号异常的用户的奖励发放，平台有权对该用户采取不发现金或奖励等措施，亦有权收回用户可领取的奖励，追讨已发放的奖励，并保留追究该用户责任的权利。如因此给用户造成的损失，平台不进行任何赔偿或补偿。不正当手段及舞弊行为包括但不限于：使用非法工具分享、下载、安装、注册、登录多个账号、恶意宣传、虚假订单等及其他不正当手段。"""


def parse_date_range(date_range: str) -> tuple:
    """解析日期区间字符串，返回(start_date, end_date)"""
    start, end = parse_date_range_token(date_range)
    return datetime.combine(start, datetime.min.time()), datetime.combine(end, datetime.min.time())


def read_mianyongka_data(source_file: Path, date_range: str = None) -> pd.DataFrame:
    """读取免佣卡sheet数据"""
    from datetime import datetime, timedelta

    wb = load_workbook(source_file, data_only=True)
    ws = wb["免佣卡"]

    def first_value(*values):
        for value in values:
            if value is not None and not (isinstance(value, float) and pd.isna(value)):
                return value
        return None

    # 解析日期范围
    start_date, end_date = None, None
    if date_range:
        start_date, end_date = parse_date_range(date_range)

    # 收集所有数据行
    rows = []
    for row in range(3, ws.max_row + 1):
        city = ws.cell(row, 1).value  # A列 城市名称
        brand = ws.cell(row, 2).value  # B列 品牌
        card_name = ws.cell(row, 5).value  # E列 免佣卡名称
        if not city or not brand or not card_name:
            continue

        # 按日期范围过滤
        if start_date and end_date:
            date_val = ws.cell(row, 6).value  # F列 日期
            if date_val and isinstance(date_val, datetime):
                date_only = date_val.replace(hour=0, minute=0, second=0, microsecond=0)
                if not (start_date <= date_only <= end_date):
                    continue

        # 读取各列数据（时段已格式化好，直接从AN-AY列读取）
        row_data = {
            "城市": city,
            "品牌": brand,
            "免佣卡名称": card_name,
            "日期": ws.cell(row, 6).value,  # F列
            "卡券天数": ws.cell(row, 8).value,  # H列
            "时段1开始": first_value(ws.cell(row, 40).value, ws.cell(row, 9).value),  # AN列 / I列
            "时段1结束": first_value(ws.cell(row, 41).value, ws.cell(row, 10).value),  # AO列 / J列
            "时段2开始": first_value(ws.cell(row, 42).value, ws.cell(row, 11).value),  # AP列 / K列
            "时段2结束": first_value(ws.cell(row, 43).value, ws.cell(row, 12).value),  # AQ列 / L列
            "时段3开始": first_value(ws.cell(row, 44).value, ws.cell(row, 13).value),  # AR列 / M列
            "时段3结束": first_value(ws.cell(row, 45).value, ws.cell(row, 14).value),  # AS列 / N列
            "时段4开始": first_value(ws.cell(row, 46).value, ws.cell(row, 15).value),  # AT列 / O列
            "时段4结束": first_value(ws.cell(row, 47).value, ws.cell(row, 16).value),  # AU列 / P列
            "时段5开始": first_value(ws.cell(row, 48).value, ws.cell(row, 17).value),  # AV列 / Q列
            "时段5结束": first_value(ws.cell(row, 49).value, ws.cell(row, 18).value),  # AW列 / R列
            "修正价格": ws.cell(row, 22).value,  # V列
            "补贴类型": ws.cell(row, 24).value,  # X列
            # 时段小时数（用于计算售卖结束时间）
            "时段1开始小时": ws.cell(row, 9).value,   # I列
            "时段1结束小时": ws.cell(row, 10).value,  # J列
            "时段2开始小时": ws.cell(row, 11).value,  # K列
            "时段2结束小时": ws.cell(row, 12).value,  # L列
            "时段3开始小时": ws.cell(row, 13).value,  # M列
            "时段3结束小时": ws.cell(row, 14).value,  # N列
            "时段4开始小时": ws.cell(row, 15).value,  # O列
            "时段4结束小时": ws.cell(row, 16).value,  # P列
            "时段5开始小时": ws.cell(row, 17).value,  # Q列
            "时段5结束小时": ws.cell(row, 18).value,  # R列
        }
        rows.append(row_data)

    wb.close()
    return pd.DataFrame(rows)


def parse_time_value(val) -> str:
    """解析时段值，返回HH:MM:SS格式字符串"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, str):
        # 已经是HH:MM:SS格式
        return val
    # 数字按小时处理
    h = int(val)
    return f"{h:02d}:00:00"


def generate_import_file(df_brand: pd.DataFrame, date_range: str, brand_name: str, output_dir: Path, template_file: Path):
    """为一个品牌生成导入文件"""
    # 加载模版
    wb_template = load_workbook(template_file)
    ws_template = wb_template["活动信息"]

    # 获取表头行（第1行）和示例行（第2行）
    # 模版结构: A=活动序号, B=活动标题, C=城市, D-Y=时段, Z=b补品类, AA=活动要求
    template_headers = [ws_template.cell(1, col).value for col in range(1, 28)]
    template_examples = [ws_template.cell(2, col).value for col in range(1, 28)]

    # 创建新工作簿
    wb_out = load_workbook(template_file)
    ws_out = wb_out["活动信息"]

    # 清空数据行（从第3行开始）
    for row in range(3, ws_out.max_row + 1):
        for col in range(1, 28):
            ws_out.cell(row, col).value = None

    # 写入数据
    seq = 1
    from datetime import datetime, timedelta
    now = datetime.now()

    for idx, (_, row) in enumerate(df_brand.iterrows()):
        out_row = idx + 3  # 数据从第3行开始

        # A: 活动序号
        ws_out.cell(out_row, 1).value = seq
        # B: 活动标题 = 免佣卡名称
        ws_out.cell(out_row, 2).value = row["免佣卡名称"]
        # C: 城市
        ws_out.cell(out_row, 3).value = row["城市"]
        # D: 卡券售卖开始时间
        # 提前1天售卖，如果那天是今天则用现在+30分钟，否则用当天0点
        effect_date = row["日期"]
        if effect_date:
            if isinstance(effect_date, str):
                effect_date = pd.to_datetime(effect_date)
            sale_start = effect_date - pd.Timedelta(days=1)
            if sale_start.date() == now.date():
                sale_start = now + timedelta(minutes=30)
            ws_out.cell(out_row, 4).value = sale_start.strftime("%Y-%m-%d %H:%M:%S")
        # E: 卡券售卖结束时间 = 有效期当天 + 最后一个时段的结束小时 - 1小时
        if effect_date:
            if isinstance(effect_date, str):
                effect_date = pd.to_datetime(effect_date)
            # 找最后一个时段的结束小时
            last_end_hour = None
            for col_name in ["时段5结束小时", "时段4结束小时", "时段3结束小时", "时段2结束小时", "时段1结束小时"]:
                end_h = row.get(col_name)
                if end_h is not None and not (isinstance(end_h, float) and pd.isna(end_h)):
                    last_end_hour = int(end_h)
                    break
            if last_end_hour is not None:
                sale_end = effect_date.replace(hour=last_end_hour) - timedelta(hours=1)
                ws_out.cell(out_row, 5).value = sale_end.strftime("%Y-%m-%d %H:%M:%S")
            else:
                ws_out.cell(out_row, 5).value = effect_date.strftime("%Y-%m-%d 23:59:59")
        # F: 卡券有效期开始时间
        if effect_date:
            if isinstance(effect_date, str):
                effect_date = pd.to_datetime(effect_date)
            ws_out.cell(out_row, 6).value = effect_date.strftime("%Y-%m-%d 00:00:00")
        # G: 卡券有效期结束时间 = 有效期开始+天数-1 23:59:59
        if effect_date and row["卡券天数"]:
            if isinstance(effect_date, str):
                effect_date = pd.to_datetime(effect_date)
            days = int(row["卡券天数"])
            end_date = effect_date + pd.Timedelta(days=days - 1)
            ws_out.cell(out_row, 7).value = end_date.strftime("%Y-%m-%d 23:59:59")
        # H: 卡标签
        ws_out.cell(out_row, 8).value = FIXED_TAG
        # I: 库存量
        ws_out.cell(out_row, 9).value = FIXED_STOCK
        # J: 获取方式
        ws_out.cell(out_row, 10).value = FIXED_GET_TYPE
        # K: 购买金额 = 修正价格
        ws_out.cell(out_row, 11).value = row["修正价格"]
        # L: 免佣规则
        ws_out.cell(out_row, 12).value = FIXED_CARD_RULE
        # M: 免佣单量数量
        ws_out.cell(out_row, 13).value = FIXED_ORDER_LIMIT
        # N: 参与人群 - 空
        ws_out.cell(out_row, 14).value = ""
        # O: 屏蔽人群 - 空
        ws_out.cell(out_row, 15).value = ""
        # P/Q: 时段1开始/结束
        ws_out.cell(out_row, 16).value = parse_time_value(row["时段1开始"])
        ws_out.cell(out_row, 17).value = parse_time_value(row["时段1结束"])
        # R/S: 时段2开始/结束
        ws_out.cell(out_row, 18).value = parse_time_value(row["时段2开始"])
        ws_out.cell(out_row, 19).value = parse_time_value(row["时段2结束"])
        # T/U: 时段3开始/结束
        ws_out.cell(out_row, 20).value = parse_time_value(row["时段3开始"])
        ws_out.cell(out_row, 21).value = parse_time_value(row["时段3结束"])
        # V/W: 时段4开始/结束
        ws_out.cell(out_row, 22).value = parse_time_value(row["时段4开始"])
        ws_out.cell(out_row, 23).value = parse_time_value(row["时段4结束"])
        # X/Y: 时段5开始/结束
        ws_out.cell(out_row, 24).value = parse_time_value(row["时段5开始"])
        ws_out.cell(out_row, 25).value = parse_time_value(row["时段5结束"])
        # Z: b补品类
        ws_out.cell(out_row, 26).value = row["补贴类型"]
        # AA: 活动要求
        ws_out.cell(out_row, 27).value = FIXED_ACTIVITY_RULE

        seq += 1

    # 保存文件
    output_file = output_dir / f"{date_range}{brand_name}免佣卡.xlsx"
    wb_out.save(output_file)
    wb_out.close()
    return output_file


def generate_import_files(date_range: str, source_file: Path, output_dir: Path, template_file: Path = TEMPLATE_FILE) -> list[Path]:
    source_file = Path(source_file)
    output_dir = Path(output_dir)

    if not source_file.exists():
        raise FileNotFoundError(f"源文件不存在 {source_file}")

    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"读取免佣卡数据: {source_file}, 日期范围: {date_range}")
    df = read_mianyongka_data(source_file, date_range)
    print(f"共读取 {len(df)} 条免佣卡记录")
    if df.empty:
        return []

    # 按品牌分组
    brands = df["品牌"].unique()
    print(f"涉及 {len(brands)} 个品牌: {list(brands)}")

    # 生成各品牌的导入文件
    files = []
    for brand in brands:
        df_brand = df[df["品牌"] == brand]
        output_file = generate_import_file(df_brand, date_range, brand, output_dir, template_file)
        files.append(output_file)
        print(f"  生成: {output_file.name} ({len(df_brand)} 条)")

    print(f"\n完成！共生成 {len(files)} 个文件")
    return files


def main():
    parser = argparse.ArgumentParser(description="生成免佣卡导入文件")
    parser.add_argument("--date-range", required=True, help="日期区间，如 0420-0423")
    parser.add_argument("--file", help="源xlsm文件路径")
    parser.add_argument("--output-dir", help="输出目录")
    parser.add_argument("--confirmed", action="store_true", help="确认生成后台导入文件")
    args = parser.parse_args()

    config = load_celuehuodong_config()
    source_file = Path(args.file).expanduser() if args.file else config["strategy_workbook_path"]
    output_dir = Path(args.output_dir).expanduser() if args.output_dir else config["import_output_dir_path"]

    if not args.confirmed:
        print("免佣卡导入文件预览")
        print(f"源文件: {source_file}")
        print(f"输出目录: {output_dir}")
        print(f"日期区间: {args.date_range}")
        print("未写入。确认无误后追加 --confirmed 执行。")
        return

    generate_import_files(args.date_range, source_file, output_dir)


if __name__ == "__main__":
    main()
