"""
共补活动sheet更新脚本

功能：从数据库读取共补数据，追加到"共补活动"sheet
- 为免佣卡sheet提供BC列公式的数据源
- 设置正确的日期格式和H列公式
- 支持批量执行和从配置文件读取城市列表

执行顺序：
1. 先执行此脚本更新共补活动sheet
2. 再执行create_mianyongka.py生成免佣卡

使用示例：
    # 从配置读取城市列表
    python3 update_gongbu_activity.py --file "城市策略活动表.xlsm" --start 2026-04-20 --end 2026-04-23

    # 指定城市
    python3 update_gongbu_activity.py --file "城市策略活动表.xlsm" --start 2026-04-20 --end 2026-04-23 --cities "福州市,郑州市"

    # 批量模式（自动从配置读取）
    python3 update_gongbu_activity.py --batch --file "城市策略活动表.xlsm" --start 2026-04-20 --end 2026-04-23
"""

import os
import sys
import argparse
from datetime import datetime, date

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 添加 lxx_share 到路径
_skills_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, _skills_dir)

from config_loader import load_celuehuodong_config
from lxx_share import DatabaseConnector


# 默认数据库配置路径
DEFAULT_DB_CONFIG_PATH = 'config/database.yaml'

# 默认skill配置路径
DEFAULT_SKILL_CONFIG_PATH = os.path.join(os.path.dirname(__file__), '..', 'config.yaml')


def load_skill_config(config_path=None):
    """加载skill配置文件"""
    if config_path is None:
        return load_celuehuodong_config()
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def get_target_cities(config):
    """从配置获取目标城市列表"""
    if config is None:
        return None

    # 优先使用 target_cities
    cities = config.get('target_cities', [])
    if cities:
        return cities

    # 兼容：从 cities 配置中提取
    return list(config.get('cities', {}).keys())


def get_gongbu_data(conn, cities, start_date, end_date):
    """从数据库获取共补数据

    返回: list of dict
    """
    cur = conn.cursor()

    # city_id 不再使用 dim 表，直接置空

    # 查询共补数据
    cur.execute('''
        SELECT city_name, strategy_type,
               start_time_1, end_time_1,
               start_time_2, end_time_2,
               start_time_3, end_time_3,
               start_time_4, end_time_4,
               start_time_5, end_time_5,
               is_isolated_city, non_b_isolated_time_range, date
        FROM hhdata.fact_gongbu_strategy
        WHERE city_name IN %s AND date >= %s AND date <= %s
        ORDER BY date, city_name, strategy_type
    ''', (tuple(cities), start_date, end_date))

    rows = cur.fetchall()
    cur.close()

    results = []
    for row in rows:
        city_name = row[0]
        strategy_type = row[1]
        is_isolated = row[12]
        non_b_time_range = row[13]
        date_val = row[14]
        city_id = 'N/A'

        # 构建时段字符串
        periods = []
        for i in range(5):
            start_t = row[2 + i*2]
            end_t = row[3 + i*2]
            if start_t and end_t:
                start_str = start_t.strftime('%H:%M:%S')
                # 如果结束时间是00:00:00，显示为24:00:00
                end_str = '24:00:00' if (end_t.hour == 0 and end_t.minute == 0 and end_t.second == 0) else end_t.strftime('%H:%M:%S')
                periods.append(f"{start_str}-{end_str}")
        time_period = ", ".join(periods) if periods else ""

        results.append({
            'city_id': city_id,
            'city_name': city_name,
            'strategy_type': strategy_type,
            'time_period': time_period,
            'is_isolated': is_isolated,
            'non_b_time_range': non_b_time_range,
            'date': date_val
        })

    return results


def update_gongbu_activity_sheet(file_path, start_date, end_date, cities=None, output_path=None):
    """更新共补活动sheet

    参数：
        file_path: xlsm文件路径
        start_date: 开始日期
        end_date: 结束日期
        cities: 城市列表（可选，默认使用TARGET_CITIES）
        output_path: 输出文件路径（可选）

    返回：
        新增记录数
    """
    if cities is None:
        cities = get_target_cities(load_skill_config())

    # 使用 DatabaseConnector 获取数据库连接
    db = DatabaseConnector()

    with db.connect() as conn:
        # 从数据库获取共补数据
        gongbu_data = get_gongbu_data(conn, cities, start_date, end_date)

    if not gongbu_data:
        print(f"无新增共补数据（{start_date} ~ {end_date}）")
        return 0

    # 加载xlsm文件
    wb = load_workbook(file_path, keep_vba=True)

    if '共补活动' not in wb.sheetnames:
        print("未找到'共补活动'sheet")
        wb.close()
        return 0

    ws = wb['共补活动']

    # 样式配置
    font = Font(name='思源黑体', size=10)
    alignment = Alignment(horizontal='center', vertical='center')

    # 找到最后一行
    last_row = ws.max_row

    # 追加数据
    for data in gongbu_data:
        new_row = last_row + 1
        last_row += 1

        # A列: city_id
        ws.cell(row=new_row, column=1, value=data['city_id'])

        # B列: 城市名称
        ws.cell(row=new_row, column=2, value=data['city_name'])

        # C列: 策略类型
        ws.cell(row=new_row, column=3, value=data['strategy_type'])

        # D列: 策略时段
        ws.cell(row=new_row, column=4, value=data['time_period'])

        # E列: 联盟端非B补隔离城市（非B补隔离填城市名，B补隔离留空）
        ws.cell(row=new_row, column=5, value=data['city_name'] if not data['is_isolated'] else None)

        # F列: 联盟端策略时段
        ws.cell(row=new_row, column=6, value=data['non_b_time_range'] if not data['is_isolated'] else None)

        # G列: 日期，格式 yyyy/m/d
        cell_g = ws.cell(row=new_row, column=7)
        cell_g.value = data['date']
        cell_g.number_format = 'yyyy/m/d;@'

        # H列: 公式 =B&C&TEXT(G,"yyyymmdd")
        ws.cell(row=new_row, column=8, value=f'=B{new_row}&C{new_row}&TEXT(G{new_row},"yyyymmdd")')

        # 应用样式
        for col in range(1, 9):
            ws.cell(row=new_row, column=col).font = font
            ws.cell(row=new_row, column=col).alignment = alignment

    # 保存文件
    save_path = output_path or file_path
    wb.save(save_path)
    wb.close()

    print(f"已追加 {len(gongbu_data)} 条共补数据到'共补活动'sheet")
    print(f"保存到: {save_path}")

    return len(gongbu_data)


def main():
    parser = argparse.ArgumentParser(description='更新共补活动sheet（为免佣卡提供数据源）')
    parser.add_argument('--file', required=True, help='xlsm文件路径')
    parser.add_argument('--start', required=True, help='开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end', required=True, help='结束日期 (YYYY-MM-DD)')
    parser.add_argument('--cities', help='城市列表（可选，逗号分隔）')
    parser.add_argument('--batch', action='store_true', help='批量模式：从配置文件读取城市列表')
    parser.add_argument('--config', help='skill配置文件路径（可选）')
    parser.add_argument('--output', help='输出文件路径（可选）')

    args = parser.parse_args()

    start_date = datetime.strptime(args.start, '%Y-%m-%d').date()
    end_date = datetime.strptime(args.end, '%Y-%m-%d').date()

    # 加载skill配置
    config = load_skill_config(args.config)

    # 确定城市列表
    if args.cities:
        cities = [c.strip() for c in args.cities.split(',')]
    elif args.batch or config:
        # 批量模式或配置文件存在时，从配置读取
        cities = get_target_cities(config)
        if cities is None:
            print("未找到城市配置，请使用 --cities 参数指定城市")
            return
    else:
        # 默认：尝试从配置读取
        cities = get_target_cities(config)
        if cities is None:
            print("请指定城市列表：使用 --cities 或 --batch 参数")
            return

    print(f"处理城市: {', '.join(cities)}")

    count = update_gongbu_activity_sheet(
        file_path=args.file,
        start_date=start_date,
        end_date=end_date,
        cities=cities,
        output_path=args.output
    )

    if count > 0:
        print(f"\n提示: 共补活动sheet已更新，现在可以执行 create_mianyongka.py 生成免佣卡")


if __name__ == '__main__':
    main()
