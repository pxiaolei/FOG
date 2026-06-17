"""
城市策略活动表 - 共补时段日历更新脚本（数据库版）

功能：从数据库读取共补时段数据，更新到城市sheet的日历中
- 支持批量执行和从配置文件读取城市列表

使用示例：
    # 单城市
    python3 update_gongbu_calendar.py --file "城市策略活动表.xlsm" --city "福州市" --start 2026-04-20 --end 2026-04-23

    # 批量执行所有配置的城市
    python3 update_gongbu_calendar.py --batch --file "城市策略活动表.xlsm" --start 2026-04-20 --end 2026-04-23

    # 指定多个城市
    python3 update_gongbu_calendar.py --cities "福州市,郑州市" --file "城市策略活动表.xlsm" --start 2026-04-20 --end 2026-04-23
"""

import os
import sys
import argparse
from datetime import datetime, date

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 添加 lxx_share 到路径
_skills_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, _skills_dir)

from config_loader import load_celuehuodong_config
from lxx_share import DatabaseConnector


# 样式配置
STYLE_CONFIG = {
    'font': Font(name='思源黑体', size=9),
    'fill': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
    'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
}

# 策略类型显示文本映射
STRATEGY_TEXT = {
    '免佣': '共补免佣',
    '流水25%': '共补流水25%'
}

# 默认数据库配置路径
DEFAULT_DB_CONFIG_PATH = 'config/database.yaml'

# 默认skill配置路径
DEFAULT_SKILL_CONFIG_PATH = os.path.join(os.path.dirname(__file__), '..', 'config.yaml')


def load_skill_config(config_path=None):
    """加载skill配置文件"""
    if config_path is None:
        return load_celuehuodong_config()
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def get_target_cities(config):
    """从配置获取目标城市列表"""
    if config is None:
        return None

    # calendar 可单独配置，避免复用共补活动 sheet 的城市范围。
    cities = config.get('calendar_cities', [])
    if cities:
        return cities

    # 兼容旧配置：未声明 calendar_cities 时沿用 target_cities。
    cities = config.get('target_cities', [])
    if cities:
        return cities

    return list(config.get('cities', {}).keys())


def get_gongbu_data(conn, city_name, start_date, end_date):
    """从数据库获取指定城市日期范围的共补数据

    返回: list of dict，每个dict包含:
        - date: 日期
        - strategy_type: 策略类型
        - periods: [(start_hour, end_hour), ...] 时段列表
        - is_isolated: 是否B补隔离
    """
    cur = conn.cursor()
    cur.execute('''
        SELECT date, strategy_type,
               start_time_1, end_time_1,
               start_time_2, end_time_2,
               start_time_3, end_time_3,
               start_time_4, end_time_4,
               start_time_5, end_time_5,
               is_isolated_city
        FROM hhdata.fact_gongbu_strategy
        WHERE city_name = %s AND date >= %s AND date <= %s
        ORDER BY date
    ''', (city_name, start_date, end_date))

    rows = cur.fetchall()
    cur.close()

    results = []
    for row in rows:
        date_val = row[0]
        strategy_type = row[1]
        is_isolated = row[12]

        # 解析时段
        periods = []
        for i in range(5):
            start_t = row[2 + i*2]
            end_t = row[3 + i*2]
            if start_t and end_t:
                start_h = start_t.hour
                end_h = end_t.hour
                # 00:00:00 表示24点（一天的结束）
                if end_h == 0:
                    end_h = 24
                periods.append((start_h, end_h))

        if periods:
            results.append({
                'date': date_val,
                'strategy_type': strategy_type,
                'periods': periods,
                'is_isolated': is_isolated or False
            })

    return results


def find_hour_row_offset(sheet):
    """动态检测时段起始行

    在B列查找"0点"或类似文本，确定时段行的起始偏移量。
    """
    for row_idx in range(1, 50):
        cell = sheet.cell(row=row_idx, column=2)
        value = str(cell.value) if cell.value else ""
        if "0点" in value or "0:00" in value:
            return row_idx
    return 9  # 默认


def find_date_column(sheet, target_date):
    """在城市sheet中找到指定日期对应的列号

    第3行是日期行，从列3开始横向展开
    """
    target = target_date if isinstance(target_date, date) else target_date.date()

    for col_idx in range(3, sheet.max_column + 1):
        cell = sheet.cell(row=3, column=col_idx)
        if cell.value is None:
            continue

        cell_date = cell.value
        if hasattr(cell_date, 'date'):
            cell_date = cell_date.date()

        if cell_date == target:
            return col_idx

    return None


def update_city_calendar(file_path, city_name, start_date, end_date, output_path=None):
    """更新城市策略活动表的共补时段日历

    参数：
        file_path: xlsm文件路径
        city_name: 城市名称（如"福州市"）
        start_date: 开始日期
        end_date: 结束日期
        output_path: 输出文件路径（可选，默认覆盖原文件）

    返回：
        更新记录列表
    """
    # 使用 DatabaseConnector 获取数据库连接
    db = DatabaseConnector()

    with db.connect() as conn:
        # 从数据库获取共补数据
        gongbu_data = get_gongbu_data(conn, city_name, start_date, end_date)

    if not gongbu_data:
        print(f"{city_name} {start_date} 到 {end_date}: 无共补数据")
        return []

    # 加载xlsm文件
    wb = load_workbook(file_path, keep_vba=True)

    if city_name not in wb.sheetnames:
        print(f"未找到城市sheet: {city_name}")
        wb.close()
        return []

    sheet = wb[city_name]

    # 动态检测时段起始行
    hour_offset = find_hour_row_offset(sheet)

    # 更新记录
    update_records = []

    for data in gongbu_data:
        target_date = data['date']
        strategy_type = data['strategy_type']
        periods = data['periods']
        is_isolated = data['is_isolated']

        # 找日期列
        col_idx = find_date_column(sheet, target_date)

        if col_idx is None:
            print(f"未找到日期 {target_date} 对应的列")
            continue

        # 先解除该列时段区域的所有合并
        existing_merges = [
            m for m in sheet.merged_cells.ranges
            if m.min_col == col_idx and m.max_col == col_idx
            and m.min_row >= hour_offset and m.max_row <= hour_offset + 23
        ]
        for m in existing_merges:
            sheet.unmerge_cells(str(m))

        # 清空该列时段区域
        for hr in range(hour_offset, hour_offset + 24):
            cell = sheet.cell(row=hr, column=col_idx)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

        # 构建显示文本
        bbu_status = "B补隔离" if is_isolated else ""
        display_text = STRATEGY_TEXT.get(strategy_type, strategy_type)

        # 写入时段
        for (start_h, end_h) in periods:
            # end_h是结束边界（不包含），如6-10表示覆盖6点到9点59分
            # 显示结束时间 = end_h - 1:59
            display_end_h = end_h - 1 if end_h > start_h else start_h

            start_row = hour_offset + start_h
            end_row = hour_offset + display_end_h

            # 显示文本格式：6:00-9:59共补免佣
            text = f"{start_h}:00-{display_end_h}:59{display_text}"
            if bbu_status:
                text += f"-{bbu_status}"

            # 合并单元格（如果跨度大于1行）
            if end_row > start_row:
                sheet.merge_cells(
                    start_row=start_row, start_column=col_idx,
                    end_row=end_row, end_column=col_idx
                )

            # 设置内容
            cell = sheet.cell(row=start_row, column=col_idx)
            cell.value = text
            cell.fill = STYLE_CONFIG['fill']
            cell.font = STYLE_CONFIG['font']
            cell.alignment = STYLE_CONFIG['alignment']

            # 记录更新
            update_records.append({
                '日期': target_date.strftime('%Y-%m-%d'),
                '列号': col_idx,
                '策略类型': strategy_type,
                '时段': f"{start_h}:00-{end_h}:00",
                '显示文本': text,
                '合并行': f"{start_row}-{end_row}"
            })

    # 保存文件
    save_path = output_path or file_path
    wb.save(save_path)
    wb.close()

    print(f"已更新 {len(update_records)} 个时段，保存到: {save_path}")

    return update_records


def main():
    parser = argparse.ArgumentParser(description='更新城市策略活动表共补时段日历（数据库版）')
    parser.add_argument('--file', required=True, help='xlsm文件路径')
    parser.add_argument('--city', help='城市名称（单城市模式）')
    parser.add_argument('--cities', help='城市列表（逗号分隔）')
    parser.add_argument('--batch', action='store_true', help='批量模式：从配置文件读取城市列表')
    parser.add_argument('--start', required=True, help='开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end', required=True, help='结束日期 (YYYY-MM-DD)')
    parser.add_argument('--output', help='输出文件路径（可选）')
    parser.add_argument('--config', help='skill配置文件路径（可选）')

    args = parser.parse_args()

    start_date = datetime.strptime(args.start, '%Y-%m-%d').date()
    end_date = datetime.strptime(args.end, '%Y-%m-%d').date()

    # 加载skill配置
    config = load_skill_config(args.config)

    # 确定城市列表
    if args.cities:
        cities = [c.strip() for c in args.cities.split(',')]
    elif args.city:
        cities = [args.city]
    elif args.batch or config:
        # 批量模式或配置文件存在时，从配置读取
        cities = get_target_cities(config)
        if cities is None:
            print("未找到城市配置，请使用 --city 或 --cities 参数指定城市")
            return
    else:
        # 默认：尝试从配置读取
        cities = get_target_cities(config)
        if cities is None:
            print("请指定城市：使用 --city、--cities 或 --batch 参数")
            return

    print(f"处理城市: {', '.join(cities)}")

    total_periods = 0
    all_records = []

    for city_name in cities:
        print(f"\n处理：{city_name}")

        records = update_city_calendar(
            file_path=args.file,
            city_name=city_name,
            start_date=start_date,
            end_date=end_date,
            output_path=args.output
        )

        total_periods += len(records)
        all_records.extend(records)

        # 打印更新记录
        for r in records:
            print(f"  {r['日期']} 列{r['列号']}: {r['显示文本']} (行{r['合并行']})")

    print(f"\n===== 汇总 =====")
    print(f"共更新 {total_periods} 个时段")
    for city in cities:
        count = sum(1 for r in all_records if city in r.get('显示文本', '') or True)  # 按城市分组
        print(f"  {city}: {sum(1 for r in all_records if r.get('城市', city) == city)}个时段")


if __name__ == '__main__':
    main()
