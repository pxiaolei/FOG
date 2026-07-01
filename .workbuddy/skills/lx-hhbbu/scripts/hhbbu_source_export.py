#!/usr/bin/env python3
"""
从公司 dataReporting 库按 date_day + city_name + brand_name 导出 hhdata B补相关金额。
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _find_project_root() -> Path:
    for candidate in [Path(__file__).resolve(), *Path(__file__).resolve().parents]:
        if (candidate / ".workbuddy" / "skills").is_dir() and (candidate / "config").is_dir():
            return candidate
    return Path(__file__).resolve().parents[3]


PROJECT_ROOT = _find_project_root()
SKILLS_DIR = PROJECT_ROOT / ".workbuddy" / "skills"
LX_SHUJUKU_SCRIPTS_DIR = SKILLS_DIR / "lx_shujuku" / "scripts"

for path in (SKILLS_DIR, LX_SHUJUKU_SCRIPTS_DIR):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from lxx_share.fog_config import get_section, resolve_project_path  # noqa: E402
from lx_shujuku import create_client  # noqa: E402


Q2 = Decimal("0.01")
IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
WRITABLE_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "workspace" / "02数据导入" / "处理日志" / "lx-hhbbu"
TARGET_TYPES = {"excel_file", "database_table"}
TARGET_COLUMN_ALIASES = {
    "date": ("日期",),
    "city_name": ("城市名称", "城市"),
    "brand_name": ("品牌名称", "品牌"),
    "total_b_subsidy": ("总b补金额", "总B补金额"),
    "merchant_b_subsidy": ("商家b补金额", "商家B补金额"),
    "card_merchant_income": ("售卡商家收入金额", "售卡商家收入"),
}
UPDATE_FIELDS = {
    "total_b_subsidy": "总b补金额",
    "merchant_b_subsidy": "商家b补金额",
    "card_merchant_income": "售卡商家收入金额",
}


@dataclass(frozen=True)
class HhdataTargetInfo:
    status: str
    target: str
    source: str
    message: str
    path: str = ""
    selected_file: str = ""
    table: str = ""
    city_dim_table: str = ""
    brand_dim_table: str = ""


@dataclass(frozen=True)
class Key:
    date_day: str
    city_name: str
    brand_name: str


@dataclass
class SourceAmounts:
    activity_total_reward: Decimal = Decimal("0")
    activity_merchant_subsidy: Decimal = Decimal("0")
    coupon_total_subsidy: Decimal = Decimal("0")
    coupon_merchant_subsidy: Decimal = Decimal("0")
    merchant_coupon_sales_revenue: Decimal = Decimal("0")

    @property
    def total_b_subsidy(self) -> Decimal:
        return money(self.activity_total_reward + self.coupon_total_subsidy)

    @property
    def merchant_b_subsidy(self) -> Decimal:
        return money(self.activity_merchant_subsidy + self.coupon_merchant_subsidy)

    @property
    def card_merchant_income(self) -> Decimal:
        return money(self.merchant_coupon_sales_revenue)

    @property
    def has_nonzero_target_amount(self) -> bool:
        return any(
            value != 0
            for value in (
                self.total_b_subsidy,
                self.merchant_b_subsidy,
                self.card_merchant_income,
            )
        )


def parse_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"日期格式必须是 YYYY-MM-DD: {value}") from exc


def iter_dates(start: date, end: date) -> list[str]:
    if start > end:
        raise ValueError("--start-date 不能晚于 --end-date")
    days: list[str] = []
    current = start
    while current <= end:
        days.append(current.isoformat())
        current += timedelta(days=1)
    return days


def to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value))


def money(value: Any) -> Decimal:
    return to_decimal(value).quantize(Q2, rounding=ROUND_HALF_UP)


def fmt_money(value: Any) -> str:
    return f"{money(value):.2f}"


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value)
    if not text:
        return ""
    text = text.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def load_hhbbu_config() -> dict[str, Any]:
    return get_section("lx_hhbbu", PROJECT_ROOT)


def resolve_runtime_path(value: str | Path) -> Path:
    path = Path(str(value)).expanduser()
    if path.is_absolute():
        return path
    return resolve_project_path(path, PROJECT_ROOT)


def is_excel_candidate(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES and not path.name.startswith("~$")


def quote_identifier(name: str) -> str:
    text = normalize_text(name)
    parts = text.split(".")
    if not text or any(not IDENTIFIER_RE.fullmatch(part) for part in parts):
        raise ValueError(f"非法数据库标识符: {name!r}")
    return ".".join(f"`{part}`" for part in parts)


def dict_config(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def inspect_hhdata_target(
    *,
    hhdata_target: str | None,
    hhdata_file: str | None,
    hhdata_table: str | None,
    city_dim_table: str | None,
    brand_dim_table: str | None,
    require_hhdata_target: bool,
    config: dict[str, Any],
) -> HhdataTargetInfo:
    local_config = dict_config(config.get("local_hhdata"))
    config_requires = bool(local_config.get("required_before_run"))
    required = require_hhdata_target or config_requires
    target = normalize_text(hhdata_target or local_config.get("target") or "excel_file")
    failure_status = "error" if required else "warning"

    if target not in TARGET_TYPES:
        return HhdataTargetInfo(
            status="error",
            target=target,
            source="fog_config.yaml:lx_hhbbu.local_hhdata.target",
            message="hhdata target 只能是 excel_file 或 database_table",
        )

    if target == "database_table":
        database_config = dict_config(local_config.get("database_table"))
        table = normalize_text(hhdata_table or database_config.get("table") or local_config.get("table"))
        city_table = normalize_text(city_dim_table or database_config.get("city_dim_table") or local_config.get("city_dim_table"))
        brand_table = normalize_text(brand_dim_table or database_config.get("brand_dim_table") or local_config.get("brand_dim_table"))
        if not table or not city_table or not brand_table:
            return HhdataTargetInfo(
                status=failure_status,
                target=target,
                source="fog_config.yaml:lx_hhbbu.local_hhdata.database_table",
                message="database_table 目标必须配置 table、city_dim_table、brand_dim_table",
                table=table,
                city_dim_table=city_table,
                brand_dim_table=brand_table,
            )
        return HhdataTargetInfo(
            status="ok",
            target=target,
            source="fog_config.yaml:lx_hhbbu.local_hhdata.database_table",
            message=f"已定位 hhdata 数据库表: {table}",
            table=table,
            city_dim_table=city_table,
            brand_dim_table=brand_table,
        )

    location_value: str | Path | None
    source: str
    explicit = False
    excel_config = dict_config(local_config.get("excel_file"))
    if hhdata_file:
        location_value = hhdata_file
        source = "--hhdata-file"
        explicit = True
    elif excel_config.get("file"):
        location_value = excel_config.get("file")
        source = "fog_config.yaml:lx_hhbbu.local_hhdata.excel_file.file"
    elif local_config.get("file"):
        location_value = local_config.get("file")
        source = "fog_config.yaml:lx_hhbbu.local_hhdata.file"
    else:
        return HhdataTargetInfo(
            status=failure_status,
            target=target,
            source="fog_config.yaml:lx_hhbbu.local_hhdata.excel_file.file",
            message="未配置唯一 hhdata Excel 文件路径；请填写 lx_hhbbu.local_hhdata.excel_file.file 或传 --hhdata-file",
        )

    path = resolve_runtime_path(location_value)
    failure_status = "error" if required or explicit else "warning"

    if is_excel_candidate(path):
        return HhdataTargetInfo(
            status="ok",
            target=target,
            source=source,
            path=str(path),
            message=f"已定位唯一本地 hhdata Excel: {path}",
            selected_file=str(path),
        )
    return HhdataTargetInfo(
        status=failure_status,
        target=target,
        source=source,
        path=str(path),
        message=f"未找到可用本地 hhdata Excel: {path}",
    )


def hhdata_target_to_dict(info: HhdataTargetInfo) -> dict[str, Any]:
    return {
        "status": info.status,
        "target": info.target,
        "source": info.source,
        "path": info.path,
        "message": info.message,
        "selected_file": info.selected_file,
        "table": info.table,
        "city_dim_table": info.city_dim_table,
        "brand_dim_table": info.brand_dim_table,
    }


def print_hhdata_target_info(info: HhdataTargetInfo) -> None:
    prefix = {"ok": "[ok]", "warning": "[warning]", "error": "[error]"}.get(info.status, "[info]")
    print(f"{prefix} hhdata 写回目标: {info.message}")
    print(f"目标类型: {info.target}")
    print(f"来源: {info.source}")
    if info.selected_file:
        print(f"写回文件: {info.selected_file}")
    if info.table:
        print(f"写回表: {info.table}")
        print(f"城市维表: {info.city_dim_table}")
        print(f"品牌维表: {info.brand_dim_table}")


def find_header(headers: dict[str, int], field: str) -> int | None:
    for alias in TARGET_COLUMN_ALIASES[field]:
        if alias in headers:
            return headers[alias]
    return None


def detect_hhdata_sheet(workbook: Any, sheet_name: str = "") -> tuple[Any, int, dict[str, int]]:
    sheets = [workbook[sheet_name]] if sheet_name else [workbook[name] for name in workbook.sheetnames]
    for ws in sheets:
        if ws.sheet_state != "visible":
            continue
        for row_idx in range(1, min(ws.max_row, 10) + 1):
            headers = {
                normalize_text(ws.cell(row=row_idx, column=col_idx).value): col_idx
                for col_idx in range(1, ws.max_column + 1)
                if normalize_text(ws.cell(row=row_idx, column=col_idx).value)
            }
            resolved = {field: find_header(headers, field) for field in TARGET_COLUMN_ALIASES}
            if all(resolved.values()):
                return ws, row_idx, {field: int(col_idx) for field, col_idx in resolved.items() if col_idx}
    target = f"工作表 {sheet_name!r}" if sheet_name else "任一可见工作表"
    raise RuntimeError(f"未在{target}找到 hhdata 表头：需要日期、城市名称、品牌名称、总b补金额、商家b补金额、售卡商家收入金额")


def source_update_values(amounts: SourceAmounts) -> dict[str, Decimal]:
    return {
        "total_b_subsidy": amounts.total_b_subsidy,
        "merchant_b_subsidy": amounts.merchant_b_subsidy,
        "card_merchant_income": amounts.card_merchant_income,
    }


def make_backup(file_path: Path, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{file_path.stem}.before_lx-hhbbu_{timestamp}{file_path.suffix}"
    shutil.copy2(file_path, backup_path)
    return backup_path


def build_hhdata_update(
    *,
    source: dict[Key, SourceAmounts],
    hhdata_target: HhdataTargetInfo,
    output_dir: Path,
    sheet_name: str = "",
    backup_dir: str = "",
    confirmed: bool = False,
) -> dict[str, Any]:
    if not hhdata_target.selected_file:
        return {
            "status": "error",
            "target": "excel_file",
            "message": "未能唯一定位要写回的 hhdata Excel；请用 --hhdata-file 或 lx_hhbbu.local_hhdata.excel_file.file 指定单个文件",
            "confirmed": confirmed,
        }

    file_path = Path(hhdata_target.selected_file)
    if file_path.suffix.lower() not in WRITABLE_EXCEL_SUFFIXES:
        return {
            "status": "error",
            "target": "excel_file",
            "message": f"当前只支持写回 .xlsx/.xlsm 文件，不支持: {file_path}",
            "confirmed": confirmed,
        }

    workbook = load_workbook(file_path)
    try:
        ws, header_row, columns = detect_hhdata_sheet(workbook, sheet_name)
        rows_by_key: dict[Key, list[int]] = defaultdict(list)
        row_meta: dict[int, Key] = {}
        for row_idx in range(header_row + 1, ws.max_row + 1):
            date_day = normalize_date(ws.cell(row=row_idx, column=columns["date"]).value)
            city_name = normalize_text(ws.cell(row=row_idx, column=columns["city_name"]).value)
            brand_name = normalize_text(ws.cell(row=row_idx, column=columns["brand_name"]).value)
            if not date_day or not city_name or not brand_name:
                continue
            key = Key(date_day, city_name, brand_name)
            rows_by_key[key].append(row_idx)
            row_meta[row_idx] = key

        duplicate_keys = {key for key, row_indexes in rows_by_key.items() if len(row_indexes) > 1}
        source_keys = {key for key, amounts in source.items() if amounts.has_nonzero_target_amount}
        changed_rows: set[int] = set()
        unchanged_rows = 0
        missing_source_rows = []
        skipped_duplicate_rows = []
        changes = []

        for row_idx, key in sorted(row_meta.items()):
            if key in duplicate_keys:
                skipped_duplicate_rows.append({
                    "row": row_idx,
                    "date": key.date_day,
                    "city_name": key.city_name,
                    "brand_name": key.brand_name,
                    "reason": "本地 Excel 存在重复 date+city_name+brand_name，跳过避免重复写入",
                })
                continue
            amounts = source.get(key)
            if amounts is None or not amounts.has_nonzero_target_amount:
                missing_source_rows.append({
                    "row": row_idx,
                    "date": key.date_day,
                    "city_name": key.city_name,
                    "brand_name": key.brand_name,
                })
                continue

            row_changed = False
            for field, display_name in UPDATE_FIELDS.items():
                col_idx = columns[field]
                cell = ws.cell(row=row_idx, column=col_idx)
                old_value = money(cell.value)
                new_value = source_update_values(amounts)[field]
                if old_value == new_value:
                    continue
                row_changed = True
                changed_rows.add(row_idx)
                changes.append({
                    "row": row_idx,
                    "date": key.date_day,
                    "city_name": key.city_name,
                    "brand_name": key.brand_name,
                    "field": field,
                    "column": display_name,
                    "old_value": fmt_money(old_value),
                    "new_value": fmt_money(new_value),
                })
                if confirmed:
                    cell.value = float(new_value)
                    cell.number_format = "0.00"
            if not row_changed:
                unchanged_rows += 1

        local_unique_keys = {key for key, row_indexes in rows_by_key.items() if len(row_indexes) == 1}
        source_not_in_local = sorted(source_keys - local_unique_keys, key=lambda key: (key.date_day, key.city_name, key.brand_name))
        backup_path = ""
        if confirmed and changes:
            backup_root = resolve_runtime_path(backup_dir) if backup_dir else output_dir / "backups"
            backup_path = str(make_backup(file_path, backup_root))
            workbook.save(file_path)

        mode = "confirmed" if confirmed else "dry_run"
        status = "updated" if confirmed and changes else "no_changes" if not changes else "dry_run"
        return {
            "status": status,
            "target": "excel_file",
            "mode": mode,
            "confirmed": confirmed,
            "message": "已写回本地 hhdata Excel" if confirmed and changes else "未保存 Excel；这是 dry-run 更新计划" if changes else "没有需要写回的单元格",
            "file": str(file_path),
            "sheet": ws.title,
            "header_row": header_row,
            "backup_path": backup_path,
            "local_key_count": len(rows_by_key),
            "duplicate_local_key_count": len(duplicate_keys),
            "skipped_duplicate_row_count": len(skipped_duplicate_rows),
            "missing_source_row_count": len(missing_source_rows),
            "source_not_in_local_count": len(source_not_in_local),
            "changed_row_count": len(changed_rows),
            "changed_cell_count": len(changes),
            "unchanged_matched_row_count": unchanged_rows,
            "changes_sample": changes[:200],
            "skipped_duplicate_rows_sample": skipped_duplicate_rows[:50],
            "missing_source_rows_sample": missing_source_rows[:50],
            "source_not_in_local_sample": [
                {"date": key.date_day, "city_name": key.city_name, "brand_name": key.brand_name}
                for key in source_not_in_local[:50]
            ],
        }
    finally:
        workbook.close()


def build_database_update(
    *,
    source: dict[Key, SourceAmounts],
    hhdata_target: HhdataTargetInfo,
    confirmed: bool = False,
) -> dict[str, Any]:
    source_keys = {key for key, amounts in source.items() if amounts.has_nonzero_target_amount}
    if not source_keys:
        return {
            "status": "no_changes",
            "target": "database_table",
            "mode": "confirmed" if confirmed else "dry_run",
            "confirmed": confirmed,
            "message": "公司源没有可写回的非零金额",
            "table": hhdata_target.table,
            "changed_row_count": 0,
            "changed_cell_count": 0,
        }

    try:
        table = quote_identifier(hhdata_target.table)
        city_dim_table = quote_identifier(hhdata_target.city_dim_table)
        brand_dim_table = quote_identifier(hhdata_target.brand_dim_table)
    except ValueError as exc:
        return {
            "status": "error",
            "target": "database_table",
            "mode": "confirmed" if confirmed else "dry_run",
            "confirmed": confirmed,
            "message": str(exc),
            "table": hhdata_target.table,
        }

    start_day = min(key.date_day for key in source_keys)
    end_day = max(key.date_day for key in source_keys)
    select_sql = f"""
        SELECT
            f.id,
            DATE_FORMAT(f.date_day, '%%Y-%%m-%%d') AS date_day,
            f.city_id,
            c.city_name,
            f.brand_id,
            b.brand_name,
            f.total_b_subsidy,
            f.merchant_b_subsidy,
            f.card_merchant_income
        FROM {table} f
        LEFT JOIN {city_dim_table} c ON f.city_id = c.city_id
        LEFT JOIN {brand_dim_table} b ON f.brand_id = b.brand_id
        WHERE f.date_day BETWEEN %s AND %s
    """
    update_sql = f"""
        UPDATE {table}
        SET total_b_subsidy = %s,
            merchant_b_subsidy = %s,
            card_merchant_income = %s
        WHERE id = %s
    """

    try:
        from lxx_share.database import DatabaseConnector  # noqa: PLC0415

        db = DatabaseConnector()
        with db.connect() as conn:
            cursor = conn.cursor()
            cursor.execute(select_sql, [start_day, end_day])
            columns = [desc[0] for desc in cursor.description or []]
            local_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

            rows_by_key: dict[Key, list[dict[str, Any]]] = defaultdict(list)
            row_meta: dict[Any, Key] = {}
            row_by_id: dict[Any, dict[str, Any]] = {}
            for row in local_rows:
                date_day = normalize_date(row.get("date_day"))
                city_name = normalize_text(row.get("city_name"))
                brand_name = normalize_text(row.get("brand_name"))
                row_id = row.get("id")
                if not row_id or not date_day or not city_name or not brand_name:
                    continue
                key = Key(date_day, city_name, brand_name)
                rows_by_key[key].append(row)
                row_meta[row_id] = key
                row_by_id[row_id] = row

            duplicate_keys = {key for key, rows in rows_by_key.items() if len(rows) > 1}
            changed_row_ids: set[Any] = set()
            unchanged_rows = 0
            missing_source_rows = []
            skipped_duplicate_rows = []
            changes = []
            row_updates: dict[Any, dict[str, Decimal]] = {}

            for row_id, key in sorted(row_meta.items(), key=lambda item: (item[1].date_day, item[1].city_name, item[1].brand_name, str(item[0]))):
                row = row_by_id[row_id]
                if key in duplicate_keys:
                    skipped_duplicate_rows.append({
                        "id": row_id,
                        "date": key.date_day,
                        "city_name": key.city_name,
                        "brand_name": key.brand_name,
                        "reason": "数据库表存在重复 date+city_name+brand_name，跳过避免重复写入",
                    })
                    continue
                amounts = source.get(key)
                if amounts is None or not amounts.has_nonzero_target_amount:
                    missing_source_rows.append({
                        "id": row_id,
                        "date": key.date_day,
                        "city_name": key.city_name,
                        "brand_name": key.brand_name,
                    })
                    continue

                update_values = source_update_values(amounts)
                row_changed = False
                for field, display_name in UPDATE_FIELDS.items():
                    old_value = money(row.get(field))
                    new_value = update_values[field]
                    if old_value == new_value:
                        continue
                    row_changed = True
                    changed_row_ids.add(row_id)
                    changes.append({
                        "id": row_id,
                        "date": key.date_day,
                        "city_name": key.city_name,
                        "brand_name": key.brand_name,
                        "field": field,
                        "column": display_name,
                        "old_value": fmt_money(old_value),
                        "new_value": fmt_money(new_value),
                    })
                if row_changed:
                    row_updates[row_id] = update_values
                else:
                    unchanged_rows += 1

            updated_row_count = 0
            if confirmed and row_updates:
                for row_id, values in row_updates.items():
                    cursor.execute(
                        update_sql,
                        [
                            values["total_b_subsidy"],
                            values["merchant_b_subsidy"],
                            values["card_merchant_income"],
                            row_id,
                        ],
                    )
                    updated_row_count += cursor.rowcount
                conn.commit()
            else:
                conn.rollback()

        local_unique_keys = {key for key, rows in rows_by_key.items() if len(rows) == 1}
        source_not_in_local = sorted(source_keys - local_unique_keys, key=lambda key: (key.date_day, key.city_name, key.brand_name))
        mode = "confirmed" if confirmed else "dry_run"
        status = "updated" if confirmed and changes else "no_changes" if not changes else "dry_run"
        return {
            "status": status,
            "target": "database_table",
            "mode": mode,
            "confirmed": confirmed,
            "message": "已写回 hhdata 数据库表" if confirmed and changes else "未写数据库；这是 dry-run 更新计划" if changes else "没有需要写回的数据库行",
            "table": hhdata_target.table,
            "city_dim_table": hhdata_target.city_dim_table,
            "brand_dim_table": hhdata_target.brand_dim_table,
            "date_range": {"start": start_day, "end": end_day},
            "local_key_count": len(rows_by_key),
            "duplicate_local_key_count": len(duplicate_keys),
            "skipped_duplicate_row_count": len(skipped_duplicate_rows),
            "missing_source_row_count": len(missing_source_rows),
            "source_not_in_local_count": len(source_not_in_local),
            "changed_row_count": len(changed_row_ids),
            "changed_cell_count": len(changes),
            "updated_row_count": updated_row_count,
            "unchanged_matched_row_count": unchanged_rows,
            "changes_sample": changes[:200],
            "skipped_duplicate_rows_sample": skipped_duplicate_rows[:50],
            "missing_source_rows_sample": missing_source_rows[:50],
            "source_not_in_local_sample": [
                {"date": key.date_day, "city_name": key.city_name, "brand_name": key.brand_name}
                for key in source_not_in_local[:50]
            ],
        }
    except Exception as exc:
        return {
            "status": "error",
            "target": "database_table",
            "mode": "confirmed" if confirmed else "dry_run",
            "confirmed": confirmed,
            "message": f"数据库写回失败: {exc}",
            "table": hhdata_target.table,
        }


def fetch_source(dates: list[str], source_limit: int) -> tuple[dict[Key, SourceAmounts], dict[str, Any]]:
    client = create_client()
    source: dict[Key, SourceAmounts] = defaultdict(SourceAmounts)
    source_counts: dict[str, dict[str, int]] = {}

    for date_day in dates:
        activity_sql = f"""
            SELECT date_day, city_name, brand_name,
                   SUM(COALESCE(total_reward_amount, 0)) AS activity_total_reward,
                   SUM(COALESCE(merchant_subsidy_amount, 0)) AS activity_merchant_subsidy
            FROM honghu_activity_marketing_data
            WHERE date_day = '{date_day}'
            GROUP BY date_day, city_name, brand_name
            LIMIT {source_limit}
        """
        coupon_sql = f"""
            SELECT date_day, city_name, brand_name,
                   SUM(COALESCE(total_subsidy_amount, 0)) AS coupon_total_subsidy,
                   SUM(COALESCE(merchant_subsidy_amount, 0)) AS coupon_merchant_subsidy,
                   SUM(COALESCE(merchant_coupon_sales_revenue, 0)) AS merchant_coupon_sales_revenue
            FROM honghu_coupon_marketing_data
            WHERE date_day = '{date_day}'
            GROUP BY date_day, city_name, brand_name
            LIMIT {source_limit}
        """
        activity_rows = client.execute(activity_sql)
        coupon_rows = client.execute(coupon_sql)
        if len(activity_rows) >= source_limit or len(coupon_rows) >= source_limit:
            raise RuntimeError(
                f"{date_day} 公司库聚合结果达到 source_limit={source_limit}，可能被截断；"
                "请提高 lx_shujuku max_limit 或缩小日期范围后重试。"
            )

        source_counts[date_day] = {
            "activity_groups": len(activity_rows),
            "coupon_groups": len(coupon_rows),
        }

        for row in activity_rows:
            key = Key(row["date_day"], row.get("city_name") or "", row.get("brand_name") or "")
            source[key].activity_total_reward += to_decimal(row.get("activity_total_reward"))
            source[key].activity_merchant_subsidy += to_decimal(row.get("activity_merchant_subsidy"))

        for row in coupon_rows:
            key = Key(row["date_day"], row.get("city_name") or "", row.get("brand_name") or "")
            source[key].coupon_total_subsidy += to_decimal(row.get("coupon_total_subsidy"))
            source[key].coupon_merchant_subsidy += to_decimal(row.get("coupon_merchant_subsidy"))
            source[key].merchant_coupon_sales_revenue += to_decimal(row.get("merchant_coupon_sales_revenue"))

    return source, {"group_counts": source_counts}


def build_rows(source: dict[Key, SourceAmounts]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for key, amounts in sorted(source.items(), key=lambda item: (item[0].date_day, item[0].city_name, item[0].brand_name)):
        if not amounts.has_nonzero_target_amount:
            continue
        rows.append({
            "date": key.date_day,
            "city_name": key.city_name,
            "brand_name": key.brand_name,
            "total_b_subsidy": fmt_money(amounts.total_b_subsidy),
            "merchant_b_subsidy": fmt_money(amounts.merchant_b_subsidy),
            "card_merchant_income": fmt_money(amounts.card_merchant_income),
            "activity_total_reward": fmt_money(amounts.activity_total_reward),
            "activity_merchant_subsidy": fmt_money(amounts.activity_merchant_subsidy),
            "coupon_total_subsidy": fmt_money(amounts.coupon_total_subsidy),
            "coupon_merchant_subsidy": fmt_money(amounts.coupon_merchant_subsidy),
            "merchant_coupon_sales_revenue": fmt_money(amounts.merchant_coupon_sales_revenue),
        })
    return rows


def summarize_by_date(source: dict[Key, SourceAmounts], dates: list[str]) -> dict[str, dict[str, str]]:
    summary: dict[str, dict[str, Decimal]] = {
        day: {
            "total_b_subsidy": Decimal("0"),
            "merchant_b_subsidy": Decimal("0"),
            "card_merchant_income": Decimal("0"),
        }
        for day in dates
    }
    for key, amounts in source.items():
        if key.date_day not in summary:
            continue
        summary[key.date_day]["total_b_subsidy"] += amounts.total_b_subsidy
        summary[key.date_day]["merchant_b_subsidy"] += amounts.merchant_b_subsidy
        summary[key.date_day]["card_merchant_income"] += amounts.card_merchant_income
    return {
        day: {metric: fmt_money(value) for metric, value in values.items()}
        for day, values in summary.items()
    }


def write_outputs(report: dict[str, Any], output_dir: Path) -> tuple[Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"{timestamp}_lx-hhbbu_source.csv"
    json_path = output_dir / f"{timestamp}_lx-hhbbu_source.json"
    md_path = output_dir / f"{timestamp}_lx-hhbbu_source.md"

    fieldnames = [
        "date",
        "city_name",
        "brand_name",
        "total_b_subsidy",
        "merchant_b_subsidy",
        "card_merchant_income",
        "activity_total_reward",
        "activity_merchant_subsidy",
        "coupon_total_subsidy",
        "coupon_merchant_subsidy",
        "merchant_coupon_sales_revenue",
    ]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report["rows"])

    report["outputs"] = {
        "csv": str(csv_path),
        "json": str(json_path),
        "markdown": str(md_path),
    }
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    md_path.write_text(render_markdown(report), encoding="utf-8")
    return csv_path, json_path, md_path


def render_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# lx-hhbbu 公司库来源导出",
        "",
        f"- 聚合键: `date + city_name + brand_name`",
        f"- 日期范围: {report['date_range']['start']} 至 {report['date_range']['end']}",
        f"- 输出行数: {report['row_count']}",
        f"- CSV: `{report['outputs']['csv']}`",
        "",
        "## hhdata 写回目标",
        "",
        f"- 状态: `{report['hhdata_target']['status']}`",
        f"- 目标类型: `{report['hhdata_target']['target']}`",
        f"- 来源: `{report['hhdata_target']['source']}`",
        f"- 路径: `{report['hhdata_target']['path']}`",
        f"- 表: `{report['hhdata_target']['table']}`",
        f"- 说明: {report['hhdata_target']['message']}",
        "",
    ]
    if report.get("target_update"):
        update = report["target_update"]
        heading = "本地 hhdata Excel 写回" if update.get("target") == "excel_file" else "hhdata 数据库表写回"
        lines.extend([
            f"## {heading}",
            "",
            f"- 状态: `{update['status']}`",
            f"- 模式: `{update['mode']}`",
            f"- 目标类型: `{update.get('target', '')}`",
            f"- 文件: `{update.get('file', '')}`",
            f"- 表: `{update.get('table', '')}`",
            f"- 工作表: `{update.get('sheet', '')}`",
            f"- 变更行数: {update.get('changed_row_count', 0)}",
            f"- 变更单元格数: {update.get('changed_cell_count', 0)}",
            f"- 数据库已更新行数: {update.get('updated_row_count', 0)}",
            f"- 跳过重复 key 行数: {update.get('skipped_duplicate_row_count', 0)}",
            f"- 本地无公司源行数: {update.get('missing_source_row_count', 0)}",
            f"- 公司源未命中本地 key 数: {update.get('source_not_in_local_count', 0)}",
            f"- 备份: `{update.get('backup_path', '')}`",
            f"- 说明: {update.get('message', '')}",
            "",
        ])
    lines.extend([
        "## 来源汇总",
        "",
        "| 日期 | 总b补金额 | 商家b补金额 | 售卡商家收入金额 |",
        "|---|---:|---:|---:|",
    ])
    for day, values in report["source_summary"].items():
        lines.append(
            f"| {day} | {values['total_b_subsidy']} | "
            f"{values['merchant_b_subsidy']} | {values['card_merchant_income']} |"
        )
    lines.extend([
        "",
        "## 字段说明",
        "",
        "- `date`、`city_name`、`brand_name`：公司库聚合键。",
        "- `total_b_subsidy`：活动总奖励金额 + 卡券总补贴金额。",
        "- `merchant_b_subsidy`：活动商家补贴金额 + 卡券商家补贴金额。",
        "- `card_merchant_income`：卡券商家券后售卡收入。",
        "- 其余 `activity_*` / `coupon_*` 字段是来源拆分金额。",
        "",
    ])
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    config = load_hhbbu_config()
    source_limit = int(config.get("source_limit") or 1000)
    parser = argparse.ArgumentParser(description="按 date + city_name + brand_name 导出 hhdata B补来源金额")
    parser.add_argument("--start-date", type=parse_date, help="开始日期 YYYY-MM-DD")
    parser.add_argument("--end-date", type=parse_date, help="结束日期 YYYY-MM-DD")
    parser.add_argument("--source-limit", type=int, default=source_limit, help="公司库单日聚合查询 LIMIT，默认 1000")
    parser.add_argument("--hhdata-target", choices=sorted(TARGET_TYPES), help="hhdata 写回目标类型；优先于 fog_config.yaml")
    parser.add_argument("--hhdata-file", help="本地 hhdata Excel 文件路径；优先于 fog_config.yaml")
    parser.add_argument("--hhdata-table", help="hhdata 数据库事实表；target=database_table 时优先于 fog_config.yaml")
    parser.add_argument("--city-dim-table", help="城市维表；target=database_table 时优先于 fog_config.yaml")
    parser.add_argument("--brand-dim-table", help="品牌维表；target=database_table 时优先于 fog_config.yaml")
    parser.add_argument("--require-local-hhdata", action="store_true", help="找不到 hhdata 写回目标时失败")
    parser.add_argument("--check-local-hhdata", action="store_true", help="只检查 hhdata 写回目标，不查询公司库")
    parser.add_argument("--update-hhdata", action="store_true", help="按公司源生成 hhdata 三列写回计划；默认不保存")
    parser.add_argument("--confirmed", action="store_true", help="确认保存 hhdata 写回；必须和 --update-hhdata 一起使用")
    parser.add_argument("--hhdata-sheet", help="本地 hhdata 工作表名称；默认自动检测")
    parser.add_argument("--backup-dir", help="确认写回前的备份目录；默认写到输出目录 backups 子目录")
    parser.add_argument(
        "--output-dir",
        help="审计文件输出目录；默认读取 fog_config.yaml 的 lx_hhbbu.output_dir",
    )
    args = parser.parse_args()
    output_dir = args.output_dir or config.get("output_dir") or DEFAULT_OUTPUT_DIR
    args.output_dir = str(resolve_runtime_path(output_dir))
    args.hhbbu_config = config
    local_config = dict_config(config.get("local_hhdata"))
    excel_config = dict_config(local_config.get("excel_file"))
    args.hhdata_target = args.hhdata_target or local_config.get("target") or "excel_file"
    args.hhdata_sheet = args.hhdata_sheet or excel_config.get("sheet_name") or local_config.get("sheet_name") or ""
    args.backup_dir = args.backup_dir or excel_config.get("backup_dir") or local_config.get("backup_dir") or ""
    if args.confirmed and not args.update_hhdata:
        parser.error("--confirmed 必须和 --update-hhdata 一起使用")
    if not args.check_local_hhdata and (not args.start_date or not args.end_date):
        parser.error("导出公司源时必须同时提供 --start-date 和 --end-date；只检查 hhdata 写回目标可使用 --check-local-hhdata")
    return args


def main() -> int:
    args = parse_args()
    hhdata_target = inspect_hhdata_target(
        hhdata_target=args.hhdata_target,
        hhdata_file=args.hhdata_file,
        hhdata_table=args.hhdata_table,
        city_dim_table=args.city_dim_table,
        brand_dim_table=args.brand_dim_table,
        require_hhdata_target=args.require_local_hhdata or args.update_hhdata,
        config=args.hhbbu_config,
    )
    if args.check_local_hhdata:
        print_hhdata_target_info(hhdata_target)
        return 1 if hhdata_target.status == "error" else 0
    if hhdata_target.status == "error":
        print_hhdata_target_info(hhdata_target)
        return 1

    dates = iter_dates(args.start_date, args.end_date)
    source, source_meta = fetch_source(dates, args.source_limit)
    source_meta["hhdata_target"] = hhdata_target_to_dict(hhdata_target)
    rows = build_rows(source)
    source_summary = summarize_by_date(source, dates)
    workbook_update = None
    target_update = None
    if args.update_hhdata:
        if hhdata_target.target == "excel_file":
            target_update = build_hhdata_update(
                source=source,
                hhdata_target=hhdata_target,
                output_dir=Path(args.output_dir),
                sheet_name=args.hhdata_sheet,
                backup_dir=args.backup_dir,
                confirmed=args.confirmed,
            )
            workbook_update = target_update
        else:
            target_update = build_database_update(
                source=source,
                hhdata_target=hhdata_target,
                confirmed=args.confirmed,
            )
        if target_update.get("status") == "error":
            print(f"[error] hhdata 写回失败: {target_update.get('message', '')}")
            return 1
    report = {
        "type": "lx-hhbbu.source_export",
        "version": 3,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "args": {
            "start_date": dates[0],
            "end_date": dates[-1],
            "source_limit": args.source_limit,
            "output_dir": args.output_dir,
            "hhdata_target": args.hhdata_target,
            "hhdata_file": args.hhdata_file,
            "hhdata_table": args.hhdata_table,
            "city_dim_table": args.city_dim_table,
            "brand_dim_table": args.brand_dim_table,
            "require_local_hhdata": args.require_local_hhdata,
            "update_hhdata": args.update_hhdata,
            "confirmed": args.confirmed,
            "hhdata_sheet": args.hhdata_sheet,
            "backup_dir": args.backup_dir,
        },
        "date_range": {"start": dates[0], "end": dates[-1], "days": len(dates)},
        "key": ["date", "city_name", "brand_name"],
        "hhdata_target": hhdata_target_to_dict(hhdata_target),
        "local_hhdata": hhdata_target_to_dict(hhdata_target),
        "target_update": target_update,
        "workbook_update": workbook_update,
        "source_meta": source_meta,
        "source_summary": source_summary,
        "row_count": len(rows),
        "rows": rows,
    }
    csv_path, json_path, md_path = write_outputs(report, Path(args.output_dir))

    print("lx-hhbbu 公司库来源导出")
    print("聚合键: date + city_name + brand_name")
    print(f"日期范围: {dates[0]} 至 {dates[-1]}")
    print(f"输出行数: {len(rows)}")
    print("来源汇总:")
    for day, values in source_summary.items():
        print(
            f"- {day}: 总b补={values['total_b_subsidy']} "
            f"商家b补={values['merchant_b_subsidy']} "
            f"售卡商家收入={values['card_merchant_income']}"
        )
    print_hhdata_target_info(hhdata_target)
    if target_update:
        print("hhdata 写回:")
        print(f"- 目标类型: {target_update.get('target', '')}")
        print(f"- 状态: {target_update['status']}")
        print(f"- 模式: {target_update['mode']}")
        print(f"- 变更行数: {target_update.get('changed_row_count', 0)}")
        print(f"- 变更单元格数: {target_update.get('changed_cell_count', 0)}")
        if target_update.get("updated_row_count"):
            print(f"- 数据库已更新行数: {target_update['updated_row_count']}")
        if target_update.get("backup_path"):
            print(f"- 备份: {target_update['backup_path']}")
        print(f"- 说明: {target_update.get('message', '')}")
    print(f"CSV: {csv_path}")
    print(f"审计 JSON: {json_path}")
    print(f"审计 Markdown: {md_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
