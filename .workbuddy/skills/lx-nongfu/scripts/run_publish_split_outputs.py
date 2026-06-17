#!/usr/bin/env python3
"""Publish already-split Excel files into operator daily-info Feishu sheets."""

from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
import time
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from run_split_publish import (
    LarkCli,
    NongfuError,
    OperatorPublish,
    PROJECT_ROOT,
    TargetWorkbook,
    a1_range,
    configured_list,
    configured_value_map,
    create_sheet,
    load_config,
    parse_annotated_csv,
    pad_row,
    resolve_lark_cli,
    resolve_operator_root_folder,
    resolve_targets,
    rows_to_csv,
    target_sheet_link,
)


@dataclass
class SplitEntry:
    batch: str
    operator: str
    sheet_name: str
    path: Path
    rows: list[list[Any]]
    column_count: int
    target: TargetWorkbook | None = None
    sheet_id: str = ""
    link: str = ""
    status: str = "pending"
    reason: str = ""


def source_label(path: Path) -> str:
    stem = path.stem if path.suffix.lower() == ".zip" else path.name
    return re.sub(r"^\d{8}_\d{4}_", "", stem)


def resolve_source_dir(path: Path, temp_dirs: list[tempfile.TemporaryDirectory[str]]) -> Path:
    if path.is_dir():
        return path
    if path.suffix.lower() != ".zip":
        raise NongfuError(f"输入不是目录或 zip: {path}")
    sibling = path.with_suffix("")
    if sibling.is_dir():
        return sibling
    tmp = tempfile.TemporaryDirectory(prefix="lx-nongfu-split-")
    temp_dirs.append(tmp)
    with zipfile.ZipFile(path) as archive:
        try:
            archive.extractall(tmp.name)
        except UnicodeDecodeError as exc:
            raise NongfuError(f"zip 文件名编码无法自动解压，请先手动解压: {path}") from exc
    return Path(tmp.name)


def person_dir(source_dir: Path, contact_person: str) -> Path:
    candidate = source_dir / contact_person
    return candidate if candidate.is_dir() else source_dir


def parse_operator_and_sheet(path: Path, default_sheet_name: str) -> tuple[str, str] | None:
    """从文件名解析运营主体和 sheet 名。

    返回 (operator, sheet_name)；若文件名没有 '_' 前缀（汇总文件），返回 None 表示跳过。
    """
    stem = path.stem
    if "_" not in stem:
        # 没有下划线分隔符，视为无运营主体前缀的汇总文件，跳过
        return None
    operator, suffix = stem.split("_", 1)
    operator = operator.strip()
    if not operator:
        return None
    return operator, (suffix.strip() or default_sheet_name)


def cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        # datetime 是 date 的子类，需优先判断
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.isoformat()
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def read_workbook_rows(path: Path) -> tuple[list[list[Any]], int]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows: list[list[Any]] = []
    max_col = max(sheet.max_column or 1, 1)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_col, values_only=True):
        rows.append([cell_value(value) for value in row])
    while rows and all(value == "" for value in rows[-1]):
        rows.pop()
    if not rows:
        raise NongfuError(f"Excel 没有读到任何行: {path}")
    return rows, max(max_col, max(len(row) for row in rows), 1)


def collect_entries(paths: list[Path], contact_person: str, include_lx: bool) -> list[SplitEntry]:
    temp_dirs: list[tempfile.TemporaryDirectory[str]] = []
    entries: list[SplitEntry] = []
    try:
        for raw_path in paths:
            path = raw_path.expanduser()
            if not path.is_absolute():
                path = PROJECT_ROOT / path
            if not path.exists():
                raise NongfuError(f"输入不存在: {path}")
            source_dir = resolve_source_dir(path, temp_dirs)
            batch_name = source_label(path)
            current_person_dir = person_dir(source_dir, contact_person)
            files = sorted(item for item in current_person_dir.glob("*.xlsx") if not item.name.startswith("~$"))
            if not files:
                raise NongfuError(f"未找到 {contact_person} 的已拆分 xlsx: {current_person_dir}")
            for file_path in files:
                parsed = parse_operator_and_sheet(file_path, batch_name)
                if parsed is None:
                    # 无运营主体前缀（汇总文件），跳过
                    continue
                operator, sheet_name = parsed
                if operator == "LX" and not include_lx:
                    continue
                rows, column_count = read_workbook_rows(file_path)
                entries.append(
                    SplitEntry(
                        batch=batch_name,
                        operator=operator,
                        sheet_name=sheet_name,
                        path=file_path,
                        rows=rows,
                        column_count=column_count,
                    )
                )
    finally:
        # Keep temp dirs alive until all workbooks are read, then clean them.
        for tmp in temp_dirs:
            tmp.cleanup()
    return entries


def write_entry(cli: LarkCli, entry: SplitEntry, delay_seconds: float) -> None:
    if entry.target is None:
        raise NongfuError(f"{entry.operator} 缺少目标表格。")
    sheet_id = create_sheet(cli, entry.target, entry.sheet_name, len(entry.rows), entry.column_count)
    time.sleep(delay_seconds)
    cli.sheets(
        [
            "+csv-put",
            "--spreadsheet-token",
            entry.target.spreadsheet_token,
            "--sheet-id",
            sheet_id,
            "--start-cell",
            "A1",
            "--csv",
            "-",
        ],
        input_text=rows_to_csv(entry.rows, entry.column_count),
    )
    time.sleep(delay_seconds)
    verify = cli.sheets(
        [
            "+csv-get",
            "--spreadsheet-token",
            entry.target.spreadsheet_token,
            "--sheet-id",
            sheet_id,
            "--range",
            a1_range(len(entry.rows), entry.column_count),
        ]
    )
    data = verify.get("data") if isinstance(verify.get("data"), dict) else {}
    actual = parse_annotated_csv(str(data.get("annotated_csv") or ""))
    expected = [pad_row(row, entry.column_count) for row in entry.rows]
    actual_padded = [pad_row(row, entry.column_count) for row in actual]
    if len(actual_padded) != len(expected):
        raise NongfuError(f"{entry.operator}/{entry.sheet_name} 写后验证行数不一致：预期 {len(expected)}，实际 {len(actual_padded)}。")
    for row_index, expected_row in enumerate(expected):
        actual_row = actual_padded[row_index]
        if [str(item) for item in actual_row[: min(3, entry.column_count)]] != [
            str(item) for item in expected_row[: min(3, entry.column_count)]
        ]:
            raise NongfuError(f"{entry.operator}/{entry.sheet_name} 写后验证失败：第 {row_index + 1} 行前 3 列不一致。")
    entry.sheet_id = sheet_id
    entry.link = target_sheet_link(entry.target, sheet_id)
    entry.status = "created"


def output_paths(config: dict[str, Any], label: str) -> Path:
    workspace = "workspace/12农夫协作"
    nongfu = config.get("lx_nongfu", {})
    if isinstance(nongfu, dict) and nongfu.get("workspace_dir"):
        workspace = str(nongfu["workspace_dir"])
    output_dir = Path(workspace)
    if not output_dir.is_absolute():
        output_dir = PROJECT_ROOT / output_dir
    output_dir = output_dir / "输出"
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", label).strip("_") or "split_outputs"
    return output_dir / f"{stamp}_{safe_label}_publish_summary.json"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="lx-nongfu 已拆分 Excel 发布到飞书日常信息表。")
    parser.add_argument("inputs", nargs="+", help="已拆分输出目录或 zip；zip 同名解压目录存在时优先使用该目录")
    parser.add_argument("--contact-person", default="", help="对接人；默认取配置 lx_nongfu.default_contact_persons[0]")
    parser.add_argument("--include-lx", action="store_true", help="包含 LX 文件；默认跳过 LX")
    parser.add_argument("--operator-root-folder-token", default="", help="运营主体文件夹所在父文件夹 token")
    parser.add_argument("--operator-root-folder-url", default="", help="运营主体文件夹所在父文件夹 URL")
    parser.add_argument("--operator-folder-template", default="", help="运营主体文件夹名模板；默认 {operator}-运营主体")
    parser.add_argument("--target-table-template", default="", help="目标表格名模板；默认 {operator}-日常信息")
    parser.add_argument("--target-sheet-name", default="", help="统一覆盖目标 sheet 名；不填时使用文件名 suffix")
    parser.add_argument("--if-sheet-exists", choices=["fail", "skip"], default="fail")
    parser.add_argument("--confirmed", action="store_true", help="实际写入飞书；不加时只 dry-run")
    parser.add_argument("--write-delay-seconds", type=float, default=2.5)
    parser.add_argument("--lark-cli", default="")
    parser.add_argument("--identity", choices=["user", "bot"], default="user")
    parser.add_argument("--timeout", type=int, default=180)
    parser.add_argument("--output-json", default="")
    parser.add_argument("--no-output-files", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    config = load_config()
    nongfu = config.get("lx_nongfu", {}) if isinstance(config.get("lx_nongfu"), dict) else {}
    operator_doc = configured_value_map(config, ["lx_nongfu", "operator_doc"])
    contact_person = args.contact_person
    if not contact_person:
        defaults = configured_list(config, ["lx_nongfu", "default_contact_persons"], [])
        contact_person = defaults[0] if defaults else ""
    if not contact_person:
        raise NongfuError("缺少 --contact-person，且配置里没有 lx_nongfu.default_contact_persons。")

    root_folder_token = resolve_operator_root_folder(args, config, contact_person)
    if not root_folder_token:
        raise NongfuError(
            "缺少运营主体根文件夹。请传 --operator-root-folder-url，"
            "或在 config/fog_config.yaml 的 lx_nongfu.operator_doc.contact_person_root_folders 中按对接人配置。"
        )

    entries = collect_entries([Path(item) for item in args.inputs], contact_person, args.include_lx)
    if not entries:
        raise NongfuError("没有可发布的已拆分 Excel。")

    # 统一覆盖 sheet 名
    if args.target_sheet_name:
        for entry in entries:
            entry.sheet_name = args.target_sheet_name

    duplicate_keys: set[tuple[str, str]] = set()
    seen_keys: set[tuple[str, str]] = set()
    for entry in entries:
        key = (entry.operator, entry.sheet_name)
        if key in seen_keys:
            duplicate_keys.add(key)
        seen_keys.add(key)
    if duplicate_keys:
        raise NongfuError("同一运营主体下存在重复 sheet 名: " + json.dumps(sorted(duplicate_keys), ensure_ascii=False))

    target_table_template = args.target_table_template or str(
        operator_doc.get("target_table_name_template") or "{operator}-日常信息"
    )
    operator_folder_template = args.operator_folder_template or str(
        operator_doc.get("operator_folder_name_template") or "{operator}-运营主体"
    )
    operators = sorted({entry.operator for entry in entries})
    cli = LarkCli(resolve_lark_cli(config, args.lark_cli), identity=args.identity, timeout=args.timeout)
    targets = resolve_targets(cli, root_folder_token, operators, operator_folder_template, target_table_template)

    for entry in entries:
        target = targets.get(entry.operator)
        if not target:
            entry.status = "blocked"
            entry.reason = "target_workbook_not_found"
            continue
        entry.target = target
        existing_sheet_id = target.existing_sheets.get(entry.sheet_name, "")
        if existing_sheet_id:
            entry.sheet_id = existing_sheet_id
            entry.link = target_sheet_link(target, existing_sheet_id)
            if args.if_sheet_exists == "skip":
                entry.status = "skipped"
                entry.reason = "target_sheet_exists"
            else:
                entry.status = "blocked"
                entry.reason = "target_sheet_exists"
            continue
        entry.link = target.url
        entry.status = "ready"

    blocked = [entry for entry in entries if entry.status == "blocked"]
    if blocked and args.confirmed:
        detail = [
            {"operator": entry.operator, "sheet_name": entry.sheet_name, "reason": entry.reason}
            for entry in blocked
        ]
        raise NongfuError("存在无法写入的条目，已停止: " + json.dumps(detail, ensure_ascii=False))

    if args.confirmed:
        for entry in entries:
            if entry.status != "ready":
                continue
            write_entry(cli, entry, args.write_delay_seconds)

    summary = [
        {
            "batch": entry.batch,
            "operator": entry.operator,
            "sheet_name": entry.sheet_name,
            "status": entry.status,
            "reason": entry.reason,
            "rows": len(entry.rows),
            "columns": entry.column_count,
            "source_file": str(entry.path),
            "target_url": entry.target.url if entry.target else "",
            "sheet_id": entry.sheet_id,
            "link": entry.link,
        }
        for entry in entries
    ]
    result = {
        "ok": not blocked,
        "dry_run": not args.confirmed,
        "contact_person": contact_person,
        "inputs": args.inputs,
        "include_lx": bool(args.include_lx),
        "entries_total": len(entries),
        "entries_ready_or_created": sum(1 for entry in entries if entry.status in {"ready", "created"}),
        "entries_skipped": sum(1 for entry in entries if entry.status == "skipped"),
        "entries_blocked": sum(1 for entry in entries if entry.status == "blocked"),
        "operators_total": len(operators),
        "summary": summary,
    }
    if not args.no_output_files:
        output_json = Path(args.output_json).expanduser() if args.output_json else output_paths(config, "split_outputs")
        if not output_json.is_absolute():
            output_json = PROJECT_ROOT / output_json
        output_json.parent.mkdir(parents=True, exist_ok=True)
        output_json.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        result["output_files"] = {"json": str(output_json)}
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["ok"] else 2


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except NongfuError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
