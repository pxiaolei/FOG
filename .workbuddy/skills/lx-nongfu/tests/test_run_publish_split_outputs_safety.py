import stat
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook


SKILL_DIR = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = SKILL_DIR / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

import run_publish_split_outputs as publish  # noqa: E402
from run_split_publish import NongfuError, TargetWorkbook  # noqa: E402


class ZipSafetyTests(unittest.TestCase):
    def assert_rejected_without_target_changes(self, members):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            archive_path = root / "input.zip"
            target = root / "target"
            target.mkdir()
            sentinel = target / "sentinel.txt"
            sentinel.write_text("unchanged", encoding="utf-8")
            with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
                for member, content in members:
                    archive.writestr(member, content)

            before = sorted(path.relative_to(target) for path in target.rglob("*"))
            with patch.object(
                publish.tempfile,
                "TemporaryDirectory",
                side_effect=AssertionError("malicious zip must be rejected before extraction"),
            ), self.assertRaises(NongfuError):
                publish.collect_entries([archive_path], "测试人", include_lx=False, confirmed=True)
            after = sorted(path.relative_to(target) for path in target.rglob("*"))

            self.assertEqual(after, before)
            self.assertEqual(sentinel.read_text(encoding="utf-8"), "unchanged")

    def test_rejects_absolute_path(self):
        self.assert_rejected_without_target_changes([("/absolute.xlsx", b"x")])

    def test_rejects_parent_traversal(self):
        self.assert_rejected_without_target_changes([("../escape.xlsx", b"x")])

    def test_rejects_any_parent_path_component(self):
        self.assert_rejected_without_target_changes([("person/../operator.xlsx", b"x")])

    def test_rejects_symbolic_link(self):
        info = zipfile.ZipInfo("linked.xlsx")
        info.create_system = 3
        info.external_attr = (stat.S_IFLNK | 0o777) << 16
        self.assert_rejected_without_target_changes([(info, b"target.xlsx")])

    def test_rejects_duplicate_normalized_output_name(self):
        self.assert_rejected_without_target_changes(
            [("person/operator.xlsx", b"a"), ("person//operator.xlsx", b"b")]
        )

    def test_rejects_excessive_member_count(self):
        with patch.object(publish, "MAX_ZIP_MEMBERS", 1):
            self.assert_rejected_without_target_changes([("a.xlsx", b"a"), ("b.xlsx", b"b")])

    def test_rejects_excessive_uncompressed_size(self):
        with patch.object(publish, "MAX_ZIP_UNCOMPRESSED_BYTES", 3):
            self.assert_rejected_without_target_changes([("a.xlsx", b"1234")])

    def test_rejects_abnormal_compression_ratio(self):
        with patch.object(publish, "MAX_ZIP_COMPRESSION_RATIO", 2.0):
            self.assert_rejected_without_target_changes([("a.xlsx", b"0" * 10_000)])


class DryRunSafetyTests(unittest.TestCase):
    def test_confirmed_and_dry_run_are_mutually_exclusive(self):
        with self.assertRaises(SystemExit):
            publish.main(["input.zip", "--confirmed", "--dry-run"])

    def test_dry_run_reads_valid_zip_without_extracting_or_writing_summary(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = root / "operator_topic.xlsx"
            workbook = Workbook()
            workbook.active.append(["品牌", "城市"])
            workbook.active.append(["A", "杭州"])
            workbook.save(workbook_path)
            archive_path = root / "batch.zip"
            with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
                archive.write(workbook_path, "测试人/测试主体_topic.xlsx")
            summary_path = root / "summary.json"
            target = TargetWorkbook(
                operator="测试主体",
                folder_token="folder",
                spreadsheet_token="sheet",
                url="https://example.invalid/sheet",
            )

            config = {
                "lx_nongfu": {
                    "default_contact_persons": ["测试人"],
                    "operator_doc": {},
                }
            }
            with patch.object(publish, "load_config", return_value=config), patch.object(
                publish, "resolve_operator_root_folder", return_value="root-token"
            ), patch.object(
                publish, "resolve_targets", return_value={"测试主体": target}
            ), patch.object(
                publish.tempfile,
                "TemporaryDirectory",
                side_effect=AssertionError("dry-run must not extract zip"),
            ):
                result = publish.main(
                    [
                        str(archive_path),
                        "--contact-person",
                        "测试人",
                        "--output-json",
                        str(summary_path),
                    ]
                )

            self.assertEqual(result, 0)
            self.assertFalse(summary_path.exists())


class PublishReadbackTests(unittest.TestCase):
    def test_write_entry_rejects_mismatch_after_third_column(self):
        class FakeCli:
            def sheets(self, args, *, input_text=None, retries=1):
                if args[0] == "+csv-put":
                    return {"ok": True}
                if args[0] == "+csv-get":
                    return {"data": {"annotated_csv": "A,B,C,D\n1,2,3,WRONG"}}
                raise AssertionError(args)

        target = TargetWorkbook(
            operator="测试主体",
            folder_token="folder",
            spreadsheet_token="sheet-token",
            url="https://example.invalid/sheet",
        )
        entry = publish.SplitEntry(
            batch="batch",
            operator="测试主体",
            sheet_name="topic",
            path=Path("topic.xlsx"),
            rows=[["A", "B", "C", "D"], ["1", "2", "3", "4"]],
            column_count=4,
            target=target,
        )

        with patch.object(publish, "create_sheet", return_value="sheet-id"), patch.object(
            publish.time, "sleep", return_value=None
        ), self.assertRaisesRegex(NongfuError, "第 2 行第 4 列"):
            publish.write_entry(FakeCli(), entry, delay_seconds=0)

        self.assertEqual(entry.status, "pending")


if __name__ == "__main__":
    unittest.main()
