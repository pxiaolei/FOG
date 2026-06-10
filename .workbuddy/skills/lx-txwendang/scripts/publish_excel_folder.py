#!/usr/bin/env python3
"""批量发布本地 Excel 目录到腾讯文档在线表格。

默认匹配文件名：{运营主体}_{描述}.xlsx

支持两种模式：
- 单 sheet 模式：--sheet-name "xxx"（发布所有 Excel 的第一个 sheet）
- 多 sheet 模式：--all-sheets（发布每个 Excel 的所有 visible sheet，以原名命名）
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

def _find_skills_dir() -> Path:
    for p in Path(__file__).resolve().parents:
        if (p / "lxx_share").is_dir():
            return p
    return Path(__file__).resolve().parents[2]


_skills_dir = _find_skills_dir()
if str(_skills_dir) not in sys.path:
    sys.path.insert(0, str(_skills_dir))

from lxx_share.cache_utils import load_entity_cache, versioned_entity_cache_template  # noqa: E402

SKILL_DIR = Path(__file__).parent.parent
CACHE_PATH = SKILL_DIR / "assets" / "entity_cache.json"
LEGACY_CACHE_PATH = _skills_dir / "lx-zhutichaibiao" / "assets" / "entity_cache.json"


def resolve_cache_path(value: str | None = None) -> Path:
    if value:
        return Path(value).expanduser()
    if CACHE_PATH.exists():
        return CACHE_PATH
    if LEGACY_CACHE_PATH.exists():
        return LEGACY_CACHE_PATH
    return CACHE_PATH


def load_cache(path: Path) -> dict[str, dict[str, Any]]:
    return load_entity_cache(path)


def find_entity_files(folder: str) -> dict[str, Path]:
    source = Path(folder).expanduser()
    if not source.exists():
        raise FileNotFoundError(f"目录不存在: {source}")
    if not source.is_dir():
        raise NotADirectoryError(f"不是目录: {source}")

    entities: dict[str, Path] = {}
    for path in sorted(source.glob("*.xlsx")):
        stem = path.stem
        if "未匹配" in stem:
            continue
        if "_" not in stem:
            continue
        entity = stem.split("_", 1)[0]
        entities[entity] = path
    return entities


def read_excel_data(path: Path, sheet_name: str | None = None) -> list[list[Any]]:
    from openpyxl import load_workbook  # noqa: WPS433

    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        data = [[cell.value for cell in row] for row in ws.iter_rows()]
    finally:
        wb.close()

    while data and all(value is None or str(value).strip() == "" for value in data[-1]):
        data.pop()
    return data


def list_sheet_names(path: Path) -> list[str]:
    """返回 Excel 中所有 visible sheet 名称。"""
    from openpyxl import load_workbook  # noqa: WPS433

    wb = load_workbook(path, read_only=True)
    try:
        return [
            name for name in wb.sheetnames
            if wb[name].sheet_state == "visible"
        ]
    finally:
        wb.close()


def excel_shape(path: Path, sheet_name: str | None = None) -> tuple[int, int]:
    from openpyxl import load_workbook  # noqa: WPS433

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        return ws.max_row, ws.max_column
    finally:
        wb.close()


def preview_targets(
    entities: dict[str, Path],
    cache: dict[str, dict[str, Any]],
    sheet_name: str,
    cache_path: Path,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    print("\n=== 腾讯文档发布预览 ===")
    print(f"Sheet 名称: {sheet_name}")
    print(f"实体数量: {len(entities)}")
    print(f"缓存路径: {cache_path}")
    print()

    for entity, path in entities.items():
        row_count, col_count = excel_shape(path)
        entry = cache.get(entity, {})
        file_id = entry.get("file_id", "") if isinstance(entry, dict) else ""
        display_file_id = "已配置" if file_id else "???"
        print(f"  {entity:<12} -> {row_count}行 x {col_count}列 -> file_id: {display_file_id}")
        rows.append({
            "entity": entity,
            "path": str(path),
            "file_id": file_id,
            "rows": row_count,
            "cols": col_count,
        })
    return rows


def publish_folder(
    folder: str,
    sheet_name: str,
    dry_run: bool = False,
    force: bool = False,
    cache_path: str | None = None,
    config_path: str | None = None,
    all_sheets: bool = False,
) -> list[dict[str, Any]]:
    entities = find_entity_files(folder)
    if not entities:
        raise RuntimeError("目录中没有找到可发布的 Excel 文件")

    resolved_cache = resolve_cache_path(cache_path)
    cache = load_cache(resolved_cache)

    # resolve sheet names per entity（all-sheets 模式 vs 单一 sheet 模式）
    entity_sheets: dict[str, list[str]] = {}
    if all_sheets:
        for entity, path in entities.items():
            entity_sheets[entity] = list_sheet_names(path)
    else:
        for entity in entities:
            entity_sheets[entity] = [sheet_name]

    # preview
    print("\n=== 腾讯文档发布预览 ===")
    print(f"模式: {'all-sheets' if all_sheets else 'single-sheet'}")
    print(f"Sheet: {sheet_name if not all_sheets else '(各Excel的visible sheets)'}")
    print(f"实体数量: {len(entities)}")
    print(f"缓存路径: {resolved_cache}")
    print()

    any_missing = False
    preview_rows: list[dict[str, Any]] = []
    for entity, path in entities.items():
        entry = cache.get(entity, {})
        file_id = entry.get("file_id", "") if isinstance(entry, dict) else ""
        sheets = entity_sheets[entity]
        for sn in sheets:
            row_count, col_count = excel_shape(path, sn)
            display_file_id = "已配置" if file_id else "???"
            print(f"  {entity:<12} [{sn:<20}] -> {row_count}行 x {col_count}列 -> file_id: {display_file_id}")
            preview_rows.append({
                "entity": entity,
                "sheet": sn,
                "path": str(path),
                "file_id": file_id,
                "rows": row_count,
                "cols": col_count,
            })
            if not file_id:
                any_missing = True

    if any_missing:
        missing = list(dict.fromkeys(
            row["entity"] for row in preview_rows if not row["file_id"]
        ))
        print("\n以下实体缺少 file_id 缓存，将跳过:")
        for entity in missing:
            print(f"  - {entity}")

    if dry_run:
        print("\n预览模式，未写入腾讯文档。")
        return [
            {**row, "status": "dry-run"}
            for row in preview_rows
        ]

    if not force:
        confirm = input("\n确认发布？[y/N] ").strip().lower()
        if confirm != "y":
            print("已取消")
            return []

    from lxx_share.tdocs_api import TdocsClient  # noqa: WPS433

    client = TdocsClient(config_path=config_path)
    results: list[dict[str, Any]] = []

    for entity, path in entities.items():
        entry = cache.get(entity, {})
        file_id = entry.get("file_id", "") if isinstance(entry, dict) else ""

        if not file_id:
            sheets = entity_sheets[entity]
            print(f"  {entity}: 缺少 file_id 缓存，跳过 ({len(sheets)} sheet)")
            for sn in sheets:
                results.append({
                    "entity": entity,
                    "sheet": sn,
                    "status": "skipped",
                    "reason": "no file_id",
                })
            continue

        for sn in entity_sheets[entity]:
            try:
                data = read_excel_data(path, sn)
                if not data:
                    print(f"  {entity} [{sn}]: 文件为空，跳过")
                    results.append({"entity": entity, "sheet": sn, "status": "skipped", "reason": "empty"})
                    continue

                row_count = len(data)
                col_count = max((len(row) for row in data), default=1)
                sheet_id = client.add_sheet(
                    file_id,
                    sn,
                    row_count=max(row_count, 200),
                    column_count=col_count,
                )
                client.write_range_auto(file_id, sheet_id, data)

                print(f"  {entity} [{sn}]: {row_count}行 -> sheet_id: {sheet_id}")
                results.append({
                    "entity": entity,
                    "sheet": sn,
                    "status": "ok",
                    "file_id": file_id,
                    "sheet_id": sheet_id,
                    "rows": row_count,
                    "cols": col_count,
                })
            except Exception as exc:
                print(f"  {entity} [{sn}]: 失败 - {exc}", file=sys.stderr)
                results.append({
                    "entity": entity,
                    "sheet": sn,
                    "status": "failed",
                    "file_id": file_id,
                    "error": str(exc),
                })

    ok = sum(1 for row in results if row["status"] == "ok")
    failed = sum(1 for row in results if row["status"] == "failed")
    skipped = sum(1 for row in results if row["status"] == "skipped")
    print("\n=== 发布完成 ===")
    print(f"成功: {ok} / 失败: {failed} / 跳过: {skipped} / 总数: {len(results)}")

    if ok:
        print("\n--- sheet_id 列表 ---")
        for row in results:
            if row["status"] == "ok":
                print(f"  {row['entity']} [{row.get('sheet', '')}]: {row['sheet_id']}")

    return results


def refresh_cache_instructions() -> None:
    print("请通过腾讯文档 MCP 或人工导航补充缓存:\n")
    print("  1. 打开腾讯文档根文件夹，获取各运营主体子文件夹")
    print("  2. 找到每个主体的目标在线表格")
    print("  3. 写入 .workbuddy/skills/lx-txwendang/assets/entity_cache.json\n")
    print("缓存格式:")
    print(json.dumps(versioned_entity_cache_template(), indent=2, ensure_ascii=False))


def sanitize_results_for_output(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """输出前隐藏 file_id，避免把线上资源 ID 打到终端。"""
    sanitized: list[dict[str, Any]] = []
    for row in results:
        item = dict(row)
        if item.get("file_id"):
            item["file_id"] = "[已隐藏]"
        sanitized.append(item)
    return sanitized


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="lx-txwendang 批量发布 Excel 到腾讯文档")
    parser.add_argument("folder", nargs="?", help="本地 Excel 输出目录")
    parser.add_argument("--sheet-name", "-s", help="新建 sheet 名称（单sheet模式，与 --all-sheets 互斥）")
    parser.add_argument("--all-sheets", action="store_true", help="自动发布 Excel 中所有 visible sheet（每个 sheet 以原名发布）")
    parser.add_argument("--dry-run", "-n", action="store_true", help="只预览，不写入")
    parser.add_argument("--force", "-y", action="store_true", help="跳过确认直接发布")
    parser.add_argument("--cache-path", help="指定 entity_cache.json 路径")
    parser.add_argument("--config-path", help="指定腾讯文档 config.yaml 路径")
    parser.add_argument("--refresh-cache", action="store_true", help="输出缓存刷新指引")
    parser.add_argument("--json", action="store_true", help="输出 JSON 结果")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.refresh_cache:
        refresh_cache_instructions()
        return 0

    if not args.folder:
        print("错误：缺少本地 Excel 输出目录", file=sys.stderr)
        return 1
    if not args.all_sheets and not args.sheet_name:
        print("错误：需要 --sheet-name（单sheet模式）或 --all-sheets（全sheet模式）", file=sys.stderr)
        return 1

    try:
        results = publish_folder(
            folder=args.folder,
            sheet_name=args.sheet_name or "",
            dry_run=args.dry_run,
            force=args.force,
            cache_path=args.cache_path,
            config_path=args.config_path,
            all_sheets=args.all_sheets,
        )
    except Exception as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 1

    if args.json:
        print("\n--- JSON ---")
        print(json.dumps(sanitize_results_for_output(results), indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
