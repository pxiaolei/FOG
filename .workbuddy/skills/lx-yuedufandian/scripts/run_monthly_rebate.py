#!/usr/bin/env python3
"""月度返点计算入口。

核心原则：
- 当月口径只读取 `规则/rules.yaml`，不从聊天记录推断。
- 源 Excel 只在 `import-sources` 阶段读取；`calculate` 默认从 PostgreSQL import 批次读取。
- `--dry-run` 不写 Excel、不写数据库。
- `import-sources --confirmed` 写入源数据 import 批次。
- `calculate --confirmed --sync-db` 写入计算结果 run 批次。
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import re
import sys
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import yaml
except ImportError as exc:  # pragma: no cover - exercised in missing env
    raise SystemExit("缺少 PyYAML，请先安装 scripts/requirements.txt 中的依赖") from exc

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - exercised in missing env
    raise SystemExit("缺少 openpyxl，请先安装 scripts/requirements.txt 中的依赖") from exc


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
PROJECT_ROOT = next(
    (
        parent
        for parent in SCRIPT_DIR.parents
        if (parent / "config" / "fog_config.yaml.example").exists()
        and (parent / ".workbuddy").is_dir()
    ),
    Path.cwd(),
)


def _find_skills_dir() -> Path:
    for parent in Path(__file__).resolve().parents:
        if (parent / "lxx_share").is_dir():
            return parent
    return SKILL_DIR.parent


SKILLS_DIR = _find_skills_dir()
for _path in [SKILLS_DIR, SKILLS_DIR / "lx_shujuku" / "scripts"]:
    if str(_path) not in sys.path:
        sys.path.insert(0, str(_path))


OLD_TOTAL_HEADERS = [
    "月份",
    "主体&品牌",
    "运营主体",
    "品牌",
    "对接人",
    "总GMV-财务报表",
    "非共补免佣GMV",
    "月度固定金额kpi",
    "月度固定共建金额占比",
    "月度运营活动金额",
    "月度运营活动占比",
    "C补共建金额",
    "C补共建金额占比",
    "新开城金额",
    "新开城金额占比",
    "节假日共建金额",
    "过程指标完成度",
    "总共建金额",
    "点位=总共建金额/非共补GMV",
]

AUDIT_HEADERS = [
    "返佣基数",
    "规模返佣返点",
    "过程考核返点",
    "T返点",
    "月度完单考核返点",
    "额外返佣返点",
    "最终返点",
    "返佣金额",
    "剔除前返佣基数",
    "开城周期剔除GMV",
    "新开城金额",
    "新开城金额占比",
    "总共建金额",
    "总点位",
    "月总完单数",
    "完单考核日均完单",
    "命中档位",
    "考核方式",
    "过程返佣系数",
    "红线状态",
    "数据状态",
    "异常说明",
]


@dataclass(frozen=True)
class ScopeRow:
    operator: str
    brand: str
    city: str
    contact_person: str


@dataclass
class BillingAgg:
    gross_all: float = 0.0
    gross_receivable: float = 0.0
    spring_service_fee: float = 0.0
    rebate_base_before_open_city: float = 0.0
    open_city_excluded_base: float = 0.0
    rebate_base: float = 0.0
    filtered_row_count: int = 0


@dataclass
class BillingImportAgg:
    gross_receivable: float = 0.0
    spring_service_fee: float = 0.0
    row_count: int = 0


@dataclass
class BillingRawRow:
    sheet_name: str
    row_no: int
    partition_date: date | None
    brand: str
    city: str
    tr_type: str
    channel: str
    gross_receivable: float
    spring_service_fee: float
    raw_payload: dict[str, Any]


@dataclass
class OpenCityRow:
    sheet_name: str
    brand: str
    city: str
    settlement_type: str
    settlement_unit: str
    settlement_item: str
    open_date: str
    incentive_period: str
    rate: float
    settlement_period: str
    rebate_basis: float
    reward_amount: float
    remark1: str
    remark2: str
    raw_payload: dict[str, Any]


@dataclass
class TravelTarget:
    operator: str
    brand: str
    first_target: float | None
    second_target: float | None
    third_target: float | None
    first_miss_label: str
    first_hit_label: str
    second_hit_label: str
    third_hit_label: str
    assessment_mode: str
    process_coefficient: float
    extra_rate: float
    raw_payload: dict[str, Any]


@dataclass
class YouxingTarget:
    merchant: str
    cities: list[str]
    tiers: list[tuple[float, float, str]]
    raw_payload: dict[str, Any]


@dataclass
class ProcessDetail:
    operator: str
    brand: str
    metric_key: str
    metric_name: str
    metric_value: str
    threshold_text: str
    passed: bool
    base_rate: float
    coefficient: float
    final_rate: float
    source_sheet: str
    reason: str = ""
    completion_base_rate: float = 0.0
    completion_earned_rate: float = 0.0


@dataclass
class ResultRow:
    month_label: str
    operator: str
    brand: str
    contact_person: str
    gross_all: float
    rebate_base_before_open_city: float
    open_city_excluded_base: float
    rebate_base: float
    scale_rate: float
    process_rate: float
    monthly_order_rate: float
    extra_rate: float
    final_rate: float
    completed_orders: float
    completed_orders_for_target: float
    tier_name: str
    assessment_mode: str
    process_coefficient: float
    redline_status: str
    status: str
    reason: str
    process_completion_rate: float = 0.0
    new_city_amount: float = 0.0
    process_details: list[ProcessDetail] = field(default_factory=list)

    @property
    def t_rate(self) -> float:
        return self.scale_rate + self.process_rate

    @property
    def rebate_amount(self) -> float:
        return self.rebate_base * self.final_rate

    @property
    def new_city_rate(self) -> float:
        if self.rebate_base == 0:
            return 0.0
        return self.new_city_amount / self.rebate_base

    @property
    def total_rebate_amount(self) -> float:
        return self.rebate_amount + self.new_city_amount

    @property
    def total_rate(self) -> float:
        if self.rebate_base == 0:
            return 0.0
        return self.total_rebate_amount / self.rebate_base


@dataclass
class ValidationItem:
    category: str
    operator: str
    brand: str
    city: str
    message: str


@dataclass
class CalculationContext:
    run_id: str
    import_id: str
    data_source: str
    month: str
    month_label: str
    work_dir: Path
    rules_path: Path
    rules_hash: str
    rules_text: str
    import_rules_hash: str
    rules: dict[str, Any]
    source_files: dict[str, Path]
    source_file_hashes: dict[str, str]
    contacts: list[str]
    exclude_operators: list[str]
    scope_rows: list[ScopeRow]
    billing_by_brand_city: dict[tuple[str, str], BillingAgg]
    billing_by_pair: dict[tuple[str, str], BillingAgg]
    completed_by_brand_city: dict[tuple[str, str], float]
    process_rows: dict[str, list[dict[str, Any]]]
    open_city_rows: list[OpenCityRow]
    open_city_by_pair: dict[tuple[str, str], float]
    travel_targets: dict[tuple[str, str], TravelTarget]
    youxing_targets: dict[str, YouxingTarget]
    target_raw_rows: list[tuple[str, str, dict[str, Any]]]
    source_sheet_counts: list[tuple[str, Path, str, int]]
    validations: list[ValidationItem]
    results: list[ResultRow]


@dataclass
class ImportContext:
    import_id: str
    import_type: str
    month: str
    month_label: str
    work_dir: Path
    rules_path: Path
    rules_hash: str
    rules_text: str
    rules: dict[str, Any]
    source_files: dict[str, Path]
    source_file_hashes: dict[str, str]
    billing_import_agg: dict[tuple[str, str, str, str, str, str], BillingImportAgg]
    process_rows: dict[str, list[dict[str, Any]]]
    open_city_rows: list[OpenCityRow]
    target_raw_rows: list[tuple[str, str, dict[str, Any]]]
    source_sheet_counts: list[tuple[str, Path, str, int]]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\ufeff", "").strip()


def norm_header(value: Any) -> str:
    return clean(value)


def safe_float(value: Any, default: float | None = 0.0) -> float | None:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value)
    if not text or text.startswith("#"):
        return default
    text = text.replace(",", "").replace("％", "%")
    try:
        if text.endswith("%"):
            return float(text[:-1]) / 100
        return float(text)
    except ValueError:
        return default


def must_float(value: Any, label: str) -> float:
    result = safe_float(value, None)
    if result is None:
        raise RuntimeError(f"无法解析数字: {label}={value!r}")
    return result


def parse_contacts(value: str | None, defaults: list[str]) -> list[str]:
    if value:
        parts = re.split(r"[,，]", value)
        return [p.strip() for p in parts if p.strip()]
    return [str(item).strip() for item in defaults if str(item).strip()]


def parse_month(month: str) -> tuple[int, int]:
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month)
    if not match:
        raise RuntimeError(f"月份格式必须是 YYYY-MM: {month}")
    year = int(match.group(1))
    month_num = int(match.group(2))
    if month_num < 1 or month_num > 12:
        raise RuntimeError(f"月份不合法: {month}")
    return year, month_num


def month_label(month: str) -> str:
    year, month_num = parse_month(month)
    return f"{year}年{month_num}月"


def previous_month(month: str) -> str:
    year, month_num = parse_month(month)
    if month_num == 1:
        return f"{year - 1}-12"
    return f"{year}-{month_num - 1:02d}"


def days_in_month(month: str) -> int:
    year, month_num = parse_month(month)
    return calendar.monthrange(year, month_num)[1]


def completed_orders_for_target(completed_orders: float, month: str, target_rules: dict[str, Any]) -> float:
    basis = clean(target_rules.get("completed_target_basis", "monthly_total"))
    if basis == "daily_average":
        return completed_orders / days_in_month(month)
    if basis == "monthly_total":
        return completed_orders
    raise RuntimeError(f"未知完单目标口径: targets.completed_target_basis={basis}")


def value_month_key(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    text = clean(value)
    if not text:
        return ""
    match = re.search(r"(\d{4})[-/年](\d{1,2})", text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    return text[:7]


def parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    match = re.search(r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})", text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def format_date_key(value: date | None) -> str:
    return value.isoformat() if value else ""


def month_dir_from_month(month: str) -> Path:
    return PROJECT_ROOT / "workspace" / "13月度返点计算" / month_label(month)


def hash_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def hash_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def read_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    if not isinstance(data, dict):
        raise RuntimeError(f"rules.yaml 顶层必须是对象: {path}")
    return data


def write_yaml(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        yaml.safe_dump(data, fh, allow_unicode=True, sort_keys=False)


def load_template(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def init_rules(args: argparse.Namespace) -> int:
    month = args.month
    label = month_label(month)
    work_dir = Path(args.work_dir).expanduser() if args.work_dir else month_dir_from_month(month)
    if not work_dir.is_absolute():
        work_dir = PROJECT_ROOT / work_dir
    rules_dir = work_dir / "规则"
    rules_dir.mkdir(parents=True, exist_ok=True)

    contacts = parse_contacts(args.contacts, ["雷维亮"])
    contact = contacts[0] if contacts else "雷维亮"

    md_path = rules_dir / f"{label}月度返点规则.md"
    yaml_path = rules_dir / "rules.yaml"

    md_text = load_template(SKILL_DIR / "references" / "monthly_rules_template.md").format(
        month=month,
        month_label=label,
        contacts="、".join(contacts),
    )
    yaml_text = load_template(SKILL_DIR / "references" / "rules_yaml_template.yaml").format(
        month=month,
        contact=contact,
    )
    yaml_data = yaml.safe_load(yaml_text)
    yaml_data["contacts"] = contacts

    for path, content, is_yaml in [
        (md_path, md_text, False),
        (yaml_path, yaml_data, True),
    ]:
        if path.exists() and not args.overwrite:
            print(f"[skip] 已存在: {path}")
            continue
        if args.dry_run:
            print(f"[dry-run] 将写入: {path}")
            continue
        if is_yaml:
            write_yaml(path, content)
        else:
            path.write_text(content, encoding="utf-8")
        print(f"[ok] 已写入: {path}")

    for subdir in ["源数据", "输出"]:
        path = work_dir / subdir
        if args.dry_run:
            print(f"[dry-run] ensure dir: {path}")
        else:
            path.mkdir(parents=True, exist_ok=True)
    return 0


def resolve_source_files(work_dir: Path, rules: dict[str, Any]) -> dict[str, Path]:
    source_dir = work_dir / "源数据"
    if not source_dir.exists():
        raise RuntimeError(f"源数据目录不存在: {source_dir}")
    result: dict[str, Path] = {}
    source_rules = rules.get("source_files") or {}
    required_roles = ["billing", "process", "target"]
    roles = required_roles + [role for role in source_rules if role not in required_roles]
    for role in roles:
        pattern = ((source_rules.get(role) or {}).get("pattern") or "").strip()
        if not pattern and role in required_roles:
            raise RuntimeError(f"rules.yaml 缺少 source_files.{role}.pattern")
        if not pattern:
            continue
        matches = sorted(p for p in source_dir.glob(pattern) if p.is_file() and not p.name.startswith(("~$", ".~")))
        if len(matches) != 1:
            names = ", ".join(p.name for p in matches) or "无"
            raise RuntimeError(f"{role} 源文件必须唯一，pattern={pattern}，实际={names}")
        result[role] = matches[0]
    return result


def resolve_process_file(work_dir: Path, pattern: str = "*LX品牌过程考核数据*.xlsx") -> Path:
    source_dir = work_dir / "源数据"
    if not source_dir.exists():
        raise RuntimeError(f"源数据目录不存在: {source_dir}")
    matches = sorted(p for p in source_dir.glob(pattern) if p.is_file() and not p.name.startswith(("~$", ".~")))
    if len(matches) != 1:
        names = ", ".join(p.name for p in matches) or "无"
        raise RuntimeError(f"process 源文件必须唯一，pattern={pattern}，实际={names}")
    return matches[0]


def workbook_sheet_counts(role: str, path: Path) -> list[tuple[str, Path, str, int]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return [(role, path, ws.title, max(ws.max_row - 1, 0)) for ws in wb.worksheets]
    finally:
        wb.close()


def require_columns(headers: list[str], required: list[str], sheet_name: str) -> dict[str, int]:
    index = {header: idx for idx, header in enumerate(headers) if header}
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"sheet `{sheet_name}` 缺少表头: {', '.join(missing)}")
    return index


def iter_sheet_dicts(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"workbook 缺少 sheet: {path} / {sheet_name}")
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [norm_header(v) for v in next(rows)]
        except StopIteration:
            return []
        result = []
        for row in rows:
            if not any(cell not in (None, "") for cell in row):
                continue
            result.append({headers[idx]: row[idx] for idx in range(min(len(headers), len(row))) if headers[idx]})
        return result
    finally:
        wb.close()


def sql_quote(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "\\'") + "'"


def load_operator_scope(contacts: list[str], exclude_operators: list[str]) -> tuple[list[ScopeRow], list[ValidationItem]]:
    if not contacts:
        raise RuntimeError("必须指定至少一个对接人")
    try:
        from lx_shujuku import create_client
    except ImportError as exc:
        raise RuntimeError("无法 import lx_shujuku，请确认 Skill 目录完整") from exc

    client = create_client()
    contact_sql = ", ".join(sql_quote(item) for item in contacts)
    conditions = [f"contact_person IN ({contact_sql})"]
    if exclude_operators:
        excluded = ", ".join(sql_quote(item) for item in exclude_operators)
        conditions.append(f"operator_entity NOT IN ({excluded})")
    where = " AND ".join(conditions)
    count_rows = client.execute(f"SELECT COUNT(*) AS cnt FROM operator_brand WHERE {where} LIMIT 1")
    expected_count = int(count_rows[0].get("cnt") or 0) if count_rows else 0
    limit = min(max(expected_count, 1), getattr(client, "max_limit", 1000))
    raw_rows = client.execute(
        "SELECT operator_entity, brand_name, city_name, contact_person "
        f"FROM operator_brand WHERE {where} "
        f"ORDER BY operator_entity, brand_name, city_name LIMIT {limit}"
    )
    validations: list[ValidationItem] = []
    if expected_count > len(raw_rows):
        validations.append(
            ValidationItem(
                "公司库范围截断",
                "",
                "",
                "",
                f"operator_brand 命中 {expected_count} 行，但当前查询限制只返回 {len(raw_rows)} 行；请调大 lx_shujuku.query.max_limit",
            )
        )
    scope = []
    for row in raw_rows:
        operator = clean(row.get("operator_entity") or row.get("运营主体"))
        brand = clean(row.get("brand_name") or row.get("品牌"))
        city = clean(row.get("city_name") or row.get("城市"))
        contact_person = clean(row.get("contact_person") or row.get("对接人"))
        if not operator or not brand or not city:
            validations.append(ValidationItem("公司库范围异常", operator, brand, city, "缺少运营主体/品牌/城市"))
            continue
        scope.append(ScopeRow(operator, brand, city, contact_person))
    if not scope:
        raise RuntimeError(f"公司库 operator_brand 未查到对接人范围: {', '.join(contacts)}")
    return scope, validations


def build_scope_indexes(scope_rows: list[ScopeRow]) -> tuple[
    dict[tuple[str, str], list[ScopeRow]],
    dict[str, list[ScopeRow]],
    dict[tuple[str, str], ScopeRow],
    list[ValidationItem],
]:
    by_pair: dict[tuple[str, str], list[ScopeRow]] = defaultdict(list)
    by_operator: dict[str, list[ScopeRow]] = defaultdict(list)
    by_brand_city: dict[tuple[str, str], list[ScopeRow]] = defaultdict(list)
    validations: list[ValidationItem] = []
    for row in scope_rows:
        by_pair[(row.operator, row.brand)].append(row)
        by_operator[row.operator].append(row)
        by_brand_city[(row.brand, row.city)].append(row)
    owner: dict[tuple[str, str], ScopeRow] = {}
    for key, rows in by_brand_city.items():
        owner[key] = sorted(rows, key=lambda item: (item.operator, item.brand, item.city))[0]
        if len(rows) > 1:
            validations.append(
                ValidationItem(
                    "公司库品牌城市多归属",
                    "",
                    key[0],
                    key[1],
                    "；".join(f"{r.operator}-{r.brand}" for r in rows),
                )
            )
    return dict(by_pair), dict(by_operator), owner, validations


def apply_scope_overrides(
    scope_rows: list[ScopeRow],
    rules: dict[str, Any],
    validations: list[ValidationItem],
) -> list[ScopeRow]:
    scope_rules = rules.get("scope") or {}
    result = list(scope_rows)

    for item in scope_rules.get("reassign_rows") or []:
        from_operator = clean(item.get("from_operator_entity"))
        from_brand = clean(item.get("from_brand_name") or item.get("brand_name"))
        from_city = clean(item.get("from_city_name") or item.get("city_name"))
        to_operator = clean(item.get("operator_entity"))
        to_brand = clean(item.get("brand_name") or from_brand)
        contact_person = clean(item.get("contact_person"))
        changed = False
        next_rows: list[ScopeRow] = []
        for row in result:
            matched = (
                row.operator == from_operator
                and row.brand == from_brand
                and row.city == from_city
            )
            if matched:
                next_rows.append(
                    ScopeRow(
                        operator=to_operator or row.operator,
                        brand=to_brand or row.brand,
                        city=row.city,
                        contact_person=contact_person or row.contact_person,
                    )
                )
                changed = True
            else:
                next_rows.append(row)
        result = next_rows
        if changed:
            validations.append(
                ValidationItem(
                    "规则覆盖-归属改写",
                    to_operator,
                    to_brand,
                    from_city,
                    f"{from_operator}-{from_brand}-{from_city} 改按 {to_operator}-{to_brand}-{from_city} 计算",
                )
            )
        else:
            validations.append(
                ValidationItem(
                    "规则覆盖-归属改写未命中",
                    to_operator,
                    to_brand,
                    from_city,
                    f"未在公司库范围中找到 {from_operator}-{from_brand}-{from_city}",
                )
            )

    seen = {(row.operator, row.brand, row.city) for row in result}
    for item in scope_rules.get("include_rows") or []:
        operator = clean(item.get("operator_entity"))
        brand = clean(item.get("brand_name"))
        city = clean(item.get("city_name"))
        contact_person = clean(item.get("contact_person"))
        if not operator or not brand or not city:
            validations.append(ValidationItem("规则覆盖-额外范围异常", operator, brand, city, "缺少运营主体/品牌/城市"))
            continue
        key = (operator, brand, city)
        if key in seen:
            continue
        result.append(ScopeRow(operator, brand, city, contact_person))
        seen.add(key)
        validations.append(ValidationItem("规则覆盖-额外纳入", operator, brand, city, "按当月规则额外纳入计算范围"))

    return result


def read_billing_import_agg(path: Path, rules: dict[str, Any]) -> dict[tuple[str, str, str, str, str, str], BillingImportAgg]:
    billing_rules = rules.get("billing_base") or {}
    required = [
        billing_rules.get("date_column", "分区日期(天)"),
        billing_rules.get("brand_column", "品牌"),
        billing_rules.get("city_column", "城市名称"),
        billing_rules.get("gross_receivable_column", "求和项:乘客应收_金额"),
        billing_rules.get("spring_service_fee_column", "求和项:乘客应收_春节服务费"),
        billing_rules.get("tr_type_column", "tr类型"),
        billing_rules.get("channel_column", "流量渠道"),
    ]
    result: dict[tuple[str, str, str, str, str, str], BillingImportAgg] = defaultdict(BillingImportAgg)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                headers = [norm_header(v) for v in next(rows)]
            except StopIteration:
                continue
            col = require_columns(headers, required, ws.title)
            date_col = col[billing_rules.get("date_column", "分区日期(天)")]
            brand_col = col[billing_rules.get("brand_column", "品牌")]
            city_col = col[billing_rules.get("city_column", "城市名称")]
            gross_col = col[billing_rules.get("gross_receivable_column", "求和项:乘客应收_金额")]
            spring_col = col[billing_rules.get("spring_service_fee_column", "求和项:乘客应收_春节服务费")]
            tr_col = col[billing_rules.get("tr_type_column", "tr类型")]
            channel_col = col[billing_rules.get("channel_column", "流量渠道")]
            for row in rows:
                partition_date = parse_date_value(row[date_col] if date_col < len(row) else None)
                brand = clean(row[brand_col] if brand_col < len(row) else "")
                city = clean(row[city_col] if city_col < len(row) else "")
                if not brand or not city:
                    continue
                gross = safe_float(row[gross_col] if gross_col < len(row) else 0, 0.0) or 0.0
                tr_type = clean(row[tr_col] if tr_col < len(row) else "")
                channel = clean(row[channel_col] if channel_col < len(row) else "")
                spring = safe_float(row[spring_col] if spring_col < len(row) else 0, 0.0) or 0.0
                agg = result[(ws.title, format_date_key(partition_date), brand, city, tr_type, channel)]
                agg.gross_receivable += gross
                agg.spring_service_fee += spring
                agg.row_count += 1
    finally:
        wb.close()
    return dict(result)


def iter_billing_raw_rows(path: Path, rules: dict[str, Any]):
    billing_rules = rules.get("billing_base") or {}
    required = [
        billing_rules.get("date_column", "分区日期(天)"),
        billing_rules.get("brand_column", "品牌"),
        billing_rules.get("city_column", "城市名称"),
        billing_rules.get("gross_receivable_column", "求和项:乘客应收_金额"),
        billing_rules.get("spring_service_fee_column", "求和项:乘客应收_春节服务费"),
        billing_rules.get("tr_type_column", "tr类型"),
        billing_rules.get("channel_column", "流量渠道"),
    ]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                headers = [norm_header(v) for v in next(rows)]
            except StopIteration:
                continue
            col = require_columns(headers, required, ws.title)
            date_col = col[billing_rules.get("date_column", "分区日期(天)")]
            brand_col = col[billing_rules.get("brand_column", "品牌")]
            city_col = col[billing_rules.get("city_column", "城市名称")]
            gross_col = col[billing_rules.get("gross_receivable_column", "求和项:乘客应收_金额")]
            spring_col = col[billing_rules.get("spring_service_fee_column", "求和项:乘客应收_春节服务费")]
            tr_col = col[billing_rules.get("tr_type_column", "tr类型")]
            channel_col = col[billing_rules.get("channel_column", "流量渠道")]
            for row_no, row in enumerate(rows, start=2):
                raw_payload = {
                    headers[idx]: json_safe(row[idx])
                    for idx in range(min(len(headers), len(row)))
                    if headers[idx]
                }
                brand = clean(row[brand_col] if brand_col < len(row) else "")
                city = clean(row[city_col] if city_col < len(row) else "")
                if not brand and not city and not raw_payload:
                    continue
                yield BillingRawRow(
                    sheet_name=ws.title,
                    row_no=row_no,
                    partition_date=parse_date_value(row[date_col] if date_col < len(row) else None),
                    brand=brand,
                    city=city,
                    tr_type=clean(row[tr_col] if tr_col < len(row) else ""),
                    channel=clean(row[channel_col] if channel_col < len(row) else ""),
                    gross_receivable=safe_float(row[gross_col] if gross_col < len(row) else 0, 0.0) or 0.0,
                    spring_service_fee=safe_float(row[spring_col] if spring_col < len(row) else 0, 0.0) or 0.0,
                    raw_payload=raw_payload,
                )
    finally:
        wb.close()


def build_billing_from_import_agg(
    import_agg: dict[tuple[str, str, str, str, str, str], BillingImportAgg],
    rules: dict[str, Any],
    brand_city_owner: dict[tuple[str, str], ScopeRow],
    open_city_periods: dict[tuple[str, str], list[tuple[date, date]]] | None = None,
) -> tuple[dict[tuple[str, str], BillingAgg], dict[tuple[str, str], BillingAgg]]:
    billing_rules = rules.get("billing_base") or {}
    exclude_tr = set(str(item) for item in billing_rules.get("exclude_tr_types") or [])
    exclude_channels = set(str(item) for item in billing_rules.get("exclude_channels") or [])
    by_brand_city: dict[tuple[str, str], BillingAgg] = defaultdict(BillingAgg)
    by_pair: dict[tuple[str, str], BillingAgg] = defaultdict(BillingAgg)
    periods = open_city_periods or {}

    for (_sheet, partition_date_text, brand, city, tr_type, channel), raw_agg in import_agg.items():
        owner = brand_city_owner.get((brand, city))
        if owner is None:
            continue
        partition_date = parse_date_value(partition_date_text)
        pair_agg = by_pair[(owner.operator, owner.brand)]
        city_agg = by_brand_city[(brand, city)]
        for agg in [pair_agg, city_agg]:
            agg.gross_all += raw_agg.gross_receivable
        if tr_type in exclude_tr or channel in exclude_channels:
            continue
        raw_base = raw_agg.gross_receivable - raw_agg.spring_service_fee
        is_open_city_period = date_in_ranges(partition_date, periods.get((brand, city), []))
        for agg in [pair_agg, city_agg]:
            agg.gross_receivable += raw_agg.gross_receivable
            agg.spring_service_fee += raw_agg.spring_service_fee
            agg.rebate_base_before_open_city += raw_base
            if is_open_city_period:
                agg.open_city_excluded_base += raw_base
            else:
                agg.rebate_base += raw_base
            agg.filtered_row_count += raw_agg.row_count
    return dict(by_brand_city), dict(by_pair)


def read_billing(
    path: Path,
    rules: dict[str, Any],
    brand_city_owner: dict[tuple[str, str], ScopeRow],
) -> tuple[dict[tuple[str, str], BillingAgg], dict[tuple[str, str], BillingAgg]]:
    return build_billing_from_import_agg(read_billing_import_agg(path, rules), rules, brand_city_owner)


def read_completed_orders(path: Path) -> dict[tuple[str, str], float]:
    rows = iter_sheet_dicts(path, "完单")
    return completed_orders_from_rows(rows)


def completed_orders_from_rows(rows: list[dict[str, Any]]) -> dict[tuple[str, str], float]:
    result: dict[tuple[str, str], float] = defaultdict(float)
    for row in rows:
        brand = clean(row.get("company_name"))
        city = clean(row.get("city_name"))
        if not brand or not city:
            continue
        result[(brand, city)] += safe_float(row.get("完单数"), 0.0) or 0.0
    return dict(result)


def completed_orders_from_process_rows(process_rows: dict[str, list[dict[str, Any]]]) -> dict[tuple[str, str], float]:
    return completed_orders_from_rows(process_rows.get("完单") or [])


def read_process_rows(path: Path) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheetnames = wb.sheetnames
    finally:
        wb.close()
    return {name: iter_sheet_dicts(path, name) for name in sheetnames}


def split_cities(value: Any) -> list[str]:
    text = clean(value)
    if not text:
        return []
    return [item.strip() for item in re.split(r"[,，、;；\s]+", text) if item.strip()]


def read_targets(path: Path, rules: dict[str, Any]) -> tuple[
    dict[tuple[str, str], TravelTarget],
    dict[str, YouxingTarget],
    list[tuple[str, str, dict[str, Any]]],
]:
    target_rules = rules.get("targets") or {}
    travel_sheet = target_rules.get("travel_sheet", "出行")
    youxing_sheet = target_rules.get("youxing_sheet", "优行")
    raw_rows: list[tuple[str, str, dict[str, Any]]] = []

    for row in iter_sheet_dicts(path, travel_sheet):
        operator = clean(row.get("运营主体"))
        brand = clean(row.get("品牌名称"))
        if not operator or not brand:
            continue
        raw_rows.append(("travel", travel_sheet, json_safe(row)))

    for row in iter_sheet_dicts(path, youxing_sheet):
        merchant = clean(row.get("商家"))
        if not merchant:
            continue
        raw_rows.append(("youxing", youxing_sheet, json_safe(row)))
    travel_targets, youxing_targets = parse_targets_from_raw_rows(raw_rows)
    return travel_targets, youxing_targets, raw_rows


def parse_targets_from_raw_rows(
    raw_rows: list[tuple[str, str, dict[str, Any]]],
) -> tuple[dict[tuple[str, str], TravelTarget], dict[str, YouxingTarget]]:
    travel_targets: dict[tuple[str, str], TravelTarget] = {}
    youxing_targets: dict[str, YouxingTarget] = {}
    for target_type, _sheet_name, row in raw_rows:
        if target_type == "travel":
            operator = clean(row.get("运营主体"))
            brand = clean(row.get("品牌名称"))
            if not operator or not brand:
                continue
            target = TravelTarget(
                operator=operator,
                brand=brand,
                first_target=safe_float(row.get("一档完单考核"), None),
                second_target=safe_float(row.get("二档完单考核"), None),
                third_target=safe_float(row.get("三档完单考核"), None),
                first_miss_label=clean(row.get("一档完单未达成")),
                first_hit_label=clean(row.get("一档完单达成")),
                second_hit_label=clean(row.get("二档完单达成")),
                third_hit_label=clean(row.get("三档完单达成")),
                assessment_mode=clean(row.get("运营主体的品牌的考核方式")),
                process_coefficient=safe_float(row.get("过程返佣系数"), 1.0) or 1.0,
                extra_rate=safe_float(row.get("额外返佣"), 0.0) or 0.0,
                raw_payload=json_safe(row),
            )
            travel_targets[(operator, brand)] = target
        elif target_type == "youxing":
            merchant = clean(row.get("商家"))
            if not merchant:
                continue
            tiers = []
            for idx in [1, 2, 3]:
                threshold = safe_float(row.get(f"阶梯{idx}"), None)
                rate = safe_float(row.get(f"激励系数{idx}"), None)
                if threshold is not None and rate is not None:
                    tiers.append((threshold, rate, f"优行{idx}档"))
            youxing_targets[merchant] = YouxingTarget(
                merchant=merchant,
                cities=split_cities(row.get("考核具体城市")),
                tiers=tiers,
                raw_payload=json_safe(row),
            )
    return travel_targets, youxing_targets


def normalize_open_city_row(sheet_name: str, row: dict[str, Any], rules: dict[str, Any]) -> OpenCityRow:
    reward_rules = rules.get("open_city_reward") or {}
    brand_col = clean(reward_rules.get("brand_column", "品牌"))
    city_col = clean(reward_rules.get("city_column", "城市"))
    settlement_type_col = clean(reward_rules.get("settlement_type_column", "结算类型"))
    settlement_unit_col = clean(reward_rules.get("settlement_unit_column", "结算单位"))
    settlement_item_col = clean(reward_rules.get("settlement_item_column", "结算项目"))
    open_date_col = clean(reward_rules.get("open_date_column", "开城时间"))
    incentive_period_col = clean(reward_rules.get("incentive_period_column", "激励期间"))
    rate_col = clean(reward_rules.get("rate_column", "比例"))
    settlement_period_col = clean(reward_rules.get("settlement_period_column", "结算周期"))
    rebate_basis_col = clean(reward_rules.get("rebate_basis_column", "返利基数"))
    reward_amount_col = clean(reward_rules.get("reward_amount_column", "返利金额"))
    remark1_col = clean(reward_rules.get("remark1_column", "备注1"))
    remark2_col = clean(reward_rules.get("remark2_column", "备注2"))
    return OpenCityRow(
        sheet_name=sheet_name,
        brand=clean(row.get(brand_col)),
        city=clean(row.get(city_col)),
        settlement_type=clean(row.get(settlement_type_col)),
        settlement_unit=clean(row.get(settlement_unit_col)),
        settlement_item=clean(row.get(settlement_item_col)),
        open_date=clean(row.get(open_date_col)),
        incentive_period=clean(row.get(incentive_period_col)),
        rate=safe_float(row.get(rate_col), 0.0) or 0.0,
        settlement_period=clean(row.get(settlement_period_col)),
        rebate_basis=safe_float(row.get(rebate_basis_col), 0.0) or 0.0,
        reward_amount=safe_float(row.get(reward_amount_col), 0.0) or 0.0,
        remark1=clean(row.get(remark1_col)),
        remark2=clean(row.get(remark2_col)),
        raw_payload=json_safe(row),
    )


def read_open_city_rows(path: Path, rules: dict[str, Any]) -> list[OpenCityRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheetnames = wb.sheetnames
    finally:
        wb.close()
    rows: list[OpenCityRow] = []
    for sheet_name in sheetnames:
        for row in iter_sheet_dicts(path, sheet_name):
            normalized = normalize_open_city_row(sheet_name, row, rules)
            if normalized.brand or normalized.city or normalized.reward_amount:
                rows.append(normalized)
    return rows


def open_city_row_payload(row: OpenCityRow) -> dict[str, Any]:
    return {
        "sheet_name": row.sheet_name,
        "brand_name": row.brand,
        "city_name": row.city,
        "settlement_type": row.settlement_type,
        "settlement_unit": row.settlement_unit,
        "settlement_item": row.settlement_item,
        "open_date": row.open_date,
        "incentive_period": row.incentive_period,
        "rate": row.rate,
        "settlement_period": row.settlement_period,
        "rebate_basis": row.rebate_basis,
        "reward_amount": row.reward_amount,
        "remark1": row.remark1,
        "remark2": row.remark2,
        "raw_payload": row.raw_payload,
    }


def build_open_city_by_pair(
    rows: list[OpenCityRow],
    rules: dict[str, Any],
    brand_city_owner: dict[tuple[str, str], ScopeRow],
    validations: list[ValidationItem],
) -> dict[tuple[str, str], float]:
    reward_rules = rules.get("open_city_reward") or {}
    settlement_type_include = {clean(item) for item in reward_rules.get("settlement_type_include") or [] if clean(item)}
    settlement_item_include = {clean(item) for item in reward_rules.get("settlement_item_include") or [] if clean(item)}
    result: dict[tuple[str, str], float] = defaultdict(float)
    for row in rows:
        if settlement_type_include and row.settlement_type not in settlement_type_include:
            continue
        if settlement_item_include and row.settlement_item not in settlement_item_include:
            continue
        if not row.brand or not row.city:
            validations.append(ValidationItem("开城奖励缺少品牌城市", "", row.brand, row.city, f"sheet={row.sheet_name}"))
            continue
        owner = brand_city_owner.get((row.brand, row.city))
        if owner is None:
            validations.append(
                ValidationItem(
                    "开城奖励无法匹配公司库",
                    "",
                    row.brand,
                    row.city,
                    f"sheet={row.sheet_name}，返利金额={row.reward_amount:.2f}，未进入主表",
                )
            )
            continue
        result[(owner.operator, owner.brand)] += row.reward_amount
    return dict(result)


def parse_date_range_text(value: Any) -> tuple[date, date] | None:
    text = clean(value)
    matches = re.findall(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", text)
    if len(matches) < 2:
        return None
    try:
        start = date(int(matches[0][0]), int(matches[0][1]), int(matches[0][2]))
        end = date(int(matches[1][0]), int(matches[1][1]), int(matches[1][2]))
    except ValueError:
        return None
    if end < start:
        start, end = end, start
    return start, end


def merge_date_ranges(ranges: list[tuple[date, date]]) -> list[tuple[date, date]]:
    if not ranges:
        return []
    sorted_ranges = sorted(ranges, key=lambda item: (item[0], item[1]))
    merged = [sorted_ranges[0]]
    for start, end in sorted_ranges[1:]:
        last_start, last_end = merged[-1]
        if start <= last_end:
            merged[-1] = (last_start, max(last_end, end))
        else:
            merged.append((start, end))
    return merged


def date_in_ranges(value: date | None, ranges: list[tuple[date, date]]) -> bool:
    if value is None:
        return False
    return any(start <= value <= end for start, end in ranges)


def build_open_city_periods(
    rows: list[OpenCityRow],
    rules: dict[str, Any],
    brand_city_owner: dict[tuple[str, str], ScopeRow],
    validations: list[ValidationItem],
) -> dict[tuple[str, str], list[tuple[date, date]]]:
    reward_rules = rules.get("open_city_reward") or {}
    if not reward_rules.get("exclude_billing_base", False):
        return {}
    settlement_type_include = {clean(item) for item in reward_rules.get("settlement_type_include") or [] if clean(item)}
    settlement_item_include = {clean(item) for item in reward_rules.get("settlement_item_include") or [] if clean(item)}
    invalid_policy = clean(reward_rules.get("invalid_period_policy", "error")) or "error"
    if invalid_policy not in {"error", "warn", "ignore"}:
        raise RuntimeError(f"open_city_reward.invalid_period_policy 只能是 error/warn/ignore，当前为: {invalid_policy}")
    period_column = clean(reward_rules.get("period_column", reward_rules.get("remark1_column", "备注1")))
    remark2_column = clean(reward_rules.get("remark2_column", "备注2"))
    by_brand_city: dict[tuple[str, str], list[tuple[date, date]]] = defaultdict(list)
    for row in rows:
        if settlement_type_include and row.settlement_type not in settlement_type_include:
            continue
        if settlement_item_include and row.settlement_item not in settlement_item_include:
            continue
        if not row.brand or not row.city:
            continue
        if (row.brand, row.city) not in brand_city_owner:
            continue
        period_text = row.remark2 if period_column == remark2_column else row.remark1
        parsed = parse_date_range_text(period_text)
        if parsed is None:
            message = f"开城奖励{period_column}无法解析日期区间: {period_text}"
            if invalid_policy == "error":
                raise RuntimeError(f"{row.brand}-{row.city} {message}")
            if invalid_policy == "warn":
                validations.append(ValidationItem("开城周期无法解析", "", row.brand, row.city, message))
            continue
        by_brand_city[(row.brand, row.city)].append(parsed)
    return {key: merge_date_ranges(value) for key, value in by_brand_city.items()}


def scale_rate_for_base(base: float, tiers: list[dict[str, Any]]) -> tuple[float, str]:
    for tier in tiers:
        min_value = must_float(tier.get("min"), "scale_rebate.tiers.min")
        max_raw = tier.get("max")
        max_value = safe_float(max_raw, None) if max_raw is not None else None
        if base >= min_value and (max_value is None or base < max_value):
            return must_float(tier.get("rate"), "scale_rebate.tiers.rate"), clean(tier.get("name"))
    return 0.0, "未达规模阶梯"


def scale_basis_from_completed_orders(completed_orders: float, month: str, scale_rules: dict[str, Any]) -> float:
    basis = clean(scale_rules.get("basis", "operator_daily_completed_orders"))
    if basis == "operator_daily_completed_orders":
        return completed_orders / days_in_month(month)
    if basis == "operator_monthly_completed_orders":
        return completed_orders
    raise RuntimeError(f"未知规模返佣口径: scale_rebate.basis={basis}")


def rate_delta_from_label(label: Any) -> float:
    if isinstance(label, (int, float)):
        return float(label)
    text = clean(label).replace("％", "%").replace(" ", "")
    if not text or text in {"-", "T"}:
        return 0.0
    match = re.fullmatch(r"T([+-])(\d+(?:\.\d+)?)%", text)
    if match:
        value = float(match.group(2)) / 100
        return value if match.group(1) == "+" else -value
    match = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)%", text)
    if match:
        return float(match.group(1)) / 100
    return 0.0


def choose_monthly_delta(completed: float, target: TravelTarget) -> tuple[float, str]:
    thresholds = [
        (target.third_target, target.third_hit_label, "三档"),
        (target.second_target, target.second_hit_label, "二档"),
        (target.first_target, target.first_hit_label, "一档"),
    ]
    for threshold, label, tier_name in thresholds:
        if threshold is not None and completed >= threshold:
            return rate_delta_from_label(label), tier_name
    return rate_delta_from_label(target.first_miss_label), "一档未达成"


def sum_completed(scope_rows: list[ScopeRow], completed_by_brand_city: dict[tuple[str, str], float]) -> float:
    return sum(completed_by_brand_city.get((row.brand, row.city), 0.0) for row in scope_rows)


def is_youxing_brand(brand: str, youxing_targets: dict[str, YouxingTarget]) -> bool:
    brand = clean(brand)
    return brand in youxing_targets or brand.endswith("优行")


def exclude_youxing_scope(scope_rows: list[ScopeRow], youxing_targets: dict[str, YouxingTarget]) -> list[ScopeRow]:
    return [row for row in scope_rows if not is_youxing_brand(row.brand, youxing_targets)]


def completed_for_threshold_basis(
    basis: str,
    completed: float,
    operator_completed: float,
    month: str,
    target_rules: dict[str, Any],
    config_key: str,
) -> float:
    if basis == "operator_total_completed_orders":
        return operator_completed
    if basis == "operator_daily_completed_orders":
        return completed_orders_for_target(operator_completed, month, target_rules)
    if basis == "assessment_scope_completed_orders":
        return completed
    if basis == "assessment_scope_daily_completed_orders":
        return completed_orders_for_target(completed, month, target_rules)
    raise RuntimeError(f"未知完单门槛口径: targets.{config_key}={basis}")


def build_brand_alias_groups(rules: dict[str, Any]) -> dict[str, set[str]]:
    groups: dict[str, set[str]] = {}
    for canonical, aliases in (rules.get("brand_aliases") or {}).items():
        names = {clean(canonical)}
        if isinstance(aliases, list):
            names.update(clean(item) for item in aliases if clean(item))
        names = {item for item in names if item}
        for name in names:
            groups[name] = set(names)
    return groups


def brand_names_for_match(brand: str, alias_groups: dict[str, set[str]]) -> set[str]:
    brand = clean(brand)
    return alias_groups.get(brand, {brand})


def matching_rows(
    rows: list[dict[str, Any]],
    scope_rows: list[ScopeRow],
    brand_column: str,
    city_column: str | None = None,
    brand_alias_groups: dict[str, set[str]] | None = None,
) -> list[dict[str, Any]]:
    alias_groups = brand_alias_groups or {}
    pairs = {
        (brand_name, row.city)
        for row in scope_rows
        for brand_name in brand_names_for_match(row.brand, alias_groups)
    }
    brands = {
        brand_name
        for row in scope_rows
        for brand_name in brand_names_for_match(row.brand, alias_groups)
    }
    result = []
    for row in rows:
        brand = clean(row.get(brand_column))
        if city_column:
            city = clean(row.get(city_column))
            if (brand, city) in pairs:
                result.append(row)
        elif brand in brands:
            result.append(row)
    return result


def evaluate_redline(
    redline: dict[str, Any],
    scope_rows: list[ScopeRow],
    process_rows: dict[str, list[dict[str, Any]]],
    brand_alias_groups: dict[str, set[str]] | None = None,
) -> tuple[str, bool, str]:
    redline_type = clean(redline.get("type"))
    threshold = must_float(redline.get("threshold", 1.0), f"{redline.get('key')}.threshold")
    if redline_type == "constant":
        value = must_float(redline.get("value"), f"{redline.get('key')}.value")
        return format_rate_value(value), value >= threshold, ""
    if redline_type != "sheet_ratio_all":
        raise RuntimeError(f"未知红线类型: {redline_type}")
    sheet = clean(redline.get("sheet"))
    rows = process_rows.get(sheet)
    if rows is None:
        raise RuntimeError(f"过程考核缺少 sheet: {sheet}")
    brand_col = clean(redline.get("brand_column"))
    city_col = clean(redline.get("city_column"))
    value_col = clean(redline.get("value_column"))
    missing_as_pass = bool(redline.get("missing_as_pass"))
    values = []
    missing = []
    invalid = []
    alias_groups = brand_alias_groups or {}
    for scope in scope_rows:
        matched = [
            row
            for row in rows
            if clean(row.get(brand_col)) in brand_names_for_match(scope.brand, alias_groups)
            and clean(row.get(city_col)) == scope.city
        ]
        if not matched:
            missing.append(f"{scope.brand}-{scope.city}")
            continue
        for row in matched:
            value = safe_float(row.get(value_col), None)
            if value is None:
                if missing_as_pass:
                    missing.append(f"{scope.brand}-{scope.city}")
                else:
                    invalid.append(f"{scope.brand}-{scope.city}:非数字")
            else:
                values.append(value)
    if missing_as_pass:
        passed = not invalid and all(value >= threshold for value in values)
    else:
        passed = bool(values) and not missing and not invalid and all(value >= threshold for value in values)

    value_parts = []
    if values:
        value_parts.append(f"min={min(values):.2%}, count={len(values)}")
    if missing and missing_as_pass:
        value_parts.append(f"缺失按100%={len(missing)}")
    value_text = "；".join(value_parts)

    reason_parts = []
    if missing:
        label = "无数据按100%通过" if missing_as_pass else "缺失"
        reason_parts.append(f"{label}: {', '.join(missing[:10])}")
    if invalid:
        reason_parts.append(f"异常: {', '.join(invalid[:10])}")
    reason = "；".join(reason_parts)
    return value_text, passed, reason


def forced_redline_pass(redline: dict[str, Any], pair: tuple[str, str]) -> tuple[bool, str, str]:
    operator, brand = pair
    for item in redline.get("force_pass") or []:
        target_operator = clean(item.get("operator_entity"))
        target_brand = clean(item.get("brand_name"))
        operator_match = not target_operator or target_operator == operator
        brand_match = not target_brand or target_brand == brand
        if operator_match and brand_match:
            value_text = clean(item.get("value_text")) or "100.00%（按规则）"
            reason = clean(item.get("reason")) or "按当月规则强制按100%通过"
            return True, value_text, reason
    return False, "", ""


def evaluate_tsh_growth(
    metric: dict[str, Any],
    month: str,
    scope_rows: list[ScopeRow],
    process_rows: dict[str, list[dict[str, Any]]],
    brand_alias_groups: dict[str, set[str]] | None = None,
) -> tuple[str, bool, float, str]:
    sheet = clean(metric.get("sheet"))
    rows = process_rows.get(sheet)
    if rows is None:
        raise RuntimeError(f"过程考核缺少 sheet: {sheet}")
    date_col = clean(metric.get("date_column"))
    brand_col = clean(metric.get("brand_column"))
    city_col = clean(metric.get("city_column"))
    value_col = clean(metric.get("value_column"))
    prev_month = previous_month(month)
    current_sum = 0.0
    previous_sum = 0.0
    for row in matching_rows(rows, scope_rows, brand_col, city_col, brand_alias_groups):
        row_month = value_month_key(row.get(date_col))
        value = safe_float(row.get(value_col), 0.0) or 0.0
        if row_month == month:
            current_sum += value
        elif row_month == prev_month:
            previous_sum += value
    current_avg = current_sum / days_in_month(month)
    previous_avg = previous_sum / days_in_month(prev_month)
    if previous_avg <= 0:
        return f"本月日均={current_avg:.2f}, 上月日均={previous_avg:.2f}", False, 0.0, "上月日均 TSH 为 0 或缺失"
    growth = (current_avg - previous_avg) / previous_avg
    threshold = must_float(metric.get("threshold"), f"{metric.get('key')}.threshold")
    return f"{growth:.2%}", growth >= threshold, growth, ""


def evaluate_ratio_from_sums(
    metric: dict[str, Any],
    scope_rows: list[ScopeRow],
    process_rows: dict[str, list[dict[str, Any]]],
    by_city: bool,
    brand_alias_groups: dict[str, set[str]] | None = None,
) -> tuple[str, bool, float, str]:
    sheet = clean(metric.get("sheet"))
    rows = process_rows.get(sheet)
    if rows is None:
        raise RuntimeError(f"过程考核缺少 sheet: {sheet}")
    brand_col = clean(metric.get("brand_column"))
    city_col = clean(metric.get("city_column")) if by_city else None
    numerator_col = clean(metric.get("numerator_column"))
    denominator_col = clean(metric.get("denominator_column"))
    numerator = 0.0
    denominator = 0.0
    for row in matching_rows(rows, scope_rows, brand_col, city_col, brand_alias_groups):
        numerator += safe_float(row.get(numerator_col), 0.0) or 0.0
        denominator += safe_float(row.get(denominator_col), 0.0) or 0.0
    if denominator <= 0:
        if metric.get("zero_denominator_as_pass"):
            return "100.00%", True, 1.0, "分母为0，按100%通过"
        return "", False, 0.0, "分母为 0 或缺失"
    ratio = numerator / denominator
    threshold = safe_float(metric.get("threshold"), None)
    passed = True if threshold is None else ratio >= threshold
    return f"{ratio:.2%}", passed, ratio, ""


def max_completion_base_rate(metric: dict[str, Any]) -> float:
    metric_type = clean(metric.get("type"))
    if metric_type == "tiered_ratio_from_sums":
        return max((must_float(tier.get("rate"), f"{metric.get('key')}.tiers.rate") for tier in metric.get("tiers") or []), default=0.0)
    return safe_float(metric.get("rate"), 0.0) or 0.0


def choose_tiered_metric_rate(metric: dict[str, Any], ratio: float, coefficient: float) -> tuple[float, float, float, str]:
    for tier in metric.get("tiers") or []:
        min_value = must_float(tier.get("min"), f"{metric.get('key')}.tiers.min")
        max_raw = tier.get("max")
        max_value = safe_float(max_raw, None) if max_raw is not None else None
        if ratio >= min_value and (max_value is None or ratio < max_value):
            rate = must_float(tier.get("rate"), f"{metric.get('key')}.tiers.rate")
            final = rate * coefficient if tier.get("apply_coefficient", True) else rate
            completion_earned_rate = rate
            extra_rate = safe_float(tier.get("extra_rate"), 0.0) or 0.0
            if extra_rate:
                final += extra_rate * coefficient if tier.get("extra_apply_coefficient", False) else extra_rate
                rate += extra_rate
            return rate, final, completion_earned_rate, clean(tier.get("name"))
    return 0.0, 0.0, 0.0, "未达档"


def evaluate_process(
    month: str,
    pair: tuple[str, str],
    scope_rows: list[ScopeRow],
    completed_orders: float,
    process_threshold_orders: float,
    process_coefficient: float,
    process_rules: dict[str, Any],
    target_rules: dict[str, Any],
    process_rows: dict[str, list[dict[str, Any]]],
    brand_alias_groups: dict[str, set[str]] | None = None,
) -> tuple[float, str, list[ProcessDetail], float]:
    operator, brand = pair
    details: list[ProcessDetail] = []
    min_orders = must_float(target_rules.get("min_orders_for_process_rebate", 2000), "min_orders_for_process_rebate")
    redline_passed = True

    for redline in process_rules.get("redlines") or []:
        forced, forced_value, forced_reason = forced_redline_pass(redline, pair)
        if forced:
            value_text, passed, reason = forced_value, True, forced_reason
        else:
            value_text, passed, reason = evaluate_redline(redline, scope_rows, process_rows, brand_alias_groups)
        redline_passed = redline_passed and passed
        details.append(
            ProcessDetail(
                operator=operator,
                brand=brand,
                metric_key=clean(redline.get("key")),
                metric_name=clean(redline.get("name")),
                metric_value=value_text,
                threshold_text=clean(redline.get("threshold_text")),
                passed=passed,
                base_rate=0.0,
                coefficient=1.0,
                final_rate=0.0,
                source_sheet=clean(redline.get("sheet")),
                reason=reason,
            )
        )

    cancellation_reason = ""
    if process_threshold_orders < min_orders:
        cancellation_reason = f"运营主体完单 {process_threshold_orders:.0f} < {min_orders:.0f}，无过程返佣"
    elif not redline_passed:
        cancellation_reason = "红线未达成，过程返佣取消"

    total_rate = 0.0
    completion_base_total = 0.0
    completion_earned_total = 0.0
    for metric in process_rules.get("metrics") or []:
        metric_type = clean(metric.get("type"))
        reason = ""
        tier_name = ""
        completion_base_rate = max_completion_base_rate(metric)
        completion_earned_rate = 0.0
        if metric_type == "tsh_growth":
            value_text, passed, _, reason = evaluate_tsh_growth(metric, month, scope_rows, process_rows, brand_alias_groups)
            base_rate = must_float(metric.get("rate"), f"{metric.get('key')}.rate") if passed else 0.0
            final_rate = base_rate * process_coefficient if metric.get("apply_coefficient", True) else base_rate
            completion_earned_rate = base_rate
        elif metric_type == "tiered_ratio_from_sums":
            value_text, passed, ratio, reason = evaluate_ratio_from_sums(metric, scope_rows, process_rows, by_city=True, brand_alias_groups=brand_alias_groups)
            base_rate, final_rate, completion_earned_rate, tier_name = choose_tiered_metric_rate(metric, ratio, process_coefficient) if passed else (0.0, 0.0, 0.0, "")
            passed = final_rate > 0
        elif metric_type == "ratio_from_sums_by_brand":
            value_text, passed, _, reason = evaluate_ratio_from_sums(metric, scope_rows, process_rows, by_city=False, brand_alias_groups=brand_alias_groups)
            base_rate = must_float(metric.get("rate"), f"{metric.get('key')}.rate") if passed else 0.0
            final_rate = base_rate * process_coefficient if metric.get("apply_coefficient", True) else base_rate
            completion_earned_rate = base_rate
        else:
            raise RuntimeError(f"未知过程指标类型: {metric_type}")

        if cancellation_reason:
            final_rate = 0.0
            reason = "; ".join(item for item in [reason, cancellation_reason] if item)
            completion_earned_rate = 0.0
        completion_base_total += completion_base_rate
        completion_earned_total += completion_earned_rate
        total_rate += final_rate
        details.append(
            ProcessDetail(
                operator=operator,
                brand=brand,
                metric_key=clean(metric.get("key")),
                metric_name=clean(metric.get("name")),
                metric_value=value_text,
                threshold_text=clean(metric.get("threshold_text")) or tier_name,
                passed=passed,
                base_rate=base_rate,
                coefficient=process_coefficient if metric.get("apply_coefficient", True) else 1.0,
                final_rate=final_rate,
                source_sheet=clean(metric.get("sheet")),
                reason=reason,
                completion_base_rate=completion_base_rate,
                completion_earned_rate=completion_earned_rate,
            )
        )
    status = "通过" if redline_passed else "未通过"
    if process_threshold_orders < min_orders:
        status = "完单不足"
    completion_rate = completion_earned_total / completion_base_total if completion_base_total else 0.0
    return total_rate, status, details, completion_rate


def aggregate_billing_for_operator(operator: str, billing_by_pair: dict[tuple[str, str], BillingAgg]) -> float:
    return sum(agg.rebate_base for (op, _), agg in billing_by_pair.items() if op == operator)


def combine_operator_aggregate_target(
    operator: str,
    items: list[TravelTarget],
    validations: list[ValidationItem],
) -> TravelTarget:
    def sum_optional(values: list[float | None]) -> float | None:
        present = [value for value in values if value is not None]
        return sum(present) if present else None

    def first_text(attr: str) -> str:
        values = [clean(getattr(item, attr)) for item in items if clean(getattr(item, attr))]
        unique = sorted(set(values))
        if len(unique) > 1:
            validations.append(ValidationItem("聚合目标字段不一致", operator, "", "", f"{attr}: {', '.join(unique)}；按第一条使用"))
        return values[0] if values else ""

    def first_number(attr: str, default: float = 0.0) -> float:
        values = [float(getattr(item, attr)) for item in items]
        unique = sorted(set(values))
        if len(unique) > 1:
            validations.append(ValidationItem("聚合目标字段不一致", operator, "", "", f"{attr}: {', '.join(str(v) for v in unique)}；按第一条使用"))
        return values[0] if values else default

    return TravelTarget(
        operator=operator,
        brand=items[0].brand,
        first_target=sum_optional([item.first_target for item in items]),
        second_target=sum_optional([item.second_target for item in items]),
        third_target=sum_optional([item.third_target for item in items]),
        first_miss_label=first_text("first_miss_label"),
        first_hit_label=first_text("first_hit_label"),
        second_hit_label=first_text("second_hit_label"),
        third_hit_label=first_text("third_hit_label"),
        assessment_mode="聚合考核",
        process_coefficient=first_number("process_coefficient", 1.0),
        extra_rate=first_number("extra_rate", 0.0),
        raw_payload={"combined_targets": [item.raw_payload for item in items]},
    )


def build_operator_aggregate_targets(
    travel_targets: dict[tuple[str, str], TravelTarget],
    split_marker: str,
    validations: list[ValidationItem],
) -> dict[str, TravelTarget]:
    by_operator: dict[str, list[TravelTarget]] = defaultdict(list)
    for target in travel_targets.values():
        if clean(target.assessment_mode) != split_marker:
            by_operator[target.operator].append(target)
    return {
        operator: combine_operator_aggregate_target(operator, targets, validations)
        for operator, targets in by_operator.items()
        if targets
    }


def build_results(ctx: CalculationContext) -> list[ResultRow]:
    by_pair, by_operator, _, scope_validations = build_scope_indexes(ctx.scope_rows)
    ctx.validations.extend(scope_validations)
    scale_rules = ctx.rules.get("scale_rebate") or {}
    scale_tiers = scale_rules.get("tiers") or []
    target_rules = ctx.rules.get("targets") or {}
    exclude_youxing_from_travel = bool(target_rules.get("exclude_youxing_from_travel_operator_orders", True))
    by_operator_travel = {
        operator: exclude_youxing_scope(scope_rows, ctx.youxing_targets) if exclude_youxing_from_travel else scope_rows
        for operator, scope_rows in by_operator.items()
    }
    process_rules = ctx.rules.get("process") or {}
    brand_alias_groups = build_brand_alias_groups(ctx.rules)
    split_marker = clean(target_rules.get("split_marker", "分开考核"))
    min_monthly_orders = must_float(target_rules.get("min_orders_for_monthly_rebate", 1000), "min_orders_for_monthly_rebate")
    results: list[ResultRow] = []
    aggregate_targets = build_operator_aggregate_targets(ctx.travel_targets, split_marker, ctx.validations)

    target_pairs = set(ctx.travel_targets)
    scope_pairs = set(by_pair)
    for operator, brand in sorted(target_pairs - scope_pairs):
        ctx.validations.append(ValidationItem("目标表不在本次对接范围", operator, brand, "", "未进入主表"))
    for operator, brand in sorted(scope_pairs - target_pairs):
        if brand not in ctx.youxing_targets and operator not in aggregate_targets:
            ctx.validations.append(ValidationItem("公司库范围缺少目标", operator, brand, "", "出行/优行目标表均未命中"))

    operator_scale_cache: dict[str, tuple[float, str]] = {}
    aggregate_eval_cache: dict[str, dict[str, Any]] = {}
    for pair in sorted(scope_pairs):
        operator, brand = pair
        scope_rows = by_pair[pair]
        contacts = sorted({row.contact_person for row in scope_rows if row.contact_person})
        contact_person = "、".join(contacts)
        billing = ctx.billing_by_pair.get(pair, BillingAgg())
        new_city_amount = ctx.open_city_by_pair.get(pair, 0.0)
        status = "正常"
        reason_parts: list[str] = []
        if billing.filtered_row_count == 0:
            status = "无账单数据"
            reason_parts.append("本次对接范围内未匹配到账单返佣基数")
        if new_city_amount and billing.rebate_base == 0:
            reason_parts.append("有新开城金额但非共补GMV为0，新开城占比按0")

        if brand in ctx.youxing_targets:
            target = ctx.youxing_targets[brand]
            target_cities = set(target.cities)
            target_scope = [row for row in scope_rows if not target_cities or row.city in target_cities]
            completed = sum_completed(target_scope, ctx.completed_by_brand_city)
            target_completed = completed_orders_for_target(completed, ctx.month, target_rules)
            hit_rate = 0.0
            tier_name = "未达优行阶梯"
            for threshold, rate, name in sorted(target.tiers, key=lambda item: item[0]):
                if target_completed >= threshold:
                    hit_rate = rate
                    tier_name = name
            result = ResultRow(
                month_label=ctx.month_label,
                operator=operator,
                brand=brand,
                contact_person=contact_person,
                gross_all=billing.gross_all,
                rebate_base_before_open_city=billing.rebate_base_before_open_city,
                open_city_excluded_base=billing.open_city_excluded_base,
                rebate_base=billing.rebate_base,
                scale_rate=0.0,
                process_rate=0.0,
                monthly_order_rate=hit_rate,
                extra_rate=0.0,
                final_rate=hit_rate,
                completed_orders=completed,
                completed_orders_for_target=target_completed,
                tier_name=tier_name,
                assessment_mode="优行",
                process_coefficient=1.0,
                redline_status="-",
                status=status,
                reason="；".join(reason_parts),
                new_city_amount=new_city_amount,
            )
            results.append(result)
            continue

        own_target = ctx.travel_targets.get(pair)
        default_mode = clean(target_rules.get("default_assessment_mode", "聚合考核"))
        aggregate_target = aggregate_targets.get(operator)
        is_split = bool(own_target and clean(own_target.assessment_mode) == split_marker)
        target = own_target
        if not is_split and aggregate_target is not None:
            target = aggregate_target
            if own_target is None:
                reason_parts.append("使用运营主体聚合目标")

        if target is None:
            missing_policy = clean(target_rules.get("missing_target_policy"))
            if missing_policy != "participate_with_defaults":
                results.append(
                    ResultRow(
                        month_label=ctx.month_label,
                        operator=operator,
                        brand=brand,
                        contact_person=contact_person,
                        gross_all=billing.gross_all,
                        rebate_base_before_open_city=billing.rebate_base_before_open_city,
                        open_city_excluded_base=billing.open_city_excluded_base,
                        rebate_base=billing.rebate_base,
                        scale_rate=0.0,
                        process_rate=0.0,
                        monthly_order_rate=0.0,
                        extra_rate=0.0,
                        final_rate=0.0,
                        completed_orders=sum_completed(scope_rows, ctx.completed_by_brand_city),
                        completed_orders_for_target=completed_orders_for_target(
                            sum_completed(scope_rows, ctx.completed_by_brand_city),
                            ctx.month,
                            target_rules,
                        ),
                        tier_name="目标缺失",
                        assessment_mode="目标缺失",
                        process_coefficient=1.0,
                        redline_status="-",
                        status="目标缺失" if status == "正常" else status,
                        reason="；".join([*reason_parts, "目标拆解表没有该运营主体品牌"]),
                        new_city_amount=new_city_amount,
                    )
                )
                continue

        operator_travel_scope = by_operator_travel.get(operator) or []
        operator_completed = sum_completed(operator_travel_scope, ctx.completed_by_brand_city)
        if operator not in operator_scale_cache:
            scale_basis = scale_basis_from_completed_orders(operator_completed, ctx.month, scale_rules)
            operator_scale_cache[operator] = scale_rate_for_base(scale_basis, scale_tiers)
        scale_rate, scale_tier_name = operator_scale_cache[operator]

        target_missing = target is None
        process_coefficient = (
            safe_float(target_rules.get("default_process_coefficient"), 1.0) or 1.0
            if target_missing
            else target.process_coefficient
        )
        extra_rate = 0.0 if target_missing else target.extra_rate
        if target_missing and default_mode == split_marker:
            is_split = True

        cache_key = f"operator:{operator}" if not is_split else f"pair:{operator}:{brand}"
        if cache_key in aggregate_eval_cache:
            eval_result = aggregate_eval_cache[cache_key]
        else:
            assessment_scope = scope_rows if is_split else operator_travel_scope
            completed = sum_completed(assessment_scope, ctx.completed_by_brand_city)
            process_min_basis = clean(target_rules.get("process_min_orders_basis", "operator_total_completed_orders"))
            process_threshold_orders = completed_for_threshold_basis(
                process_min_basis,
                completed,
                operator_completed,
                ctx.month,
                target_rules,
                "process_min_orders_basis",
            )
            evaluation_pair = (operator, target.brand if target else brand)
            process_rate, redline_status, details, process_completion_rate = evaluate_process(
                ctx.month,
                evaluation_pair,
                assessment_scope,
                completed,
                process_threshold_orders,
                process_coefficient,
                process_rules,
                target_rules,
                ctx.process_rows,
                brand_alias_groups,
            )
            monthly_delta = 0.0
            monthly_tier = "目标缺失-无完单考核" if target_missing else "完单不足"
            target_completed = completed_orders_for_target(completed, ctx.month, target_rules)
            eval_reason_parts: list[str] = []
            if target_missing:
                eval_reason_parts.append("目标拆解表没有该运营主体品牌，月度完单考核和额外返佣按0")
            monthly_min_basis = clean(target_rules.get("monthly_min_orders_basis", "operator_daily_completed_orders"))
            monthly_threshold_orders = completed_for_threshold_basis(
                monthly_min_basis,
                completed,
                operator_completed,
                ctx.month,
                target_rules,
                "monthly_min_orders_basis",
            )
            if target_missing:
                pass
            elif monthly_threshold_orders >= min_monthly_orders:
                monthly_delta, monthly_tier = choose_monthly_delta(target_completed, target)
            else:
                eval_reason_parts.append(f"门槛完单 {monthly_threshold_orders:.2f} < {min_monthly_orders:.0f}，无完单返佣")
            eval_result = {
                "completed": completed,
                "target_completed": target_completed,
                "process_rate": process_rate,
                "redline_status": redline_status,
                "details": details,
                "process_completion_rate": process_completion_rate,
                "monthly_delta": monthly_delta,
                "monthly_tier": monthly_tier,
                "reason_parts": eval_reason_parts,
            }
            aggregate_eval_cache[cache_key] = eval_result

        completed = float(eval_result["completed"])
        target_completed = float(eval_result["target_completed"])
        process_rate = float(eval_result["process_rate"])
        redline_status = str(eval_result["redline_status"])
        details = list(eval_result["details"])
        process_completion_rate = float(eval_result["process_completion_rate"])
        monthly_delta = float(eval_result["monthly_delta"])
        monthly_tier = str(eval_result["monthly_tier"])
        reason_parts.extend(eval_result["reason_parts"])
        final_rate = scale_rate + process_rate + monthly_delta + extra_rate
        tier_name = f"{scale_tier_name}/{monthly_tier}"
        assessment_mode = "分开考核" if is_split else "聚合考核"
        results.append(
            ResultRow(
                month_label=ctx.month_label,
                operator=operator,
                brand=brand,
                contact_person=contact_person,
                gross_all=billing.gross_all,
                rebate_base_before_open_city=billing.rebate_base_before_open_city,
                open_city_excluded_base=billing.open_city_excluded_base,
                rebate_base=billing.rebate_base,
                scale_rate=scale_rate,
                process_rate=process_rate,
                monthly_order_rate=monthly_delta,
                extra_rate=extra_rate,
                final_rate=final_rate,
                completed_orders=completed,
                completed_orders_for_target=target_completed,
                tier_name=tier_name,
                assessment_mode=assessment_mode,
                process_coefficient=process_coefficient,
                redline_status=redline_status,
                status=status,
                reason="；".join(reason_parts),
                process_completion_rate=process_completion_rate,
                new_city_amount=new_city_amount,
                process_details=details,
            )
        )
    return results


def format_rate_value(value: float) -> str:
    return f"{value:.2%}"


def write_output_workbook(path: Path, results: list[ResultRow], process_rules: dict[str, Any], validations: list[ValidationItem]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_total = wb.active
    ws_total.title = "总考核"
    ws_process = wb.create_sheet("过程考核")

    write_total_sheet(ws_total, results)
    write_process_sheet(ws_process, results, process_rules, validations)
    wb.save(path)


def style_header(ws, col_count: int) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(name="Arial", bold=True)
    for cell in ws[1][:col_count]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}1"


def write_total_sheet(ws, results: list[ResultRow]) -> None:
    headers = OLD_TOTAL_HEADERS + AUDIT_HEADERS
    ws.append(headers)
    for result in results:
        point_amount = result.rebate_amount
        total_amount = result.total_rebate_amount
        ws.append(
            [
                result.month_label,
                f"{result.operator}-{result.brand}",
                result.operator,
                result.brand,
                result.contact_person,
                result.gross_all,
                result.rebate_base,
                point_amount,
                result.final_rate,
                0,
                0,
                0,
                0,
                result.new_city_amount,
                result.new_city_rate,
                0,
                result.process_completion_rate,
                total_amount,
                result.total_rate,
                result.rebate_base,
                result.scale_rate,
                result.process_rate,
                result.t_rate,
                result.monthly_order_rate,
                result.extra_rate,
                result.final_rate,
                point_amount,
                result.rebate_base_before_open_city,
                result.open_city_excluded_base,
                result.new_city_amount,
                result.new_city_rate,
                total_amount,
                result.total_rate,
                result.completed_orders,
                result.completed_orders_for_target,
                result.tier_name,
                result.assessment_mode,
                result.process_coefficient,
                result.redline_status,
                result.status,
                result.reason,
            ]
        )
    style_header(ws, len(headers))
    rate_cols = [9, 11, 13, 15, 17, 19, 21, 22, 23, 24, 25, 26, 31, 33]
    amount_cols = [6, 7, 8, 10, 12, 14, 16, 18, 20, 27, 28, 29, 30, 32, 34, 35]
    for row in ws.iter_rows(min_row=2):
        for idx in amount_cols:
            row[idx - 1].number_format = "#,##0.00"
        for idx in rate_cols:
            row[idx - 1].number_format = "0.00%"
    widths = {
        "A": 12,
        "B": 24,
        "C": 16,
        "D": 16,
        "E": 12,
        "F": 16,
        "G": 16,
        "R": 14,
        "AH": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col_idx in range(8, len(headers) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, 14)


def write_process_sheet(ws, results: list[ResultRow], process_rules: dict[str, Any], validations: list[ValidationItem]) -> None:
    metric_rules = list(process_rules.get("redlines") or []) + list(process_rules.get("metrics") or [])
    base_headers = [
        "月份",
        "主体&品牌",
        "运营主体",
        "品牌",
        "对接人",
        "考核方式",
        "过程系数",
        "月总完单数",
        "完单考核日均完单",
        "红线状态",
        "过程考核返点汇总",
    ]
    metric_headers: list[str] = []
    for metric in metric_rules:
        name = clean(metric.get("name"))
        metric_headers.extend([f"{name}数值", f"{name}达标", f"{name}基础返点", f"{name}最终返点", f"{name}原因"])
    headers = base_headers + metric_headers
    ws.append(headers)
    for result in results:
        details = {detail.metric_key: detail for detail in result.process_details}
        row = [
            result.month_label,
            f"{result.operator}-{result.brand}",
            result.operator,
            result.brand,
            result.contact_person,
            result.assessment_mode,
            result.process_coefficient,
            result.completed_orders,
            result.completed_orders_for_target,
            result.redline_status,
            result.process_rate,
        ]
        for metric in metric_rules:
            detail = details.get(clean(metric.get("key")))
            if detail is None:
                row.extend(["", "", 0, 0, ""])
            else:
                row.extend([
                    detail.metric_value,
                    "是" if detail.passed else "否",
                    detail.base_rate,
                    detail.final_rate,
                    detail.reason,
                ])
        ws.append(row)

    style_header(ws, len(headers))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[6].number_format = "0.00"
        row[7].number_format = "#,##0"
        row[8].number_format = "#,##0.00"
        row[10].number_format = "0.00%"
        first_metric_base_rate_col = len(base_headers) + 3
        for idx in range(first_metric_base_rate_col, len(headers) + 1, 5):
            row[idx - 1].number_format = "0.00%"
            if idx < len(row):
                row[idx].number_format = "0.00%"
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.column_dimensions["B"].width = 24

    start_row = ws.max_row + 3
    ws.cell(start_row, 1, "校验明细")
    ws.cell(start_row, 1).font = Font(name="Arial", bold=True)
    validation_headers = ["校验类型", "运营主体", "品牌", "城市", "说明"]
    for idx, header in enumerate(validation_headers, start=1):
        cell = ws.cell(start_row + 1, idx, header)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
        cell.font = Font(name="Arial", bold=True)
    if validations:
        for item in validations:
            ws.append([item.category, item.operator, item.brand, item.city, item.message])
    else:
        ws.append(["无异常", "", "", "", ""])


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [json_safe(v) for v in value]
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def process_raw_db_rows(process_rows: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    result = []
    for sheet, rows in process_rows.items():
        for row in rows:
            result.append(
                {
                    "sheet_name": sheet,
                    "brand_name": clean(
                        row.get("品牌名称")
                        or row.get("运力品牌")
                        or row.get("company_name")
                        or row.get("商家名称")
                    ),
                    "city_name": clean(row.get("城市名称") or row.get("城市") or row.get("city_name")),
                    "operator_name": clean(row.get("运营方") or row.get("运营方ID") or row.get("agent_level1")),
                    "metric_date": clean(row.get("日期(月)") or row.get("年") or row.get("月份") or row.get("月")),
                    "raw_payload": json_safe(row),
                }
            )
    return result


def ensure_database_schema(cursor: Any) -> None:
    ddl_path = SCRIPT_DIR / "db_schema.sql"
    ddl_statements = [stmt.strip() for stmt in ddl_path.read_text(encoding="utf-8").split(";") if stmt.strip()]
    for statement in ddl_statements:
        cursor.execute(statement)


def cursor_rows_as_dicts(cursor: Any) -> list[dict[str, Any]]:
    columns = [desc[0] for desc in cursor.description or []]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def json_payload(value: Any) -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        parsed = json.loads(value)
        if isinstance(parsed, dict):
            return parsed
    return dict(value)


def sheet_key(role: str, path: Path | str, sheet_name: str) -> tuple[str, str, str]:
    return clean(role), str(path), clean(sheet_name)


def require_src_sheet_id(
    sheet_ids: dict[tuple[str, str, str], int],
    role: str,
    path: Path | str,
    sheet_name: str,
) -> int:
    key = sheet_key(role, path, sheet_name)
    if key not in sheet_ids:
        raise RuntimeError(f"未找到源 sheet 登记: role={role}, file={path}, sheet={sheet_name}")
    return sheet_ids[key]


def sync_import_database(ctx: ImportContext) -> None:
    from lxx_share.database import DatabaseConnector

    db = DatabaseConnector()
    with db.connect() as conn:
        cursor = conn.cursor()
        try:
            ensure_database_schema(cursor)
            cursor.execute(
                """
                INSERT INTO lxfandian.imports
                    (id, month, import_type, work_dir, rules_hash, rules_text, status, confirmed_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, now())
                """,
                (
                    ctx.import_id,
                    ctx.month,
                    ctx.import_type,
                    str(ctx.work_dir),
                    ctx.rules_hash,
                    ctx.rules_text,
                    "confirmed",
                ),
            )
            sheet_ids: dict[tuple[str, str, str], int] = {}
            for role, path, sheet, row_count in ctx.source_sheet_counts:
                cursor.execute(
                    """
                    INSERT INTO lxfandian.src_sheets
                        (import_id, file_role, file_path, file_hash, sheet_name, row_count)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (ctx.import_id, role, str(path), ctx.source_file_hashes[role], sheet, row_count),
                )
                sheet_ids[sheet_key(role, path, sheet)] = int(cursor.fetchone()[0])
            billing_path = ctx.source_files.get("billing")
            if billing_path:
                batch = []
                for row in iter_billing_raw_rows(billing_path, ctx.rules):
                    batch.append(
                        (
                            ctx.import_id,
                            require_src_sheet_id(sheet_ids, "billing", billing_path, row.sheet_name),
                            row.row_no,
                            row.partition_date,
                            row.brand,
                            row.city,
                            row.tr_type,
                            row.channel,
                            row.gross_receivable,
                            row.spring_service_fee,
                            json.dumps(row.raw_payload, ensure_ascii=False),
                        )
                    )
                    if len(batch) >= 1000:
                        cursor.executemany(
                            """
                            INSERT INTO lxfandian.bill_raw
                                (import_id, src_sheet_id, row_no, partition_date, brand_name, city_name,
                                 tr_type, channel, gross_receivable, spring_service_fee, raw_payload)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                            """,
                            batch,
                        )
                        batch.clear()
                if batch:
                    cursor.executemany(
                        """
                        INSERT INTO lxfandian.bill_raw
                            (import_id, src_sheet_id, row_no, partition_date, brand_name, city_name,
                             tr_type, channel, gross_receivable, spring_service_fee, raw_payload)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                        """,
                        batch,
                    )
            for (sheet, partition_date, brand, city, tr_type, channel), agg in ctx.billing_import_agg.items():
                cursor.execute(
                    """
                    INSERT INTO lxfandian.bill_agg
                        (import_id, src_sheet_id, partition_date, brand_name, city_name, tr_type, channel, gross_receivable,
                         spring_service_fee, row_count)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        ctx.import_id,
                        require_src_sheet_id(sheet_ids, "billing", billing_path or "", sheet),
                        parse_date_value(partition_date),
                        brand,
                        city,
                        tr_type,
                        channel,
                        agg.gross_receivable,
                        agg.spring_service_fee,
                        agg.row_count,
                    ),
                )
            process_path = ctx.source_files.get("process")
            for row in process_raw_db_rows(ctx.process_rows):
                cursor.execute(
                    """
                    INSERT INTO lxfandian.proc_raw
                        (import_id, src_sheet_id, brand_name, city_name, operator_name, metric_date, raw_payload)
                    VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb)
                    """,
                    (
                        ctx.import_id,
                        require_src_sheet_id(sheet_ids, "process", process_path or "", row["sheet_name"]),
                        row["brand_name"],
                        row["city_name"],
                        row["operator_name"],
                        row["metric_date"],
                        json.dumps(row["raw_payload"], ensure_ascii=False),
                    ),
                )
            target_path = ctx.source_files.get("target")
            for target_type, sheet_name, payload in ctx.target_raw_rows:
                cursor.execute(
                    """
                    INSERT INTO lxfandian.targets
                        (import_id, src_sheet_id, target_type, operator_entity, brand_name, cities_text, target_payload)
                    VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb)
                    """,
                    (
                        ctx.import_id,
                        require_src_sheet_id(sheet_ids, "target", target_path or "", sheet_name),
                        target_type,
                        clean(payload.get("运营主体")),
                        clean(payload.get("品牌名称") or payload.get("商家")),
                        clean(payload.get("考核具体城市")),
                        json.dumps(payload, ensure_ascii=False),
                    ),
                )
            open_city_path = ctx.source_files.get("open_city")
            for row in ctx.open_city_rows:
                cursor.execute(
                    """
                    INSERT INTO lxfandian.open_city
                        (import_id, src_sheet_id, brand_name, city_name, settlement_type, settlement_unit,
                         settlement_item, open_date, incentive_period, rate, settlement_period,
                         rebate_basis, reward_amount, remark1, remark2, raw_payload)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                    """,
                    (
                        ctx.import_id,
                        require_src_sheet_id(sheet_ids, "open_city", open_city_path or "", row.sheet_name),
                        row.brand,
                        row.city,
                        row.settlement_type,
                        row.settlement_unit,
                        row.settlement_item,
                        row.open_date,
                        row.incentive_period,
                        row.rate,
                        row.settlement_period,
                        row.rebate_basis,
                        row.reward_amount,
                        row.remark1,
                        row.remark2,
                        json.dumps(row.raw_payload, ensure_ascii=False),
                    ),
                )
            conn.commit()
        except Exception:
            conn.rollback()
            raise


def latest_confirmed_import_id(
    month: str,
    import_types: list[str] | None = None,
    required: bool = True,
) -> str:
    from lxx_share.database import DatabaseConnector

    import_types = import_types or ["full"]
    db = DatabaseConnector()
    with db.connect() as conn:
        cursor = conn.cursor()
        ensure_database_schema(cursor)
        conn.commit()
        placeholders = ", ".join(["%s"] * len(import_types))
        cursor.execute(
            f"""
            SELECT id
            FROM lxfandian.imports
            WHERE month = %s AND status = 'confirmed' AND import_type IN ({placeholders})
            ORDER BY confirmed_at DESC, created_at DESC
            LIMIT 1
            """,
            (month, *import_types),
        )
        row = cursor.fetchone()
    if not row:
        if required:
            raise RuntimeError(f"未找到 {month} 的 confirmed import 批次，请先运行 import-sources --confirmed")
        return ""
    return str(row[0])


def load_import_payload_from_database(import_id: str) -> tuple[
    str,
    str,
    dict[str, Path],
    dict[str, str],
    list[tuple[str, Path, str, int]],
    dict[tuple[str, str, str, str, str, str], BillingImportAgg],
    dict[str, list[dict[str, Any]]],
    list[OpenCityRow],
    list[tuple[str, str, dict[str, Any]]],
]:
    from lxx_share.database import DatabaseConnector

    db = DatabaseConnector()
    with db.connect() as conn:
        cursor = conn.cursor()
        ensure_database_schema(cursor)
        conn.commit()
        cursor.execute(
            """
            SELECT id, month, rules_hash
            FROM lxfandian.imports
            WHERE id = %s AND status = 'confirmed'
            """,
            (import_id,),
        )
        import_row = cursor.fetchone()
        if not import_row:
            raise RuntimeError(f"未找到 confirmed import 批次: {import_id}")
        import_month = str(import_row[1])
        import_rules_hash = str(import_row[2])

        cursor.execute(
            """
            SELECT id, file_role, file_path, file_hash, sheet_name, row_count
            FROM lxfandian.src_sheets
            WHERE import_id = %s
            ORDER BY file_role, sheet_name
            """,
            (import_id,),
        )
        source_rows = cursor_rows_as_dicts(cursor)

        cursor.execute(
            """
            SELECT s.sheet_name, b.partition_date, b.brand_name, b.city_name, b.tr_type, b.channel,
                   b.gross_receivable, b.spring_service_fee, b.row_count
            FROM lxfandian.bill_agg b
            JOIN lxfandian.src_sheets s ON s.id = b.src_sheet_id
            WHERE b.import_id = %s
            """,
            (import_id,),
        )
        billing_rows = cursor_rows_as_dicts(cursor)

        cursor.execute(
            """
            SELECT s.sheet_name, p.raw_payload
            FROM lxfandian.proc_raw p
            JOIN lxfandian.src_sheets s ON s.id = p.src_sheet_id
            WHERE p.import_id = %s
            ORDER BY p.id
            """,
            (import_id,),
        )
        process_db_rows = cursor_rows_as_dicts(cursor)

        cursor.execute(
            """
            SELECT s.sheet_name, t.target_type, t.target_payload
            FROM lxfandian.targets t
            JOIN lxfandian.src_sheets s ON s.id = t.src_sheet_id
            WHERE t.import_id = %s
            ORDER BY t.id
            """,
            (import_id,),
        )
        target_db_rows = cursor_rows_as_dicts(cursor)

        cursor.execute(
            """
            SELECT s.sheet_name, o.brand_name, o.city_name, o.settlement_type, o.settlement_unit,
                   o.settlement_item, o.open_date, o.incentive_period, o.rate, o.settlement_period,
                   o.rebate_basis, o.reward_amount, o.remark1, o.remark2, o.raw_payload
            FROM lxfandian.open_city o
            JOIN lxfandian.src_sheets s ON s.id = o.src_sheet_id
            WHERE o.import_id = %s
            ORDER BY o.id
            """,
            (import_id,),
        )
        open_city_db_rows = cursor_rows_as_dicts(cursor)

    source_files: dict[str, Path] = {}
    source_file_hashes: dict[str, str] = {}
    source_sheet_counts: list[tuple[str, Path, str, int]] = []
    for row in source_rows:
        role = clean(row.get("file_role"))
        path = Path(clean(row.get("file_path")))
        source_files.setdefault(role, path)
        source_file_hashes.setdefault(role, clean(row.get("file_hash")))
        source_sheet_counts.append((role, path, clean(row.get("sheet_name")), int(row.get("row_count") or 0)))

    billing_import_agg: dict[tuple[str, str, str, str, str, str], BillingImportAgg] = {}
    for row in billing_rows:
        billing_import_agg[
            (
                clean(row.get("sheet_name")),
                format_date_key(parse_date_value(row.get("partition_date"))),
                clean(row.get("brand_name")),
                clean(row.get("city_name")),
                clean(row.get("tr_type")),
                clean(row.get("channel")),
            )
        ] = BillingImportAgg(
            gross_receivable=float(row.get("gross_receivable") or 0.0),
            spring_service_fee=float(row.get("spring_service_fee") or 0.0),
            row_count=int(row.get("row_count") or 0),
        )

    process_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in process_db_rows:
        process_rows[clean(row.get("sheet_name"))].append(json_payload(row.get("raw_payload")))

    target_raw_rows = [
        (clean(row.get("target_type")), clean(row.get("sheet_name")), json_payload(row.get("target_payload")))
        for row in target_db_rows
    ]

    open_city_rows = [
        OpenCityRow(
            sheet_name=clean(row.get("sheet_name")),
            brand=clean(row.get("brand_name")),
            city=clean(row.get("city_name")),
            settlement_type=clean(row.get("settlement_type")),
            settlement_unit=clean(row.get("settlement_unit")),
            settlement_item=clean(row.get("settlement_item")),
            open_date=clean(row.get("open_date")),
            incentive_period=clean(row.get("incentive_period")),
            rate=float(row.get("rate") or 0.0),
            settlement_period=clean(row.get("settlement_period")),
            rebate_basis=float(row.get("rebate_basis") or 0.0),
            reward_amount=float(row.get("reward_amount") or 0.0),
            remark1=clean(row.get("remark1")),
            remark2=clean(row.get("remark2")),
            raw_payload=json_payload(row.get("raw_payload")),
        )
        for row in open_city_db_rows
    ]

    return (
        import_month,
        import_rules_hash,
        source_files,
        source_file_hashes,
        source_sheet_counts,
        billing_import_agg,
        dict(process_rows),
        open_city_rows,
        target_raw_rows,
    )


def append_previous_tsh_rows(
    month: str,
    process_rows: dict[str, list[dict[str, Any]]],
    validations: list[ValidationItem],
) -> None:
    prev_month = previous_month(month)
    prev_import_id = latest_confirmed_import_id(prev_month, ["process_only", "full"], required=False)
    if not prev_import_id:
        validations.append(ValidationItem("上月过程数据缺失", "", "", "", f"{prev_month} 未找到 process_only/full import，TSH增速可能无法得分"))
        return
    (
        _import_month,
        _import_rules_hash,
        _source_files,
        _source_file_hashes,
        _source_sheet_counts,
        _billing_import_agg,
        previous_process_rows,
        _open_city_rows,
        _target_raw_rows,
    ) = load_import_payload_from_database(prev_import_id)
    previous_tsh_rows = previous_process_rows.get("TSH") or []
    process_rows.setdefault("TSH", []).extend(previous_tsh_rows)
    validations.append(
        ValidationItem(
            "上月过程数据已载入",
            "",
            "",
            "",
            f"{prev_month} import_id={prev_import_id}，追加 TSH 行数={len(previous_tsh_rows)}",
        )
    )


def sync_database(ctx: CalculationContext) -> None:
    from lxx_share.database import DatabaseConnector

    db = DatabaseConnector()
    now_status = "confirmed"
    with db.connect() as conn:
        cursor = conn.cursor()
        try:
            ensure_database_schema(cursor)
            cursor.execute(
                """
                INSERT INTO lxfandian.runs
                    (id, import_id, data_source, month, contacts, exclude_operators,
                     rules_hash, rules_text, status, confirmed_at)
                VALUES (%s, %s, %s, %s, %s::jsonb, %s::jsonb, %s, %s, %s, now())
                """,
                (
                    ctx.run_id,
                    ctx.import_id or None,
                    ctx.data_source,
                    ctx.month,
                    json.dumps(ctx.contacts, ensure_ascii=False),
                    json.dumps(ctx.exclude_operators, ensure_ascii=False),
                    ctx.rules_hash,
                    ctx.rules_text,
                    now_status,
                ),
            )
            for row in ctx.scope_rows:
                cursor.execute(
                    """
                    INSERT INTO lxfandian.scope
                        (run_id, operator_entity, brand_name, city_name, contact_person)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (ctx.run_id, row.operator, row.brand, row.city, row.contact_person),
                )
            for (brand, city), agg in ctx.billing_by_brand_city.items():
                cursor.execute(
                    """
                    INSERT INTO lxfandian.base_agg
                        (run_id, brand_name, city_name, gross_all, gross_receivable,
                         spring_service_fee, rebate_base_before_open_city, open_city_excluded_base,
                         rebate_base, filtered_row_count)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        ctx.run_id,
                        brand,
                        city,
                        agg.gross_all,
                        agg.gross_receivable,
                        agg.spring_service_fee,
                        agg.rebate_base_before_open_city,
                        agg.open_city_excluded_base,
                        agg.rebate_base,
                        agg.filtered_row_count,
                    ),
                )
            for result in ctx.results:
                cursor.execute(
                    """
                    INSERT INTO lxfandian.results
                        (run_id, month, operator_entity, brand_name, contact_person, gross_all,
                         rebate_base_before_open_city, open_city_excluded_base, rebate_base,
                         completed_orders, completed_orders_for_target, scale_rate, process_rate,
                         process_completion_rate, monthly_order_rate, extra_rate, final_rate, point_rebate_amount,
                         new_city_amount, new_city_rate, total_rebate_amount, total_rate, tier_name, status, reason)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        ctx.run_id,
                        ctx.month,
                        result.operator,
                        result.brand,
                        result.contact_person,
                        result.gross_all,
                        result.rebate_base_before_open_city,
                        result.open_city_excluded_base,
                        result.rebate_base,
                        result.completed_orders,
                        result.completed_orders_for_target,
                        result.scale_rate,
                        result.process_rate,
                        result.process_completion_rate,
                        result.monthly_order_rate,
                        result.extra_rate,
                        result.final_rate,
                        result.rebate_amount,
                        result.new_city_amount,
                        result.new_city_rate,
                        result.total_rebate_amount,
                        result.total_rate,
                        result.tier_name,
                        result.status,
                        result.reason,
                    ),
                )
                for detail in result.process_details:
                    cursor.execute(
                        """
                        INSERT INTO lxfandian.proc_detail
                            (run_id, operator_entity, brand_name, metric_key, metric_name, metric_value,
                             threshold_text, passed, base_rate, coefficient, final_rate, source_sheet, reason)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            ctx.run_id,
                            result.operator,
                            result.brand,
                            detail.metric_key,
                            detail.metric_name,
                            detail.metric_value,
                            detail.threshold_text,
                            detail.passed,
                            detail.base_rate,
                            detail.coefficient,
                            detail.final_rate,
                            detail.source_sheet,
                            detail.reason,
                        ),
                    )
            conn.commit()
        except Exception:
            conn.rollback()
            raise


def resolve_work_dir(month: str, work_dir_arg: str) -> Path:
    work_dir = Path(work_dir_arg).expanduser() if work_dir_arg else month_dir_from_month(month)
    if not work_dir.is_absolute():
        work_dir = PROJECT_ROOT / work_dir
    return work_dir


def load_month_rules(month: str, work_dir_arg: str) -> tuple[str, Path, Path, str, str, dict[str, Any]]:
    label = month_label(month)
    work_dir = resolve_work_dir(month, work_dir_arg)
    rules_path = work_dir / "规则" / "rules.yaml"
    if not rules_path.exists():
        raise RuntimeError(f"rules.yaml 不存在，请先运行 init-rules: {rules_path}")
    rules_text = rules_path.read_text(encoding="utf-8")
    rules = read_yaml(rules_path)
    return label, work_dir, rules_path, hash_text(rules_text), rules_text, rules


def build_import_context(args: argparse.Namespace) -> ImportContext:
    month = args.month
    label, work_dir, rules_path, rules_hash, rules_text, rules = load_month_rules(month, args.work_dir)
    source_files = resolve_source_files(work_dir, rules)
    source_file_hashes = {role: hash_file(path) for role, path in source_files.items()}
    billing_import_agg = read_billing_import_agg(source_files["billing"], rules)
    process_rows = read_process_rows(source_files["process"])
    open_city_rows = read_open_city_rows(source_files["open_city"], rules) if "open_city" in source_files else []
    _, _, target_raw_rows = read_targets(source_files["target"], rules)
    source_sheet_counts: list[tuple[str, Path, str, int]] = []
    for role, path in source_files.items():
        source_sheet_counts.extend(workbook_sheet_counts(role, path))
    return ImportContext(
        import_id=str(uuid.uuid4()),
        import_type="full",
        month=month,
        month_label=label,
        work_dir=work_dir,
        rules_path=rules_path,
        rules_hash=rules_hash,
        rules_text=rules_text,
        rules=rules,
        source_files=source_files,
        source_file_hashes=source_file_hashes,
        billing_import_agg=billing_import_agg,
        process_rows=process_rows,
        open_city_rows=open_city_rows,
        target_raw_rows=target_raw_rows,
        source_sheet_counts=source_sheet_counts,
    )


def build_process_import_context(args: argparse.Namespace) -> ImportContext:
    month = args.month
    label = month_label(month)
    work_dir = resolve_work_dir(month, args.work_dir)
    pattern = clean(getattr(args, "process_pattern", "") or "*LX品牌过程考核数据*.xlsx")
    process_file = resolve_process_file(work_dir, pattern)
    source_files = {"process": process_file}
    source_file_hashes = {"process": hash_file(process_file)}
    process_rows = read_process_rows(process_file)
    rules_text = f"import_type: process_only\nprocess_pattern: {pattern}\n"
    rules_hash = hash_text(rules_text)
    return ImportContext(
        import_id=str(uuid.uuid4()),
        import_type="process_only",
        month=month,
        month_label=label,
        work_dir=work_dir,
        rules_path=work_dir / "规则" / "rules.yaml",
        rules_hash=rules_hash,
        rules_text=rules_text,
        rules={},
        source_files=source_files,
        source_file_hashes=source_file_hashes,
        billing_import_agg={},
        process_rows=process_rows,
        open_city_rows=[],
        target_raw_rows=[],
        source_sheet_counts=workbook_sheet_counts("process", process_file),
    )


def build_context(args: argparse.Namespace) -> CalculationContext:
    month = args.month
    label, work_dir, rules_path, rules_hash, rules_text, rules = load_month_rules(month, args.work_dir)
    contacts = parse_contacts(args.contacts, rules.get("contacts") or [])
    exclude_operators = [clean(item) for item in (rules.get("exclude_operators") or []) if clean(item)]
    if args.exclude_operator:
        exclude_operators.extend(clean(item) for item in args.exclude_operator if clean(item))
        exclude_operators = sorted(set(exclude_operators))

    data_source = clean(getattr(args, "source", "db") or "db")
    validations: list[ValidationItem] = []
    scope_rows, scope_validations = load_operator_scope(contacts, exclude_operators)
    validations.extend(scope_validations)
    scope_rows = apply_scope_overrides(scope_rows, rules, validations)
    by_pair, _, brand_city_owner, scope_index_validations = build_scope_indexes(scope_rows)
    validations.extend(scope_index_validations)

    import_id = ""
    import_rules_hash = rules_hash
    if data_source == "excel":
        source_files = resolve_source_files(work_dir, rules)
        source_file_hashes = {role: hash_file(path) for role, path in source_files.items()}
        open_city_rows = read_open_city_rows(source_files["open_city"], rules) if "open_city" in source_files else []
        open_city_periods = build_open_city_periods(open_city_rows, rules, brand_city_owner, validations)
        billing_import_agg = read_billing_import_agg(source_files["billing"], rules)
        billing_by_brand_city, billing_by_pair = build_billing_from_import_agg(
            billing_import_agg,
            rules,
            brand_city_owner,
            open_city_periods,
        )
        process_rows = read_process_rows(source_files["process"])
        open_city_by_pair = build_open_city_by_pair(open_city_rows, rules, brand_city_owner, validations)
        completed_by_brand_city = completed_orders_from_process_rows(process_rows)
        travel_targets, youxing_targets, target_raw_rows = read_targets(source_files["target"], rules)
        source_sheet_counts: list[tuple[str, Path, str, int]] = []
        for role, path in source_files.items():
            source_sheet_counts.extend(workbook_sheet_counts(role, path))
    elif data_source == "db":
        import_id = clean(getattr(args, "import_id", "") or "") or latest_confirmed_import_id(month)
        (
            import_month,
            import_rules_hash,
            source_files,
            source_file_hashes,
            source_sheet_counts,
            billing_import_agg,
            process_rows,
            open_city_rows,
            target_raw_rows,
        ) = load_import_payload_from_database(import_id)
        if import_month != month:
            raise RuntimeError(f"import 批次月份不匹配: import_id={import_id}, import_month={import_month}, month={month}")
        open_city_periods = build_open_city_periods(open_city_rows, rules, brand_city_owner, validations)
        billing_by_brand_city, billing_by_pair = build_billing_from_import_agg(
            billing_import_agg,
            rules,
            brand_city_owner,
            open_city_periods,
        )
        open_city_by_pair = build_open_city_by_pair(open_city_rows, rules, brand_city_owner, validations)
        completed_by_brand_city = completed_orders_from_process_rows(process_rows)
        travel_targets, youxing_targets = parse_targets_from_raw_rows(target_raw_rows)
    else:
        raise RuntimeError(f"--source 只支持 db 或 excel: {data_source}")

    append_previous_tsh_rows(month, process_rows, validations)

    ctx = CalculationContext(
        run_id=str(uuid.uuid4()),
        import_id=import_id,
        data_source=data_source,
        month=month,
        month_label=label,
        work_dir=work_dir,
        rules_path=rules_path,
        rules_hash=rules_hash,
        rules_text=rules_text,
        import_rules_hash=import_rules_hash,
        rules=rules,
        source_files=source_files,
        source_file_hashes=source_file_hashes,
        contacts=contacts,
        exclude_operators=exclude_operators,
        scope_rows=scope_rows,
        billing_by_brand_city=billing_by_brand_city,
        billing_by_pair=billing_by_pair,
        completed_by_brand_city=completed_by_brand_city,
        process_rows=process_rows,
        open_city_rows=open_city_rows,
        open_city_by_pair=open_city_by_pair,
        travel_targets=travel_targets,
        youxing_targets=youxing_targets,
        target_raw_rows=target_raw_rows,
        source_sheet_counts=source_sheet_counts,
        validations=validations,
        results=[],
    )
    # Ensure scope pairs with no billing still appear in output.
    for pair in by_pair:
        ctx.billing_by_pair.setdefault(pair, BillingAgg())
    ctx.results = build_results(ctx)
    return ctx


def print_summary(ctx: CalculationContext, output_path: Path | None) -> None:
    print("月度返点计算预览")
    print("=" * 40)
    print(f"run_id: {ctx.run_id}")
    if ctx.data_source == "db":
        print(f"数据来源: PostgreSQL import_id={ctx.import_id}")
        if ctx.import_rules_hash != ctx.rules_hash:
            print(f"规则hash: 当前={ctx.rules_hash[:12]} / 导入时={ctx.import_rules_hash[:12]}")
    else:
        print("数据来源: Excel 源文件")
    print(f"月份: {ctx.month_label}")
    print(f"对接人: {', '.join(ctx.contacts)}")
    print(f"排除运营主体: {', '.join(ctx.exclude_operators) or '-'}")
    print(f"规则: {ctx.rules_path}")
    for role, path in ctx.source_files.items():
        print(f"源文件[{role}]: {path}")
    print(f"公司库范围: {len(ctx.scope_rows)} 个品牌城市")
    print(f"输出主表: {len(ctx.results)} 行")
    print(f"开城奖励原始行: {len(ctx.open_city_rows)} 行")
    print(f"开城奖励匹配主体品牌: {len(ctx.open_city_by_pair)} 个")
    print(f"校验明细: {len(ctx.validations)} 条")
    if output_path:
        print(f"输出文件: {output_path}")
    status_counts: dict[str, int] = defaultdict(int)
    for result in ctx.results:
        status_counts[result.status] += 1
    print("状态分布: " + ", ".join(f"{k}={v}" for k, v in sorted(status_counts.items())))
    if ctx.validations:
        print("校验样例:")
        for item in ctx.validations[:10]:
            print(f"  - [{item.category}] {item.operator}-{item.brand}-{item.city}: {item.message}")


def print_import_summary(ctx: ImportContext) -> None:
    print("月度返点源数据导入预览")
    print("=" * 40)
    print(f"import_id: {ctx.import_id}")
    print(f"import_type: {ctx.import_type}")
    print(f"月份: {ctx.month_label}")
    print(f"规则: {ctx.rules_path}")
    for role, path in ctx.source_files.items():
        print(f"源文件[{role}]: {path}")
    billing_rows = sum(row_count for role, _path, _sheet, row_count in ctx.source_sheet_counts if role == "billing")
    print(f"账单原始行: {billing_rows} 行")
    print(f"账单原始聚合: {len(ctx.billing_import_agg)} 个 日期+品牌+城市+tr类型+流量渠道 组合")
    print(f"过程考核原始行: {sum(len(rows) for rows in ctx.process_rows.values())} 行")
    print(f"开城奖励原始行: {len(ctx.open_city_rows)} 行")
    print(f"目标拆解原始行: {len(ctx.target_raw_rows)} 行")
    print(f"源文件sheet: {len(ctx.source_sheet_counts)} 个")


def import_sources(args: argparse.Namespace) -> int:
    if not args.dry_run and not args.confirmed:
        raise RuntimeError("写入源数据 import 批次必须显式传 --confirmed；预览请传 --dry-run")
    ctx = build_import_context(args)
    print_import_summary(ctx)
    if args.dry_run:
        return 0
    sync_import_database(ctx)
    print(f"[ok] 已导入 PostgreSQL: import_id={ctx.import_id}")
    return 0


def import_process(args: argparse.Namespace) -> int:
    if not args.dry_run and not args.confirmed:
        raise RuntimeError("写入过程数据 import 批次必须显式传 --confirmed；预览请传 --dry-run")
    ctx = build_process_import_context(args)
    print_import_summary(ctx)
    if args.dry_run:
        return 0
    sync_import_database(ctx)
    print(f"[ok] 已导入 PostgreSQL: import_id={ctx.import_id}")
    return 0


def calculate(args: argparse.Namespace) -> int:
    if not args.dry_run and not args.confirmed:
        raise RuntimeError("写 Excel 必须显式传 --confirmed；预览请传 --dry-run")
    if args.dry_run and args.sync_db:
        raise RuntimeError("--dry-run 不能和 --sync-db 同时使用")
    if args.sync_db and not args.confirmed:
        raise RuntimeError("同步数据库必须显式传 --confirmed --sync-db")

    ctx = build_context(args)
    output_path = Path(args.output).expanduser() if args.output else ctx.work_dir / "输出" / f"{ctx.month_label}月度返点-输出.xlsx"
    if not output_path.is_absolute():
        output_path = PROJECT_ROOT / output_path
    print_summary(ctx, None if args.dry_run else output_path)
    if args.dry_run:
        return 0
    write_output_workbook(output_path, ctx.results, ctx.rules.get("process") or {}, ctx.validations)
    print(f"[ok] 已写入 Excel: {output_path}")
    if args.sync_db:
        sync_database(ctx)
        print(f"[ok] 已同步 PostgreSQL: run_id={ctx.run_id}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="月度返点计算 Skill")
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_parser = subparsers.add_parser("init-rules", help="生成当月规则模板")
    init_parser.add_argument("--month", required=True, help="月份 YYYY-MM")
    init_parser.add_argument("--work-dir", default="", help="月度工作目录")
    init_parser.add_argument("--contacts", default="雷维亮", help="对接人，逗号分隔")
    init_parser.add_argument("--overwrite", action="store_true", help="覆盖已有规则文件")
    init_parser.add_argument("--dry-run", action="store_true", help="只预览，不写入")
    init_parser.set_defaults(func=init_rules)

    import_parser = subparsers.add_parser("import-sources", help="把当月源 Excel 导入 PostgreSQL import 批次")
    import_parser.add_argument("--month", required=True, help="月份 YYYY-MM")
    import_parser.add_argument("--work-dir", default="", help="月度工作目录")
    import_parser.add_argument("--dry-run", action="store_true", help="只预览，不写数据库")
    import_parser.add_argument("--confirmed", action="store_true", help="确认写入 PostgreSQL import 批次")
    import_parser.set_defaults(func=import_sources)

    process_import_parser = subparsers.add_parser("import-process", help="只把当月过程考核 Excel 导入 PostgreSQL process_only 批次")
    process_import_parser.add_argument("--month", required=True, help="月份 YYYY-MM")
    process_import_parser.add_argument("--work-dir", default="", help="月度工作目录")
    process_import_parser.add_argument("--process-pattern", default="*LX品牌过程考核数据*.xlsx", help="过程考核源文件匹配规则")
    process_import_parser.add_argument("--dry-run", action="store_true", help="只预览，不写数据库")
    process_import_parser.add_argument("--confirmed", action="store_true", help="确认写入 PostgreSQL process_only import 批次")
    process_import_parser.set_defaults(func=import_process)

    calc_parser = subparsers.add_parser("calculate", help="计算月度返点")
    calc_parser.add_argument("--month", required=True, help="月份 YYYY-MM")
    calc_parser.add_argument("--work-dir", default="", help="月度工作目录")
    calc_parser.add_argument("--contacts", default="", help="对接人，逗号分隔；不传则读取 rules.yaml")
    calc_parser.add_argument("--exclude-operator", action="append", default=[], help="排除运营主体，可重复传")
    calc_parser.add_argument("--source", choices=["db", "excel"], default="db", help="计算数据来源，默认读取 PostgreSQL import 批次")
    calc_parser.add_argument("--import-id", default="", help="指定 PostgreSQL import_id；不传则读取当月最新 confirmed import")
    calc_parser.add_argument("--output", default="", help="输出 xlsx 路径")
    calc_parser.add_argument("--dry-run", action="store_true", help="只预览，不写 Excel/数据库")
    calc_parser.add_argument("--confirmed", action="store_true", help="确认写 Excel")
    calc_parser.add_argument("--sync-db", action="store_true", help="同步本地 PostgreSQL")
    calc_parser.set_defaults(func=calculate)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return args.func(args)
    except Exception as exc:
        print(f"[error] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
