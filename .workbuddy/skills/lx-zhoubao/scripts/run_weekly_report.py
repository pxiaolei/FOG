#!/usr/bin/env python3
"""Generate hhdata weekly report workbooks."""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _find_project_root() -> Path:
    for candidate in [Path(__file__).resolve(), *Path(__file__).resolve().parents]:
        if (candidate / ".workbuddy").is_dir() and (candidate / "config").is_dir():
            return candidate
    return Path(__file__).resolve().parents[4]


PROJECT_ROOT = _find_project_root()
SKILLS_DIR = PROJECT_ROOT / ".workbuddy" / "skills"
LX_SHUJUKU_SCRIPTS = SKILLS_DIR / "lx_shujuku" / "scripts"
for path in (SKILLS_DIR, LX_SHUJUKU_SCRIPTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from lxx_share.database import DatabaseConnector  # noqa: E402


MISSING_DISPLAY = " / "
DIM_SLASH = "/"
DEFAULT_TEMPLATE = PROJECT_ROOT / "workspace/03数据报表/周报/hhdata周报模版.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "workspace/03数据报表/周报"
WEEKLY_METRICS_TABLE = "hhdata__agg_weekly_metrics"
DAILY_BRAND_CITY_METRICS_TABLE = "hhdata__agg_daily_brand_city_metrics"

SUM_FIELDS = [
    "passenger_order_count",
    "broadcast_orders",
    "match_count",
    "response_count",
    "completed_order_count",
    "cancelled_before_answer",
    "cancelled_by_passenger",
    "cancelled_by_driver",
    "online_duration_hour",
    "online_driver_count",
    "completed_drivers",
    "peak_online_driver_count",
    "peak_valid_driver_count",
    "valid_driver_count",
    "approved_drivers",
    "first_online_drivers",
    "first_completion_driver_count",
    "gmv",
    "total_b_subsidy",
    "merchant_b_subsidy",
    "total_commission",
    "brand_commission",
    "card_merchant_income",
]

DIMENSION_CODE_BY_NAME = {
    "大盘维度": "all",
    "主体纬度": "operator",
    "品牌维度": "brand",
    "品牌城市维度": "brand_city",
    "城市维度": "city",
}
DIMENSION_NAME_BY_CODE = {value: key for key, value in DIMENSION_CODE_BY_NAME.items()}
DIMENSION_ORDER_CODES = ["all", "operator", "brand", "brand_city", "city"]
WEEKLY_AGG_COLUMNS = [
    "week_start",
    "week_end",
    "day_count",
    "dimension_code",
    "dimension_name",
    "contact_person",
    "operator_name",
    "brand_name",
    "city_name",
    "mapping_status",
    "source_row_count",
    *SUM_FIELDS,
]
DAILY_AGG_COLUMNS = [
    "metric_date",
    "brand_name",
    "city_name",
    "source_row_count",
    *SUM_FIELDS,
]


@dataclass(frozen=True)
class Period:
    start: date
    end: date

    @property
    def day_count(self) -> int:
        return (self.end - self.start).days + 1

    @property
    def label(self) -> str:
        return f"{self.start:%m%d}-{self.end:%m%d}"

    @property
    def file_label(self) -> str:
        return f"{self.start:%Y%m%d}-{self.end:%Y%m%d}"


@dataclass(frozen=True)
class Periods:
    previous: Period
    current: Period


@dataclass(frozen=True)
class MetricSpec:
    label: str
    columns: tuple[int, int, int]
    metric_type: str
    expr: Callable[[str, dict[str, str]], str]


def parse_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"日期格式必须是 YYYY-MM-DD: {value}") from exc


def latest_complete_week(latest: date) -> Period:
    days_since_sunday = (latest.weekday() + 1) % 7
    sunday = latest - timedelta(days=days_since_sunday)
    return Period(sunday - timedelta(days=6), sunday)


def is_complete_natural_week(period: Period) -> bool:
    return period.day_count == 7 and period.start.weekday() == 0 and period.end.weekday() == 6


def resolve_periods(
    start: date | None,
    end: date | None,
    latest_data_date: date | None,
) -> Periods:
    if (start is None) != (end is None):
        raise ValueError("--start 和 --end 必须同时传入")
    if start and end:
        if end < start:
            raise ValueError("--end 不能早于 --start")
        current = Period(start, end)
    else:
        if latest_data_date is None:
            raise ValueError("数据库没有可用 hhdata 日期，无法推断默认周报周期")
        current = latest_complete_week(latest_data_date)
    previous = Period(current.start - timedelta(days=7), current.end - timedelta(days=7))
    return Periods(previous=previous, current=current)


def period_date_range(periods: Periods) -> tuple[date, date]:
    return periods.previous.start, periods.current.end


def fmt_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)[:10]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def make_key(*parts: str) -> tuple[str, ...]:
    return tuple(clean_text(part) for part in parts)


def load_latest_data_date(db: DatabaseConnector) -> date | None:
    df = db.execute("SELECT MAX(date_day) AS latest_date FROM hhdata__fact_daily_metrics")
    if df.empty or pd.isna(df.iloc[0]["latest_date"]):
        return None
    return datetime.strptime(fmt_date(df.iloc[0]["latest_date"]), "%Y-%m-%d").date()


def load_hhdata(db: DatabaseConnector, periods: Periods) -> pd.DataFrame:
    fields = ", ".join(f"f.{field}" for field in SUM_FIELDS)
    sql = f"""
        SELECT f.date_day AS date, c.city_name, b.brand_name, {fields}
        FROM hhdata__fact_daily_metrics f
        JOIN mabiao__dim_cities c ON f.city_id = c.city_id
        JOIN mabiao__dim_brands b ON f.brand_id = b.brand_id
        WHERE (f.date_day >= %s AND f.date_day <= %s)
           OR (f.date_day >= %s AND f.date_day <= %s)
    """
    df = db.execute(
        sql,
        [
            periods.previous.start.isoformat(),
            periods.previous.end.isoformat(),
            periods.current.start.isoformat(),
            periods.current.end.isoformat(),
        ],
    )
    if df.empty:
        return df
    df = df.copy()
    df["date"] = df["date"].map(lambda value: datetime.strptime(fmt_date(value), "%Y-%m-%d").date())
    df["period"] = ""
    df.loc[(df["date"] >= periods.previous.start) & (df["date"] <= periods.previous.end), "period"] = "previous"
    df.loc[(df["date"] >= periods.current.start) & (df["date"] <= periods.current.end), "period"] = "current"
    df = df[df["period"].isin(["previous", "current"])].copy()
    for field in SUM_FIELDS:
        df[field] = pd.to_numeric(df[field], errors="coerce").fillna(0)
    df["brand_name"] = df["brand_name"].map(clean_text)
    df["city_name"] = df["city_name"].map(clean_text)
    return df


def load_operator_brand_rows(limit: int = 1000) -> list[dict[str, str]]:
    from lx_shujuku import create_client

    client = create_client()
    return client.get_operator_brands(limit=limit)


def build_mapping(rows: list[dict[str, str]]) -> dict[str, Any]:
    pair_map: dict[tuple[str, str], dict[str, set[str]]] = defaultdict(
        lambda: {"operators": set(), "persons": set()}
    )
    operator_persons: dict[str, set[str]] = defaultdict(set)
    all_pairs: set[tuple[str, str]] = set()
    all_brands: set[str] = set()
    all_cities: set[str] = set()
    all_operators: set[str] = set()

    for row in rows:
        operator = clean_text(row.get("operator") or row.get("运营主体"))
        brand = clean_text(row.get("brand") or row.get("品牌"))
        city = clean_text(row.get("city") or row.get("城市"))
        person = clean_text(row.get("contact_person") or row.get("对接人"))
        if not operator or not brand or not city:
            continue
        key = (brand, city)
        all_pairs.add(key)
        all_brands.add(brand)
        all_cities.add(city)
        all_operators.add(operator)
        pair_map[key]["operators"].add(operator)
        if person:
            pair_map[key]["persons"].add(person)
            operator_persons[operator].add(person)

    return {
        "pair_map": pair_map,
        "operator_persons": operator_persons,
        "all_pairs": all_pairs,
        "all_brands": all_brands,
        "all_cities": all_cities,
        "all_operators": all_operators,
    }


def join_values(values: set[str]) -> str:
    cleaned = sorted(v for v in values if clean_text(v))
    return "、".join(cleaned) if cleaned else DIM_SLASH


def choose_one(values: set[str]) -> str:
    cleaned = sorted(v for v in values if clean_text(v))
    return cleaned[0] if cleaned else ""


def add_operator_columns(df: pd.DataFrame, mapping: dict[str, Any], gaps: list[dict[str, Any]]) -> pd.DataFrame:
    if df.empty:
        return df
    pair_map = mapping["pair_map"]
    operators = []
    persons = []
    for _, row in df.iterrows():
        key = (row["brand_name"], row["city_name"])
        owner = pair_map.get(key)
        if not owner:
            operators.append("")
            persons.append("")
            continue
        if len(owner["operators"]) > 1:
            gaps.append(
                {
                    "type": "mapping_conflict",
                    "dimension": "品牌城市维度",
                    "brand": key[0],
                    "city": key[1],
                    "operators": sorted(owner["operators"]),
                    "contact_persons": sorted(owner["persons"]),
                }
            )
        operator = choose_one(owner["operators"])
        operators.append(operator)
        persons.append(join_values(mapping["operator_persons"].get(operator, set())))
    result = df.copy()
    result["operator_name"] = operators
    result["contact_person"] = persons
    return result


def aggregate_df(df: pd.DataFrame, group_cols: list[str]) -> dict[tuple[str, ...], dict[str, dict[str, float]]]:
    result: dict[tuple[str, ...], dict[str, dict[str, float]]] = defaultdict(dict)
    if df.empty:
        return result
    for period_name in ["previous", "current"]:
        part = df[df["period"] == period_name]
        if part.empty:
            continue
        if group_cols:
            grouped = part.groupby(group_cols, dropna=False)
            sums = grouped[SUM_FIELDS].sum()
            counts = grouped.size()
            for key, values in sums.iterrows():
                if not isinstance(key, tuple):
                    key = (key,)
                record = {field: float(values[field]) for field in SUM_FIELDS}
                record["row_count"] = int(counts.loc[key if len(key) > 1 else key[0]])
                result[tuple(clean_text(x) for x in key)][period_name] = record
        else:
            record = {field: float(part[field].sum()) for field in SUM_FIELDS}
            record["row_count"] = int(len(part))
            result[()][period_name] = record
    return result


def metric_expr(kind: str, numerator: str = "", denominator: str = "", extra: str = "") -> Callable[[str, dict[str, str]], str]:
    def expr(prefix: str, cols: dict[str, str]) -> str:
        if kind == "daily":
            return f"{cols[prefix + numerator]}/{cols[prefix + 'day_count']}"
        if kind == "sum":
            return cols[prefix + numerator]
        if kind == "ratio":
            return f"{cols[prefix + numerator]}/{cols[prefix + denominator]}"
        if kind == "sum_ratio":
            left, right = extra.split("+")
            return f"({cols[prefix + left]}+{cols[prefix + right]})/{cols[prefix + denominator]}"
        if kind == "margin":
            return (
                f"{cols[prefix + 'brand_commission']}/{cols[prefix + 'gmv']}"
                f"+{cols[prefix + 'card_merchant_income']}/{cols[prefix + 'gmv']}"
                f"-{cols[prefix + 'merchant_b_subsidy']}/{cols[prefix + 'gmv']}-0.01"
            )
        raise ValueError(f"未知指标类型: {kind}")

    return expr


def metric_requirements(label: str) -> tuple[list[str], list[str]]:
    specs = {
        "完单": (["completed_order_count"], []),
        "发单": (["passenger_order_count"], []),
        "在线司机数": (["online_driver_count"], []),
        "TSH": (["online_duration_hour"], []),
        "TR": (["brand_commission"], ["gmv"]),
        "线上毛利": (["brand_commission", "card_merchant_income", "merchant_b_subsidy"], ["gmv"]),
        "卡劵收入": (["card_merchant_income"], ["gmv"]),
        "B补率": (["merchant_b_subsidy"], ["gmv"]),
        "单价": (["gmv"], ["completed_order_count"]),
        "完单司机数": (["completed_drivers"], []),
        "首次完单司机数": (["first_completion_driver_count"], []),
        "人均完单": (["completed_order_count"], ["completed_drivers"]),
        "匹配": (["match_count"], []),
        "匹配/发单": (["match_count"], ["passenger_order_count"]),
        "应答": (["response_count"], []),
        "应答率（应答/发单": (["response_count"], ["passenger_order_count"]),
        "应答匹配PK率": (["response_count"], ["match_count"]),
        "成交率": (["completed_order_count"], ["passenger_order_count"]),
        "司机取消率": (["cancelled_by_driver"], ["response_count"]),
        "乘客取消率": (["cancelled_by_passenger"], ["response_count"]),
        "司乘取消率": (["cancelled_by_driver", "cancelled_by_passenger"], ["response_count"]),
        "完单/应答（成交率": (["completed_order_count"], ["response_count"]),
        "gmv": (["gmv"], []),
        "周期总gmv": (["gmv"], []),
    }
    return specs[label]


METRICS: list[MetricSpec] = [
    MetricSpec("完单", (6, 7, 8), "volume", metric_expr("daily", "completed_order_count")),
    MetricSpec("发单", (9, 10, 11), "volume", metric_expr("daily", "passenger_order_count")),
    MetricSpec("在线司机数", (12, 13, 14), "volume", metric_expr("daily", "online_driver_count")),
    MetricSpec("TSH", (15, 16, 17), "volume", metric_expr("daily", "online_duration_hour")),
    MetricSpec("TR", (18, 19, 20), "rate", metric_expr("ratio", "brand_commission", "gmv")),
    MetricSpec("线上毛利", (21, 22, 23), "rate", metric_expr("margin")),
    MetricSpec("卡劵收入", (24, 25, 26), "rate", metric_expr("ratio", "card_merchant_income", "gmv")),
    MetricSpec("B补率", (27, 28, 29), "rate", metric_expr("ratio", "merchant_b_subsidy", "gmv")),
    MetricSpec("单价", (30, 31, 32), "volume", metric_expr("ratio", "gmv", "completed_order_count")),
    MetricSpec("完单司机数", (33, 34, 35), "volume", metric_expr("daily", "completed_drivers")),
    MetricSpec("首次完单司机数", (36, 37, 38), "volume", metric_expr("daily", "first_completion_driver_count")),
    MetricSpec("人均完单", (39, 40, 41), "volume", metric_expr("ratio", "completed_order_count", "completed_drivers")),
    MetricSpec("匹配", (42, 43, 44), "volume", metric_expr("daily", "match_count")),
    MetricSpec("匹配/发单", (45, 46, 47), "rate", metric_expr("ratio", "match_count", "passenger_order_count")),
    MetricSpec("应答", (48, 49, 50), "volume", metric_expr("daily", "response_count")),
    MetricSpec("应答率（应答/发单", (51, 52, 53), "rate", metric_expr("ratio", "response_count", "passenger_order_count")),
    MetricSpec("应答匹配PK率", (54, 55, 56), "rate", metric_expr("ratio", "response_count", "match_count")),
    MetricSpec("成交率", (57, 58, 59), "rate", metric_expr("ratio", "completed_order_count", "passenger_order_count")),
    MetricSpec("司机取消率", (60, 61, 62), "rate", metric_expr("ratio", "cancelled_by_driver", "response_count")),
    MetricSpec("乘客取消率", (63, 64, 65), "rate", metric_expr("ratio", "cancelled_by_passenger", "response_count")),
    MetricSpec("司乘取消率", (66, 67, 68), "rate", metric_expr("sum_ratio", denominator="response_count", extra="cancelled_by_driver+cancelled_by_passenger")),
    MetricSpec("完单/应答（成交率", (69, 70, 71), "rate", metric_expr("ratio", "completed_order_count", "response_count")),
    MetricSpec("gmv", (72, 73, 74), "volume", metric_expr("daily", "gmv")),
    MetricSpec("周期总gmv", (75, 76, 77), "volume", metric_expr("sum", "gmv")),
]


def default_source(periods: Periods) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for prefix, period in (("previous_", periods.previous), ("current_", periods.current)):
        result[prefix + "row_count"] = 0
        result[prefix + "day_count"] = period.day_count
        for field in SUM_FIELDS:
            result[prefix + field] = 0.0
    return result


def fill_period_source(source: dict[str, Any], period_name: str, values: dict[str, float]) -> None:
    prefix = "previous_" if period_name == "previous" else "current_"
    source[prefix + "row_count"] = int(values.get("row_count", 0))
    for field in SUM_FIELDS:
        source[prefix + field] = float(values.get(field, 0) or 0)


def current_sort_value(agg: dict[tuple[str, ...], dict[str, dict[str, float]]], key: tuple[str, ...]) -> float:
    return float(agg.get(key, {}).get("current", {}).get("completed_order_count", 0) or 0)


def period_sort_value(agg: dict[tuple[str, ...], dict[str, dict[str, float]]], key: tuple[str, ...], period_name: str) -> float:
    return float(agg.get(key, {}).get(period_name, {}).get("completed_order_count", 0) or 0)


def empty_weekly_source(period: Period) -> dict[str, Any]:
    source: dict[str, Any] = {"row_count": 0, "day_count": period.day_count}
    for field in SUM_FIELDS:
        source[field] = 0.0
    return source


def weekly_source_from_agg(
    agg: dict[tuple[str, ...], dict[str, dict[str, float]]],
    key: tuple[str, ...],
    period: Period,
    period_name: str,
) -> dict[str, Any]:
    source = empty_weekly_source(period)
    values = agg.get(key, {}).get(period_name, {})
    source["row_count"] = int(values.get("row_count", 0) or 0)
    for field in SUM_FIELDS:
        source[field] = float(values.get(field, 0) or 0)
    return source


def build_weekly_aggregate_rows(
    df: pd.DataFrame,
    mapping: dict[str, Any],
    period: Period,
    period_name: str = "current",
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    gaps: list[dict[str, Any]] = []
    df = add_operator_columns(df, mapping, gaps)

    agg_all = aggregate_df(df, [])
    agg_brand_city = aggregate_df(df, ["brand_name", "city_name"])
    agg_brand = aggregate_df(df, ["brand_name"])
    agg_city = aggregate_df(df, ["city_name"])
    agg_operator = aggregate_df(df[df.get("operator_name", "") != ""], ["operator_name"]) if not df.empty else {}

    data_pairs = set(agg_brand_city.keys())
    data_brands = {key[0] for key in agg_brand.keys()}
    data_cities = {key[0] for key in agg_city.keys()}
    rows: list[dict[str, Any]] = []

    def append_row(
        contact_person: str,
        dimension: str,
        operator: str,
        brand: str,
        city: str,
        source: dict[str, Any],
        mapping_status: str,
    ) -> None:
        rows.append(
            {
                "contact_person": contact_person or DIM_SLASH,
                "dimension": dimension,
                "dimension_code": DIMENSION_CODE_BY_NAME[dimension],
                "operator": operator or DIM_SLASH,
                "brand": brand or "all",
                "city": city or "all",
                "mapping_status": mapping_status,
                "source": source,
            }
        )

    append_row(
        DIM_SLASH,
        "大盘维度",
        DIM_SLASH,
        "all",
        "all",
        weekly_source_from_agg(agg_all, (), period, period_name),
        "not_applicable",
    )

    operator_keys = sorted(
        mapping["all_operators"],
        key=lambda operator: (-period_sort_value(agg_operator, (operator,), period_name), operator),
    )
    for operator in operator_keys:
        person = join_values(mapping["operator_persons"].get(operator, set()))
        append_row(
            person,
            "主体纬度",
            operator,
            "all",
            "all",
            weekly_source_from_agg(agg_operator, (operator,), period, period_name),
            "not_applicable",
        )

    brand_keys = sorted(
        mapping["all_brands"] | data_brands,
        key=lambda brand: (-period_sort_value(agg_brand, (brand,), period_name), brand),
    )
    for brand in brand_keys:
        append_row(
            DIM_SLASH,
            "品牌维度",
            DIM_SLASH,
            brand,
            "all",
            weekly_source_from_agg(agg_brand, (brand,), period, period_name),
            "not_applicable",
        )

    brand_city_keys = sorted(
        mapping["all_pairs"] | data_pairs,
        key=lambda key: (-period_sort_value(agg_brand_city, key, period_name), key[0], key[1]),
    )
    for brand, city in brand_city_keys:
        owner = mapping["pair_map"].get((brand, city))
        if owner:
            operator = join_values(owner["operators"])
            person = join_values(owner["persons"])
            mapping_status = "conflict" if len(owner["operators"]) > 1 else "matched"
        else:
            operator = DIM_SLASH
            person = DIM_SLASH
            mapping_status = "unmapped"
            gaps.append(
                {
                    "type": "unmapped_brand_city",
                    "dimension": "品牌城市维度",
                    "dimension_code": "brand_city",
                    "brand": brand,
                    "city": city,
                    "reason": "operator_brand 未匹配",
                }
            )
        append_row(
            person,
            "品牌城市维度",
            operator,
            brand,
            city,
            weekly_source_from_agg(agg_brand_city, (brand, city), period, period_name),
            mapping_status,
        )

    city_keys = sorted(
        mapping["all_cities"] | data_cities,
        key=lambda city: (-period_sort_value(agg_city, (city,), period_name), city),
    )
    for city in city_keys:
        append_row(
            DIM_SLASH,
            "城市维度",
            DIM_SLASH,
            "all",
            city,
            weekly_source_from_agg(agg_city, (city,), period, period_name),
            "not_applicable",
        )

    return rows, gaps


def table_exists(db: DatabaseConnector, table_name: str) -> bool:
    value = db.execute_scalar(
        """
        SELECT COUNT(*)
        FROM information_schema.tables
        WHERE table_schema = DATABASE() AND table_name = %s
        """,
        [table_name],
    )
    return bool(value)


def load_weekly_agg(db: DatabaseConnector, periods: Periods) -> pd.DataFrame:
    columns = ", ".join(WEEKLY_AGG_COLUMNS)
    sql = f"""
        SELECT {columns}
        FROM {WEEKLY_METRICS_TABLE}
        WHERE week_start IN (%s, %s)
    """
    df = db.execute(sql, [periods.previous.start.isoformat(), periods.current.start.isoformat()])
    if df.empty:
        return df
    df = df.copy()
    df["week_start"] = df["week_start"].map(lambda value: datetime.strptime(fmt_date(value), "%Y-%m-%d").date())
    df["week_end"] = df["week_end"].map(lambda value: datetime.strptime(fmt_date(value), "%Y-%m-%d").date())
    for field in ["day_count", "source_row_count"]:
        df[field] = pd.to_numeric(df[field], errors="coerce").fillna(0).astype(int)
    for field in SUM_FIELDS:
        df[field] = pd.to_numeric(df[field], errors="coerce").fillna(0)
    for field in ["dimension_code", "dimension_name", "contact_person", "operator_name", "brand_name", "city_name", "mapping_status"]:
        df[field] = df[field].map(clean_text)
    return df


def weekly_agg_has_periods(df: pd.DataFrame, periods: Periods) -> bool:
    if df.empty:
        return False
    starts = set(df["week_start"])
    return periods.previous.start in starts and periods.current.start in starts


def date_set(start: date, end: date) -> set[date]:
    return {start + timedelta(days=offset) for offset in range((end - start).days + 1)}


def load_daily_agg(db: DatabaseConnector, periods: Periods) -> pd.DataFrame:
    columns = ", ".join(DAILY_AGG_COLUMNS)
    sql = f"""
        SELECT {columns}
        FROM {DAILY_BRAND_CITY_METRICS_TABLE}
        WHERE (metric_date >= %s AND metric_date <= %s)
           OR (metric_date >= %s AND metric_date <= %s)
    """
    df = db.execute(
        sql,
        [
            periods.previous.start.isoformat(),
            periods.previous.end.isoformat(),
            periods.current.start.isoformat(),
            periods.current.end.isoformat(),
        ],
    )
    if df.empty:
        return df
    df = df.copy()
    df["date"] = df["metric_date"].map(lambda value: datetime.strptime(fmt_date(value), "%Y-%m-%d").date())
    df["period"] = ""
    df.loc[(df["date"] >= periods.previous.start) & (df["date"] <= periods.previous.end), "period"] = "previous"
    df.loc[(df["date"] >= periods.current.start) & (df["date"] <= periods.current.end), "period"] = "current"
    df = df[df["period"].isin(["previous", "current"])].copy()
    for field in ["source_row_count", *SUM_FIELDS]:
        df[field] = pd.to_numeric(df[field], errors="coerce").fillna(0)
    df["brand_name"] = df["brand_name"].map(clean_text)
    df["city_name"] = df["city_name"].map(clean_text)
    return df


def daily_agg_has_periods(df: pd.DataFrame, periods: Periods) -> bool:
    if df.empty:
        return False
    dates = set(df["date"])
    expected = date_set(periods.previous.start, periods.previous.end) | date_set(periods.current.start, periods.current.end)
    return expected.issubset(dates)


def prefixed_source_from_weekly_row(row: pd.Series) -> dict[str, Any]:
    values: dict[str, Any] = {"row_count": int(row.get("source_row_count", 0) or 0)}
    for field in SUM_FIELDS:
        values[field] = float(row.get(field, 0) or 0)
    return values


def build_report_rows_from_weekly_agg(
    df: pd.DataFrame,
    periods: Periods,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    gaps: list[dict[str, Any]] = []
    period_by_start = {
        periods.previous.start: "previous",
        periods.current.start: "current",
    }
    records: dict[tuple[str, str, str, str], dict[str, Any]] = {}

    for _, row in df.iterrows():
        period_name = period_by_start.get(row["week_start"])
        if not period_name:
            continue
        dimension_code = row["dimension_code"]
        key = (
            dimension_code,
            clean_text(row["operator_name"]) or DIM_SLASH,
            clean_text(row["brand_name"]) or "all",
            clean_text(row["city_name"]) or "all",
        )
        record = records.setdefault(
            key,
            {
                "dimension_code": dimension_code,
                "dimension": row["dimension_name"] or DIMENSION_NAME_BY_CODE.get(dimension_code, dimension_code),
                "contact_person": DIM_SLASH,
                "operator": key[1],
                "brand": key[2],
                "city": key[3],
                "source": default_source(periods),
                "mapping_statuses": set(),
            },
        )
        if period_name == "current" or record["contact_person"] == DIM_SLASH:
            record["contact_person"] = row["contact_person"] or DIM_SLASH
            record["operator"] = key[1]
            record["brand"] = key[2]
            record["city"] = key[3]
        record["mapping_statuses"].add(row["mapping_status"])
        fill_period_source(record["source"], period_name, prefixed_source_from_weekly_row(row))

    rows = list(records.values())
    for row in rows:
        if row["dimension_code"] == "brand_city" and "unmapped" in row.pop("mapping_statuses", set()):
            gaps.append(
                {
                    "type": "unmapped_brand_city",
                    "dimension": row["dimension"],
                    "dimension_code": row["dimension_code"],
                    "brand": row["brand"],
                    "city": row["city"],
                    "reason": "聚合表 mapping_status=unmapped",
                }
            )
        else:
            row.pop("mapping_statuses", None)

    order = {code: idx for idx, code in enumerate(DIMENSION_ORDER_CODES)}
    rows.sort(
        key=lambda row: (
            order.get(row["dimension_code"], 99),
            -float(row["source"].get("current_completed_order_count", 0) or 0),
            row["operator"],
            row["brand"],
            row["city"],
        )
    )
    add_calculation_gaps(rows, gaps)
    return rows, gaps


def build_report_rows(
    df: pd.DataFrame,
    mapping: dict[str, Any],
    periods: Periods,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    gaps: list[dict[str, Any]] = []
    df = add_operator_columns(df, mapping, gaps)

    agg_all = aggregate_df(df, [])
    agg_brand_city = aggregate_df(df, ["brand_name", "city_name"])
    agg_brand = aggregate_df(df, ["brand_name"])
    agg_city = aggregate_df(df, ["city_name"])
    agg_operator = aggregate_df(df[df.get("operator_name", "") != ""], ["operator_name"]) if not df.empty else {}

    data_pairs = set(agg_brand_city.keys())
    data_brands = {key[0] for key in agg_brand.keys()}
    data_cities = {key[0] for key in agg_city.keys()}
    rows: list[dict[str, Any]] = []

    def make_source(agg: dict[tuple[str, ...], dict[str, dict[str, float]]], key: tuple[str, ...]) -> dict[str, Any]:
        source = default_source(periods)
        for period_name, values in agg.get(key, {}).items():
            fill_period_source(source, period_name, values)
        return source

    def append_row(contact_person: str, dimension: str, operator: str, brand: str, city: str, source: dict[str, Any]) -> None:
        rows.append(
            {
                "contact_person": contact_person or DIM_SLASH,
                "dimension": dimension,
                "operator": operator or DIM_SLASH,
                "brand": brand or "all",
                "city": city or "all",
                "source": source,
            }
        )

    append_row(DIM_SLASH, "大盘维度", DIM_SLASH, "all", "all", make_source(agg_all, ()))

    operator_keys = sorted(
        mapping["all_operators"],
        key=lambda operator: (-current_sort_value(agg_operator, (operator,)), operator),
    )
    for operator in operator_keys:
        person = join_values(mapping["operator_persons"].get(operator, set()))
        append_row(person, "主体纬度", operator, "all", "all", make_source(agg_operator, (operator,)))

    brand_keys = sorted(
        mapping["all_brands"] | data_brands,
        key=lambda brand: (-current_sort_value(agg_brand, (brand,)), brand),
    )
    for brand in brand_keys:
        append_row(DIM_SLASH, "品牌维度", DIM_SLASH, brand, "all", make_source(agg_brand, (brand,)))

    brand_city_keys = sorted(
        mapping["all_pairs"] | data_pairs,
        key=lambda key: (-current_sort_value(agg_brand_city, key), key[0], key[1]),
    )
    for brand, city in brand_city_keys:
        owner = mapping["pair_map"].get((brand, city))
        if owner:
            operator = join_values(owner["operators"])
            person = join_values(owner["persons"])
        else:
            operator = DIM_SLASH
            person = DIM_SLASH
            gaps.append(
                {
                    "type": "unmapped_brand_city",
                    "dimension": "品牌城市维度",
                    "brand": brand,
                    "city": city,
                }
            )
        append_row(person, "品牌城市维度", operator, brand, city, make_source(agg_brand_city, (brand, city)))

    city_keys = sorted(
        mapping["all_cities"] | data_cities,
        key=lambda city: (-current_sort_value(agg_city, (city,)), city),
    )
    for city in city_keys:
        append_row(DIM_SLASH, "城市维度", DIM_SLASH, "all", city, make_source(agg_city, (city,)))

    add_calculation_gaps(rows, gaps)
    return rows, gaps


def has_period_data(row: dict[str, Any], prefix: str) -> bool:
    return int(row["source"].get(prefix + "row_count", 0) or 0) > 0


def can_compute(row: dict[str, Any], prefix: str, metric: MetricSpec) -> tuple[bool, str]:
    source = row["source"]
    if not has_period_data(row, prefix):
        return False, "缺周期数据"
    _, denominators = metric_requirements(metric.label)
    for denominator in denominators:
        if float(source.get(prefix + denominator, 0) or 0) == 0:
            return False, f"分母为0: {denominator}"
    return True, ""


def metric_value(row: dict[str, Any], prefix: str, metric: MetricSpec) -> float | None:
    ok, _ = can_compute(row, prefix, metric)
    if not ok:
        return None
    source = row["source"]
    if metric.label == "完单":
        return source[prefix + "completed_order_count"] / source[prefix + "day_count"]
    if metric.label == "发单":
        return source[prefix + "passenger_order_count"] / source[prefix + "day_count"]
    if metric.label == "在线司机数":
        return source[prefix + "online_driver_count"] / source[prefix + "day_count"]
    if metric.label == "TSH":
        return source[prefix + "online_duration_hour"] / source[prefix + "day_count"]
    if metric.label == "TR":
        return source[prefix + "brand_commission"] / source[prefix + "gmv"]
    if metric.label == "线上毛利":
        return (
            source[prefix + "brand_commission"] / source[prefix + "gmv"]
            + source[prefix + "card_merchant_income"] / source[prefix + "gmv"]
            - source[prefix + "merchant_b_subsidy"] / source[prefix + "gmv"]
            - 0.01
        )
    if metric.label == "卡劵收入":
        return source[prefix + "card_merchant_income"] / source[prefix + "gmv"]
    if metric.label == "B补率":
        return source[prefix + "merchant_b_subsidy"] / source[prefix + "gmv"]
    if metric.label == "单价":
        return source[prefix + "gmv"] / source[prefix + "completed_order_count"]
    if metric.label == "完单司机数":
        return source[prefix + "completed_drivers"] / source[prefix + "day_count"]
    if metric.label == "首次完单司机数":
        return source[prefix + "first_completion_driver_count"] / source[prefix + "day_count"]
    if metric.label == "人均完单":
        return source[prefix + "completed_order_count"] / source[prefix + "completed_drivers"]
    if metric.label == "匹配":
        return source[prefix + "match_count"] / source[prefix + "day_count"]
    if metric.label == "匹配/发单":
        return source[prefix + "match_count"] / source[prefix + "passenger_order_count"]
    if metric.label == "应答":
        return source[prefix + "response_count"] / source[prefix + "day_count"]
    if metric.label == "应答率（应答/发单":
        return source[prefix + "response_count"] / source[prefix + "passenger_order_count"]
    if metric.label == "应答匹配PK率":
        return source[prefix + "response_count"] / source[prefix + "match_count"]
    if metric.label == "成交率":
        return source[prefix + "completed_order_count"] / source[prefix + "passenger_order_count"]
    if metric.label == "司机取消率":
        return source[prefix + "cancelled_by_driver"] / source[prefix + "response_count"]
    if metric.label == "乘客取消率":
        return source[prefix + "cancelled_by_passenger"] / source[prefix + "response_count"]
    if metric.label == "司乘取消率":
        return (
            source[prefix + "cancelled_by_driver"] + source[prefix + "cancelled_by_passenger"]
        ) / source[prefix + "response_count"]
    if metric.label == "完单/应答（成交率":
        return source[prefix + "completed_order_count"] / source[prefix + "response_count"]
    if metric.label == "gmv":
        return source[prefix + "gmv"] / source[prefix + "day_count"]
    if metric.label == "周期总gmv":
        return source[prefix + "gmv"]
    raise ValueError(f"未知指标: {metric.label}")


def metric_change_value(row: dict[str, Any], metric: MetricSpec) -> float | None:
    prev = metric_value(row, "previous_", metric)
    curr = metric_value(row, "current_", metric)
    if prev is None or curr is None:
        return None
    if metric.metric_type == "volume":
        if prev == 0:
            return None
        return (curr - prev) / prev
    return curr - prev


def display_value(value: float | None) -> float | str:
    return MISSING_DISPLAY if value is None else value


def add_calculation_gaps(rows: list[dict[str, Any]], gaps: list[dict[str, Any]]) -> None:
    for row in rows:
        for metric in METRICS:
            for period_name, prefix in (("上期", "previous_"), ("本期", "current_")):
                ok, reason = can_compute(row, prefix, metric)
                if not ok:
                    gaps.append(
                        {
                            "type": "metric_missing",
                            "period": period_name,
                            "dimension": row["dimension"],
                            "operator": row["operator"],
                            "brand": row["brand"],
                            "city": row["city"],
                            "field": metric.label,
                            "reason": reason,
                        }
                    )
            prev_ok, prev_reason = can_compute(row, "previous_", metric)
            curr_ok, curr_reason = can_compute(row, "current_", metric)
            reason = curr_reason or prev_reason
            if metric.metric_type == "volume" and prev_ok:
                prev_value = metric_value(row, "previous_", metric)
                if prev_value == 0:
                    prev_ok = False
                    reason = "上期值为0"
            if not prev_ok or not curr_ok:
                gaps.append(
                    {
                        "type": "change_missing",
                        "dimension": row["dimension"],
                        "operator": row["operator"],
                        "brand": row["brand"],
                        "city": row["city"],
                        "field": metric.label,
                        "reason": reason,
                    }
                )


def source_columns() -> list[str]:
    columns = ["contact_person", "dimension", "operator", "brand", "city"]
    for prefix in ("previous_", "current_"):
        columns.extend([prefix + "row_count", prefix + "day_count"])
        columns.extend(prefix + field for field in SUM_FIELDS)
    return columns


def copy_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def formula_if(period_prefix: str, metric: MetricSpec, cols: dict[str, str]) -> str:
    checks = [f"{cols[period_prefix + 'row_count']}=0"]
    _, denominators = metric_requirements(metric.label)
    checks.extend(f"{cols[period_prefix + denominator]}=0" for denominator in denominators)
    check = ",".join(checks)
    return f'=IFERROR(IF(OR({check}),"{MISSING_DISPLAY}",{metric.expr(period_prefix, cols)}),"{MISSING_DISPLAY}")'


def formula_change(metric: MetricSpec, cols: dict[str, str]) -> str:
    prev_checks = [f"{cols['previous_row_count']}=0"]
    curr_checks = [f"{cols['current_row_count']}=0"]
    _, denominators = metric_requirements(metric.label)
    prev_checks.extend(f"{cols['previous_' + denominator]}=0" for denominator in denominators)
    curr_checks.extend(f"{cols['current_' + denominator]}=0" for denominator in denominators)
    if metric.metric_type == "volume":
        prev_expr = metric.expr("previous_", cols)
        curr_expr = metric.expr("current_", cols)
        expr = f'IF(({prev_expr})=0,"{MISSING_DISPLAY}",(({curr_expr})-({prev_expr}))/({prev_expr}))'
    else:
        expr = f"({metric.expr('current_', cols)})-({metric.expr('previous_', cols)})"
    checks = ",".join(prev_checks + curr_checks)
    return f'=IFERROR(IF(OR({checks}),"{MISSING_DISPLAY}",{expr}),"{MISSING_DISPLAY}")'


def write_source_sheet(wb, rows: list[dict[str, Any]]) -> dict[str, dict[str, str]]:
    if "__source" in wb.sheetnames:
        del wb["__source"]
    ws = wb.create_sheet("__source")
    ws.sheet_state = "hidden"
    columns = source_columns()
    for col_idx, name in enumerate(columns, 1):
        ws.cell(1, col_idx, name)
    row_refs: dict[str, dict[str, str]] = {}
    for row_idx, row in enumerate(rows, 2):
        values = {
            "contact_person": row["contact_person"],
            "dimension": row["dimension"],
            "operator": row["operator"],
            "brand": row["brand"],
            "city": row["city"],
            **row["source"],
        }
        for col_idx, name in enumerate(columns, 1):
            ws.cell(row_idx, col_idx, values.get(name, ""))
        refs = {
            name: f"'__source'!${get_column_letter(col_idx)}${row_idx}"
            for col_idx, name in enumerate(columns, 1)
        }
        row_refs[str(row_idx - 2)] = refs
    return row_refs


def write_report_workbook(
    template_path: Path,
    output_path: Path,
    periods: Periods,
    rows: list[dict[str, Any]],
    formula_mode: bool = False,
) -> None:
    wb = load_workbook(template_path)
    ws = wb.active
    template_styles = [copy(ws.cell(3, col)._style) for col in range(1, ws.max_column + 1)]
    template_row_height = ws.row_dimensions[3].height
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    ws["A1"] = "上期日期"
    ws["B1"] = periods.previous.label
    ws["C1"] = "本期日期"
    ws["D1"] = periods.current.label

    if formula_mode:
        refs_by_row = write_source_sheet(wb, rows)
    else:
        if "__source" in wb.sheetnames:
            del wb["__source"]
        refs_by_row = {}
    for idx, row in enumerate(rows, 3):
        row_key = str(idx - 3)
        refs = refs_by_row.get(row_key, {})
        values = [row["contact_person"], row["dimension"], row["operator"], row["brand"], row["city"]]
        for col_idx, value in enumerate(values, 1):
            ws.cell(idx, col_idx, value)
        for metric in METRICS:
            prev_col, curr_col, change_col = metric.columns
            if formula_mode:
                ws.cell(idx, prev_col, formula_if("previous_", metric, refs))
                ws.cell(idx, curr_col, formula_if("current_", metric, refs))
                ws.cell(idx, change_col, formula_change(metric, refs))
            else:
                ws.cell(idx, prev_col, display_value(metric_value(row, "previous_", metric)))
                ws.cell(idx, curr_col, display_value(metric_value(row, "current_", metric)))
                ws.cell(idx, change_col, display_value(metric_change_value(row, metric)))
            for col in metric.columns:
                if metric.metric_type == "rate" or col == change_col:
                    ws.cell(idx, col).number_format = "0.00%"
                else:
                    ws.cell(idx, col).number_format = "#,##0.00"
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(idx, col_idx)._style = copy(template_styles[col_idx - 1])
        for metric in METRICS:
            for col in metric.columns:
                if metric.metric_type == "rate" or col == metric.columns[2]:
                    ws.cell(idx, col).number_format = "0.00%"
                else:
                    ws.cell(idx, col).number_format = "#,##0.00"
        if template_row_height:
            ws.row_dimensions[idx].height = template_row_height

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def summarize(
    periods: Periods,
    rows: list[dict[str, Any]],
    gaps: list[dict[str, Any]],
    output_path: Path | None,
    gap_file: Path | None = None,
    source_mode: str = "",
    write_mode: str = "",
    fallback_reason: str = "",
) -> dict[str, Any]:
    dim_counts = Counter(row["dimension"] for row in rows)
    gap_counts = Counter(gap["type"] for gap in gaps)
    missing_by_field = Counter(
        f"{gap.get('field')}|{gap.get('reason')}"
        for gap in gaps
        if gap["type"] in {"metric_missing", "change_missing"}
    )
    return {
        "previous_period": {
            "start": periods.previous.start.isoformat(),
            "end": periods.previous.end.isoformat(),
            "day_count": periods.previous.day_count,
        },
        "current_period": {
            "start": periods.current.start.isoformat(),
            "end": periods.current.end.isoformat(),
            "day_count": periods.current.day_count,
        },
        "output_file": str(output_path) if output_path else "",
        "source_mode": source_mode,
        "write_mode": write_mode,
        "fallback_reason": fallback_reason,
        "row_count": len(rows),
        "dimension_counts": dict(dim_counts),
        "gap_counts": dict(gap_counts),
        "missing_by_field": dict(missing_by_field),
        "gap_file": str(gap_file) if gap_file else "",
        "gap_samples": gaps[:30],
    }


def output_paths(args: argparse.Namespace, periods: Periods) -> tuple[Path, Path]:
    if args.output:
        output_file = Path(args.output).expanduser()
        if not output_file.is_absolute():
            output_file = PROJECT_ROOT / output_file
    else:
        output_dir = Path(args.output_dir).expanduser() if args.output_dir else DEFAULT_OUTPUT_DIR
        if not output_dir.is_absolute():
            output_dir = PROJECT_ROOT / output_dir
        output_file = output_dir / f"hhdata周报_{periods.current.file_label}.xlsx"
    summary_file = output_file.with_name(f"{output_file.stem}_summary.json")
    return output_file, summary_file


def gap_path_for(output_file: Path) -> Path:
    return output_file.with_name(f"{output_file.stem}_gaps.csv")


def write_gaps_csv(path: Path, gaps: list[dict[str, Any]]) -> None:
    if not gaps:
        path.write_text("", encoding="utf-8")
        return
    pd.DataFrame(gaps).to_csv(path, index=False, encoding="utf-8-sig")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="生成 hhdata 周报 Excel")
    parser.add_argument("--start", type=parse_date, help="本期开始日期 YYYY-MM-DD")
    parser.add_argument("--end", type=parse_date, help="本期结束日期 YYYY-MM-DD")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="周报模板路径")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--output", default="", help="指定输出 xlsx 路径")
    parser.add_argument(
        "--source",
        choices=["auto", "weekly", "daily", "fact"],
        default="auto",
        help="数据来源：auto 优先聚合表；weekly 强制周聚合；daily 强制日粒度品牌城市聚合；fact 强制原始事实表",
    )
    parser.add_argument("--formula-mode", action="store_true", help="写 Excel 公式；默认直接写计算后的值")
    parser.add_argument("--dry-run", action="store_true", help="只预览，不写 Excel")
    return parser


def build_rows_from_fact(db: DatabaseConnector, periods: Periods) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    df = load_hhdata(db, periods)
    mapping_rows = load_operator_brand_rows()
    mapping = build_mapping(mapping_rows)
    return build_report_rows(df, mapping, periods)


def build_rows_from_daily(db: DatabaseConnector, periods: Periods) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not table_exists(db, DAILY_BRAND_CITY_METRICS_TABLE):
        raise RuntimeError(f"日聚合表不存在: {DAILY_BRAND_CITY_METRICS_TABLE}")
    df = load_daily_agg(db, periods)
    if not daily_agg_has_periods(df, periods):
        raise RuntimeError("日聚合表缺少本期或上期日期，请先运行 refresh_daily_brand_city_agg.py")
    mapping_rows = load_operator_brand_rows()
    mapping = build_mapping(mapping_rows)
    return build_report_rows(df, mapping, periods)


def build_rows_from_weekly(db: DatabaseConnector, periods: Periods) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not table_exists(db, WEEKLY_METRICS_TABLE):
        raise RuntimeError(f"周聚合表不存在: {WEEKLY_METRICS_TABLE}")
    df = load_weekly_agg(db, periods)
    if not weekly_agg_has_periods(df, periods):
        raise RuntimeError(
            "周聚合表缺少本期或上期数据，请先运行 "
            f"refresh_weekly_agg.py --week-start {periods.previous.start.isoformat()} "
            f"和 --week-start {periods.current.start.isoformat()}"
        )
    return build_report_rows_from_weekly_agg(df, periods)


def build_rows_auto(db: DatabaseConnector, periods: Periods) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, str]:
    reasons: list[str] = []
    if is_complete_natural_week(periods.current):
        try:
            rows, gaps = build_rows_from_weekly(db, periods)
            return rows, gaps, "weekly", ""
        except RuntimeError as exc:
            reasons.append(f"weekly: {exc}")
    else:
        reasons.append("非完整自然周，跳过 weekly")

    try:
        rows, gaps = build_rows_from_daily(db, periods)
        return rows, gaps, "daily", "; ".join(reasons)
    except RuntimeError as exc:
        reasons.append(f"daily: {exc}")

    rows, gaps = build_rows_from_fact(db, periods)
    return rows, gaps, "fact", "; ".join(reasons)


def run(args: argparse.Namespace) -> dict[str, Any]:
    db = DatabaseConnector()
    latest = load_latest_data_date(db)
    periods = resolve_periods(args.start, args.end, latest)
    template_path = Path(args.template).expanduser()
    if not template_path.is_absolute():
        template_path = PROJECT_ROOT / template_path
    if not template_path.exists():
        raise FileNotFoundError(f"周报模板不存在: {template_path}")

    if args.source == "auto":
        rows, gaps, source_mode, fallback_reason = build_rows_auto(db, periods)
    elif args.source == "weekly":
        fallback_reason = ""
        if not is_complete_natural_week(periods.current):
            raise ValueError("weekly 来源只支持完整自然周；非 7 天周期请使用 auto 或 daily")
        rows, gaps = build_rows_from_weekly(db, periods)
        source_mode = "weekly"
    elif args.source == "daily":
        fallback_reason = ""
        rows, gaps = build_rows_from_daily(db, periods)
        source_mode = "daily"
    else:
        fallback_reason = ""
        rows, gaps = build_rows_from_fact(db, periods)
        source_mode = "fact"

    output_file, summary_file = output_paths(args, periods)
    gap_file = gap_path_for(output_file)
    summary = summarize(
        periods,
        rows,
        gaps,
        None if args.dry_run else output_file,
        None if args.dry_run else gap_file,
        source_mode=source_mode,
        write_mode="formula" if args.formula_mode else "values",
        fallback_reason=fallback_reason,
    )

    if args.dry_run:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return summary

    write_report_workbook(template_path, output_file, periods, rows, formula_mode=args.formula_mode)
    write_gaps_csv(gap_file, gaps)
    summary_file.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return summary


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    try:
        run(args)
    except Exception as exc:
        print(f"❌ 周报生成失败: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
