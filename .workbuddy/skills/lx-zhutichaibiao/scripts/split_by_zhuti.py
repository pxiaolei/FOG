#!/usr/bin/env python3
"""
按运营主体/城市/品牌拆表工具

拆分逻辑：
- 按运营主体拆：根据原表列智能匹配（有品牌+城市用双条件，有单列用单条件）
- 按城市拆：按城市匹配运营主体，输出运营主体文件
- 按品牌拆：按品牌匹配运营主体，输出运营主体文件
- 纯品牌拆：按品牌独立输出，不映射运营主体

依赖：
- lxx_share.excel_utils（列检测、对接人筛选）
- lx_shujuku（公司库 operator_brand 码表映射）
- openpyxl + pyyaml

使用方法：
  python split_by_zhuti.py                    # 正常运行（交互式）
  python split_by_zhuti.py --config           # 重新配置
  python split_by_zhuti.py -m 1 -p 雷维亮        # 命令行指定模式和对人
"""

import sys
import shutil
import zipfile
import argparse
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# 确保 lxx_share / lx_shujuku 可 import（向上查找 skills 目录）
def _find_skills_dir():
    from pathlib import Path
    for p in Path(__file__).resolve().parents:
        if (p / "lxx_share").is_dir() and (p / "lx_shujuku").is_dir():
            return p
    return Path(__file__).resolve().parents[2]

_skills_dir = _find_skills_dir()
import sys
if str(_skills_dir) not in sys.path:
    sys.path.insert(0, str(_skills_dir))
_lx_shujuku_scripts_dir = _skills_dir / "lx_shujuku" / "scripts"
if str(_lx_shujuku_scripts_dir) not in sys.path:
    sys.path.insert(0, str(_lx_shujuku_scripts_dir))

from lxx_share.excel_utils import (
    CITY_FIELDS,
    BRAND_FIELDS,
    PERSON_FIELDS,
    OPERATOR_FIELDS,
    find_column,
    detect_columns,
    filter_by_person,
)
from lxx_share.fog_config import (
    find_project_root,
    load_fog_config,
    resolve_project_path,
    save_fog_config,
)
from lx_shujuku import create_client


# ========================================
# 路径与默认值
# ========================================

SKILL_DIR = Path(__file__).parent.parent
DEFAULT_PERSONS = ["雷维亮"]


# ========================================
# CLI 参数
# ========================================

def parse_args():
    parser = argparse.ArgumentParser(description="按运营主体/城市/品牌拆表工具")
    parser.add_argument("--mode", "-m", type=int, choices=[1, 2, 3, 4],
                        help="拆分维度: 1=运营主体, 2=城市, 3=品牌→运营主体, 4=纯品牌")
    parser.add_argument("--person", "-p", type=str,
                        help="对接人范围: 'all' 或逗号分隔的对接人列表")
    parser.add_argument("--keep-sheets", "-k", type=str,
                        help="保留的sheet名称，逗号分隔")
    parser.add_argument("--config", action="store_true",
                        help="重新配置")
    return parser.parse_args()


# ========================================
# 配置管理
# ========================================

def interactive_setup():
    """交互式配置（首次运行）。"""
    print("\n" + "=" * 60)
    print("按运营主体/城市/品牌拆表工具 - 首次配置")
    print("=" * 60)

    print("\n【重要】码表将从公司 dataReporting 的 operator_brand 表读取")
    print("  请先在 config/fog_config.yaml 的 lx_shujuku.api 段填写 dataReporting 账号")
    print("-" * 60)

    print("\n检测到未配置，请提供以下信息：\n")

    # 1. 项目根目录
    current_dir = Path.cwd()
    print(f"1. 项目根目录")
    print(f"   当前目录: {current_dir}")
    user_input = input("   回车使用当前目录，或输入路径: ").strip()
    project_root = Path(user_input) if user_input else current_dir
    project_root = project_root.resolve()

    # 2. 默认对接人
    print(f"\n2. 默认对接人")
    print("   多个对接人用逗号分隔，如: 雷维亮,陈双")
    user_input = input(f"   输入默认对接人（默认: {','.join(DEFAULT_PERSONS)}）: ").strip()
    if user_input:
        default_persons = [p.strip() for p in user_input.split(",")]
    else:
        default_persons = DEFAULT_PERSONS

    # 3. 工作目录
    print(f"\n3. 工作目录")
    print("   目录下需包含: 输入/、输出/、原表存档/")
    user_input = input("   输入工作目录名称（默认: workspace/01主体拆表）: ").strip()
    work_dir_name = user_input if user_input else "workspace/01主体拆表"
    work_dir = project_root / work_dir_name

    # 确认
    print("\n" + "-" * 60)
    print("配置确认:")
    print(f"  项目根目录: {project_root}")
    print("  码表来源: lx_shujuku / operator_brand")
    print(f"  默认对接人: {default_persons}")
    print(f"  工作目录: {work_dir}")
    print("-" * 60)

    user_input = input("\n确认保存配置？(Y/n): ").strip().lower()
    if user_input == 'n':
        print("配置已取消")
        return None

    config = {
        "项目根目录": str(project_root),
        "码表来源": "lx_shujuku.operator_brand",
        "默认对接人": default_persons,
        "工作目录": str(work_dir),
        "默认": {
            "城市字段": CITY_FIELDS,
            "品牌字段": BRAND_FIELDS,
            "处理sheet": [],
        },
        "特定配置": [],
    }

    fog_config = load_fog_config(project_root)
    fog_config.setdefault("project", {})
    fog_config["project"]["root"] = str(project_root)
    fog_config["project"].setdefault("workspace_root", "workspace")
    fog_config.setdefault("lx_zhutichaibiao", {})
    fog_config["lx_zhutichaibiao"].update({
        "work_dir": work_dir_name,
        "default_persons": default_persons,
        "mabiao_source": "lx_shujuku.operator_brand",
        "default": {
            "city_fields": CITY_FIELDS,
            "brand_fields": BRAND_FIELDS,
            "process_sheets": [],
        },
        "special_configs": [],
    })
    config_path = save_fog_config(fog_config, project_root)

    print(f"\n✅ 配置已保存到: {config_path}")
    return build_config_from_fog(fog_config, project_root)


def load_config():
    """加载配置文件。"""
    project_root = find_project_root(Path(__file__))
    fog_config = load_fog_config(project_root)
    if isinstance(fog_config.get("lx_zhutichaibiao"), dict):
        return build_config_from_fog(fog_config, project_root)
    return None


def build_config_from_fog(fog_config, project_root):
    """把 fog_config.yaml 的 lx_zhutichaibiao 段转换为旧脚本内部结构。"""
    zhutichaibiao = fog_config.get("lx_zhutichaibiao", {})
    if not isinstance(zhutichaibiao, dict):
        zhutichaibiao = {}
    default = zhutichaibiao.get("default", {})
    if not isinstance(default, dict):
        default = {}

    work_dir = resolve_project_path(
        zhutichaibiao.get("work_dir"),
        project_root,
        default="workspace/01主体拆表",
    )
    return {
        "项目根目录": str(project_root),
        "码表来源": zhutichaibiao.get("mabiao_source", "lx_shujuku.operator_brand"),
        "默认对接人": zhutichaibiao.get("default_persons", []),
        "工作目录": str(work_dir),
        "默认": {
            "城市字段": default.get("city_fields") or default.get("城市字段") or CITY_FIELDS,
            "品牌字段": default.get("brand_fields") or default.get("品牌字段") or BRAND_FIELDS,
            "处理sheet": default.get("process_sheets") or default.get("处理sheet") or [],
            "表头行数": default.get("header_rows") or default.get("表头行数") or {},
        },
        "特定配置": zhutichaibiao.get("special_configs") or zhutichaibiao.get("特定配置") or [],
    }


def load_company_mabiao():
    """从 lx_shujuku 加载公司库码表映射。"""
    client = create_client()
    return client.load_mabiao_mapping()


def get_file_config(config, filename):
    """根据文件名获取配置。"""
    default = config.get("默认", {
        "城市字段": CITY_FIELDS,
        "品牌字段": BRAND_FIELDS,
        "处理sheet": [],
        "表头行数": {},
    })

    for cfg in config.get("特定配置", []):
        if cfg.get("匹配") and cfg["匹配"] in filename:
            result = default.copy()
            result.update(cfg)
            return result

    return default


# ========================================
# 核心拆分逻辑
# ========================================

def process_file_optimized(file_path, mapping, file_config, split_mode, col_info):
    """处理单个文件（优化版：一次遍历收集所有数据）。

    Args:
        split_mode: '运营主体' / '城市' / '品牌' / '纯品牌'

    Returns:
        (result_files, stats, unmatch_data, header_info)
    """
    print(f"\n处理文件: {file_path.name}")
    print(f"  拆分模式: {split_mode}")
    print(f"  检测结果: 城市={col_info['has_city']}, 品牌={col_info['has_brand']}")

    # 获取保留 sheet 列表
    keep_sheets = file_config.get("保留 sheet", [])
    if keep_sheets:
        print(f"  保留sheet: {keep_sheets}")

    # 检查必要列
    if split_mode == '城市' and not col_info['has_city']:
        print(f"  ❌ 缺少城市列，无法按城市拆分")
        return {}, {}, {}, {}
    if split_mode == '品牌' and not col_info['has_brand']:
        print(f"  ❌ 缺少品牌列，无法按品牌拆分")
        return {}, {}, {}, {}

    city_fields = file_config.get("城市字段", CITY_FIELDS)
    brand_fields = file_config.get("品牌字段", BRAND_FIELDS)

    # 数据结构
    data_by_operator = defaultdict(lambda: defaultdict(list))
    row_styles_by_operator = defaultdict(lambda: defaultdict(list))
    header_info = {}
    keep_sheets_data = {}
    unmatch_data = defaultdict(list)

    src_wb = load_workbook(file_path, data_only=False)
    total_rows = 0

    for sheet_name in src_wb.sheetnames:
        if src_wb[sheet_name].sheet_state != 'visible':
            continue

        src_ws = src_wb[sheet_name]
        max_row = src_ws.max_row
        max_col = src_ws.max_column

        is_keep_sheet = sheet_name in keep_sheets

        header_info[sheet_name] = {
            'rows': [], 'styles': [], 'col_widths': {},
            'header_rows': 1, 'merged_cells': [],
        }

        sheet_header_rows = 1

        # 检查配置中的表头行数
        header_rows_config = file_config.get("表头行数", {})
        if sheet_name in header_rows_config:
            sheet_header_rows = header_rows_config[sheet_name]
            print(f"  sheet '{sheet_name}' 使用配置表头行数: {sheet_header_rows}")
        elif not is_keep_sheet:
            for row_idx in range(1, min(5, max_row + 1)):
                for col_idx in range(1, max_col + 1):
                    cell_value = src_ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        col_str = str(cell_value).strip()
                        if col_str in city_fields or col_str in brand_fields:
                            sheet_header_rows = max(sheet_header_rows, row_idx)

        # 收集表头
        for row_idx in range(1, sheet_header_rows + 1):
            row_data = []
            row_styles = []
            for col_idx in range(1, max_col + 1):
                cell = src_ws.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)
                row_styles.append({
                    'number_format': cell.number_format,
                    'font': copy(cell.font) if cell.has_style else None,
                    'border': copy(cell.border) if cell.has_style else None,
                    'fill': copy(cell.fill) if cell.has_style else None,
                    'alignment': copy(cell.alignment) if cell.has_style else None,
                })
            header_info[sheet_name]['rows'].append(row_data)
            header_info[sheet_name]['styles'].append(row_styles)

        header_info[sheet_name]['header_rows'] = sheet_header_rows

        # 表头区域的合并单元格
        for merged_range in src_ws.merged_cells.ranges:
            if merged_range.min_row <= sheet_header_rows:
                header_info[sheet_name]['merged_cells'].append(merged_range)

        # 列宽
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in src_ws.column_dimensions:
                header_info[sheet_name]['col_widths'][col_letter] = src_ws.column_dimensions[col_letter].width

        # 保留 sheet：完整复制
        if is_keep_sheet:
            keep_sheets_data[sheet_name] = {
                'rows': [], 'styles': [],
                'max_row': max_row, 'max_col': max_col,
                'merged_cells': list(src_ws.merged_cells.ranges),
                'row_heights': {},
            }
            for row_idx in range(1, max_row + 1):
                row_data = []
                row_styles = []
                for col_idx in range(1, max_col + 1):
                    cell = src_ws.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                    row_styles.append({
                        'number_format': cell.number_format,
                        'font': copy(cell.font) if cell.has_style else None,
                        'border': copy(cell.border) if cell.has_style else None,
                        'fill': copy(cell.fill) if cell.has_style else None,
                        'alignment': copy(cell.alignment) if cell.has_style else None,
                    })
                keep_sheets_data[sheet_name]['rows'].append(row_data)
                keep_sheets_data[sheet_name]['styles'].append(row_styles)
                if row_idx in src_ws.row_dimensions:
                    keep_sheets_data[sheet_name]['row_heights'][row_idx] = src_ws.row_dimensions[row_idx].height
            print(f"  保留sheet '{sheet_name}': {max_row}行（完整复制，含{len(keep_sheets_data[sheet_name]['merged_cells'])}个合并单元格）")
            continue

        # 非保留 sheet：独立检测列位置
        sheet_city_col_idx = None
        sheet_brand_col_idx = None
        for row_idx in range(1, sheet_header_rows + 1):
            for col_idx in range(1, max_col + 1):
                cell_value = src_ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    col_str = str(cell_value).strip()
                    if col_str in city_fields and sheet_city_col_idx is None:
                        sheet_city_col_idx = col_idx
                    if col_str in brand_fields and sheet_brand_col_idx is None:
                        sheet_brand_col_idx = col_idx

        if split_mode == '城市' and sheet_city_col_idx is None:
            print(f"  ⚠️ Sheet '{sheet_name}' 缺少城市列，跳过")
            continue
        if split_mode == '品牌' and sheet_brand_col_idx is None:
            print(f"  ⚠️ Sheet '{sheet_name}' 缺少品牌列，跳过")
            continue

        # 一次遍历数据行
        data_rows = list(range(sheet_header_rows + 1, max_row + 1))
        row_count = len(data_rows)

        for idx, row_idx in enumerate(data_rows):
            if row_count > 500 and idx % 100 == 0 and idx > 0:
                print(f"    处理进度: {idx}/{row_count} 行 ({idx*100//row_count}%)")

            city_value = src_ws.cell(row=row_idx, column=sheet_city_col_idx).value if sheet_city_col_idx else None
            brand_value = src_ws.cell(row=row_idx, column=sheet_brand_col_idx).value if sheet_brand_col_idx else None

            city_str = str(city_value).strip() if city_value else None
            brand_str = str(brand_value).strip() if brand_value else None

            matched_targets = set()
            unmatch_reason = None

            if split_mode == '纯品牌':
                if brand_str:
                    matched_targets.add(brand_str)
                else:
                    unmatch_reason = "品牌列为空"

            elif split_mode == '城市':
                if city_str and city_str in mapping['city_to_zhuti']:
                    matched_targets.update(mapping['city_to_zhuti'][city_str])
                elif city_str:
                    unmatch_reason = f"城市 '{city_str}' 不在码表中"
                else:
                    unmatch_reason = "城市列为空"

            elif split_mode == '品牌':
                if brand_str and brand_str in mapping['brand_to_zhuti']:
                    matched_targets.update(mapping['brand_to_zhuti'][brand_str])
                elif brand_str:
                    unmatch_reason = f"品牌 '{brand_str}' 不在码表中"
                else:
                    unmatch_reason = "品牌列为空"

            else:  # 运营主体
                if brand_str and city_str:
                    key = (brand_str, city_str)
                    if key in mapping['brand_city_to_zhuti']:
                        matched_targets.update(mapping['brand_city_to_zhuti'][key])
                    else:
                        unmatch_reason = f"品牌+城市组合 '{brand_str}+{city_str}' 不在码表中"
                elif city_str and city_str in mapping['city_to_zhuti']:
                    matched_targets.update(mapping['city_to_zhuti'][city_str])
                elif brand_str and brand_str in mapping['brand_to_zhuti']:
                    matched_targets.update(mapping['brand_to_zhuti'][brand_str])
                elif not brand_str and not city_str:
                    unmatch_reason = "品牌和城市列都为空"
                elif not brand_str:
                    unmatch_reason = "品牌列为空"
                else:
                    unmatch_reason = "城市列为空"

            # 收集数据行
            if matched_targets:
                total_rows += 1
                row_data = []
                row_styles = []
                for col_idx in range(1, max_col + 1):
                    cell = src_ws.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                    row_styles.append({
                        'number_format': cell.number_format,
                        'font': copy(cell.font) if cell.has_style else None,
                        'border': copy(cell.border) if cell.has_style else None,
                        'fill': copy(cell.fill) if cell.has_style else None,
                        'alignment': copy(cell.alignment) if cell.has_style else None,
                    })

                for target in matched_targets:
                    if split_mode == '纯品牌' or target in mapping.get('all_zhuti', []):
                        data_by_operator[target][sheet_name].append(row_data)
                        row_styles_by_operator[target][sheet_name].append(row_styles)
            elif unmatch_reason:
                row_data = [src_ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
                unmatch_data[sheet_name].append((row_data, brand_str, city_str, unmatch_reason))

    src_wb.close()

    print(f"  总数据行: {total_rows}")

    # 创建临时目录
    temp_dir = file_path.parent / f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    temp_dir.mkdir(exist_ok=True)

    result_files = {}
    stats = {}

    # 为每个有数据的运营主体生成文件
    for operator in sorted(data_by_operator.keys()):
        sheets_data = data_by_operator[operator]
        if not sheets_data:
            continue

        operator_rows = sum(len(rows) for rows in sheets_data.values())
        stats[operator] = operator_rows

        output_name = f"{operator}_{file_path.name}"
        output_path = temp_dir / output_name

        dst_wb = Workbook()
        dst_wb.remove(dst_wb.active)

        for sheet_name in sheets_data.keys():
            rows_data = sheets_data[sheet_name]
            rows_styles = row_styles_by_operator[operator][sheet_name]

            dst_ws = dst_wb.create_sheet(title=sheet_name[:31])

            # 写入表头
            if sheet_name in header_info:
                h_info = header_info[sheet_name]
                for row_idx, (row_data, row_styles) in enumerate(zip(h_info['rows'], h_info['styles']), 1):
                    for col_idx, (value, style) in enumerate(zip(row_data, row_styles), 1):
                        cell = dst_ws.cell(row=row_idx, column=col_idx, value=value)
                        if style:
                            if style['font']:
                                cell.font = style['font']
                            if style['border']:
                                cell.border = style['border']
                            if style['fill']:
                                cell.fill = style['fill']
                            if style['alignment']:
                                cell.alignment = style['alignment']
                            cell.number_format = style['number_format']

                for col_letter, width in h_info['col_widths'].items():
                    dst_ws.column_dimensions[col_letter].width = width

                if h_info.get('merged_cells'):
                    for merged_range in h_info['merged_cells']:
                        dst_ws.merge_cells(str(merged_range))

            # 写入数据行
            sheet_hdr_rows = h_info.get('header_rows', 1) if sheet_name in header_info else 1
            for row_idx, (row_data, row_styles) in enumerate(zip(rows_data, rows_styles), sheet_hdr_rows + 1):
                for col_idx, (value, style) in enumerate(zip(row_data, row_styles), 1):
                    cell = dst_ws.cell(row=row_idx, column=col_idx, value=value)
                    if style:
                        if style['font']:
                            cell.font = style['font']
                        if style['border']:
                            cell.border = style['border']
                        if style['fill']:
                            cell.fill = style['fill']
                        if style['alignment']:
                            cell.alignment = style['alignment']
                        cell.number_format = style['number_format']

        # 写入保留 sheet
        for keep_sheet_name, keep_data in keep_sheets_data.items():
            dst_ws = dst_wb.create_sheet(title=keep_sheet_name[:31])
            for row_idx, (row_data, row_styles) in enumerate(zip(keep_data['rows'], keep_data['styles']), 1):
                for col_idx, (value, style) in enumerate(zip(row_data, row_styles), 1):
                    cell = dst_ws.cell(row=row_idx, column=col_idx, value=value)
                    if style:
                        if style['font']:
                            cell.font = style['font']
                        if style['border']:
                            cell.border = style['border']
                        if style['fill']:
                            cell.fill = style['fill']
                        if style['alignment']:
                            cell.alignment = style['alignment']
                        cell.number_format = style['number_format']
            if keep_sheet_name in header_info:
                for col_letter, width in header_info[keep_sheet_name]['col_widths'].items():
                    dst_ws.column_dimensions[col_letter].width = width
            if 'merged_cells' in keep_data:
                for merged_range in keep_data['merged_cells']:
                    dst_ws.merge_cells(str(merged_range))
            if 'row_heights' in keep_data:
                for row_idx, height in keep_data['row_heights'].items():
                    if height is not None:
                        dst_ws.row_dimensions[row_idx].height = height

        dst_wb.save(output_path)
        dst_wb.close()
        result_files[operator] = output_path
        print(f"  生成: {output_name} ({operator_rows}行)")

    if unmatch_data:
        total_unmatch = sum(len(rows) for rows in unmatch_data.values())
        print(f"  未匹配数据: {total_unmatch}行")
    else:
        print(f"  未匹配数据: 0行")

    return result_files, stats, unmatch_data, header_info


# ========================================
# 报告与打包
# ========================================

def output_unmatch_report(unmatch_data, file_path, output_dir, header_info):
    """输出未匹配数据报告。"""
    if not unmatch_data:
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    report_name = f"{timestamp}_{file_path.stem}_未匹配.xlsx"
    report_path = output_dir / report_name

    try:
        wb = Workbook()
        wb.remove(wb.active)
        total_rows = 0

        for sheet_name, rows in unmatch_data.items():
            if not rows:
                continue

            ws = wb.create_sheet(title=sheet_name[:31])
            h_info = header_info.get(sheet_name, {})
            header_rows_count = h_info.get('header_rows', 1)

            if h_info.get('rows'):
                orig_col_count = 0
                for row_idx, (row_data, row_styles) in enumerate(zip(h_info['rows'], h_info['styles']), 1):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    orig_col_count = max(orig_col_count, len(row_data))
                ws.cell(row=header_rows_count, column=orig_col_count + 1, value="品牌")
                ws.cell(row=header_rows_count, column=orig_col_count + 2, value="城市")
                ws.cell(row=header_rows_count, column=orig_col_count + 3, value="未匹配原因")
            else:
                if rows:
                    orig_col_count = len(rows[0][0])
                    for i in range(1, orig_col_count + 1):
                        ws.cell(row=1, column=i, value=f"列{i}")
                    ws.cell(row=1, column=orig_col_count + 1, value="品牌")
                    ws.cell(row=1, column=orig_col_count + 2, value="城市")
                    ws.cell(row=1, column=orig_col_count + 3, value="未匹配原因")
                    header_rows_count = 1

            start_row = header_rows_count + 1
            for row_data, brand, city, reason in rows:
                orig_col_count = len(row_data)
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=start_row, column=col_idx, value=value)
                ws.cell(row=start_row, column=orig_col_count + 1, value=brand)
                ws.cell(row=start_row, column=orig_col_count + 2, value=city)
                ws.cell(row=start_row, column=orig_col_count + 3, value=reason)
                start_row += 1
                total_rows += 1

        wb.save(report_path)
        wb.close()

        print(f"  未匹配报告: {report_name} ({total_rows}行)")
        return report_path

    except Exception as e:
        print(f"  ⚠️ 未匹配报告生成失败: {e}")
        try:
            wb.close()
        except Exception:
            pass
        return None


def pack_results(result_files, stats, mapping, file_path, output_dir, person_list, split_mode='运营主体'):
    """打包结果。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    zip_name = f"{timestamp}_{file_path.stem}.zip"
    zip_path = output_dir / zip_name

    need_person_folders = len(person_list) > 1 and split_mode != '纯品牌'
    target_to_person = mapping.get('zhuti_to_person', {})

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(file_path, file_path.name)

        if need_person_folders:
            for person in sorted(person_list):
                for target, op_file in result_files.items():
                    persons_for_target = target_to_person.get(target, [])
                    if person in persons_for_target:
                        zf.write(op_file, f"{person}/{op_file.name}")
        else:
            for target, op_file in result_files.items():
                zf.write(op_file, op_file.name)

    print(f"\n打包完成: {zip_name}")
    print(f"  分文件数: {len(result_files)}")
    print(f"  总数据行: {sum(stats.values())}")
    if need_person_folders:
        print(f"  按对接人分文件夹")
    else:
        print(f"  直接打包")

    # 清理临时文件
    for op_file in result_files.values():
        if op_file.exists():
            op_file.unlink()
    if result_files:
        temp_dir = list(result_files.values())[0].parent
        if temp_dir.exists():
            shutil.rmtree(temp_dir)

    return zip_path


# ========================================
# 处理日志
# ========================================

def write_processing_log(log_dir, log_entries, split_mode, person_list, default_persons):
    """在处理日志文件夹写入 Markdown 格式的拆分日志。

    Args:
        log_dir: 处理日志目录（Path 对象）
        log_entries: [{'file_path', 'zip_path', 'archive_path', 'stats', 'unmatch_count'}, ...]
        split_mode: 拆分模式名称
        person_list: 对接人列表或 '全部'
        default_persons: 默认对接人列表
    """
    log_dir.mkdir(parents=True, exist_ok=True)

    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    log_name = f"{timestamp}_lx-zhutichaibiao_处理日志.md"
    log_path = log_dir / log_name

    total_rows = 0
    total_files = 0
    total_unmatch = 0

    lines = []
    lines.append(f"# lx-zhutichaibiao 拆分处理日志")
    lines.append(f"")
    lines.append(f"**处理时间**: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**拆分模式**: {split_mode}")
    lines.append(f"**对接人**: {person_list if isinstance(person_list, str) else ', '.join(person_list)}")
    lines.append(f"**处理文件数**: {len(log_entries)}")
    lines.append(f"")

    for i, entry in enumerate(log_entries, 1):
        fp = entry['file_path']
        stats = entry.get('stats', {})
        unmatch_count = entry.get('unmatch_count', 0)
        file_rows = sum(stats.values())

        total_rows += file_rows
        total_files += len(stats)
        total_unmatch += unmatch_count

        lines.append(f"## 文件 {i}: {fp.name}")
        lines.append(f"")
        lines.append(f"- **输出 ZIP**: {entry['zip_path'].name if entry['zip_path'] else 'N/A'}")
        lines.append(f"- **原表存档**: {entry['archive_path'].name if entry['archive_path'] else 'N/A'}")
        lines.append(f"- **运营主体/品牌数**: {len(stats)}")
        lines.append(f"- **有效数据行**: {file_rows}")
        lines.append(f"- **未匹配行**: {unmatch_count}")
        lines.append(f"")

        if stats:
            lines.append(f"### 各主体行数")
            lines.append(f"")
            lines.append(f"| 运营主体/品牌 | 数据行数 |")
            lines.append(f"|-------------|---------|")
            for operator, row_count in sorted(stats.items(), key=lambda x: -x[1]):
                lines.append(f"| {operator} | {row_count} |")
            lines.append(f"")

    lines.append(f"---")
    lines.append(f"")
    lines.append(f"## 汇总")
    lines.append(f"")
    lines.append(f"| 指标 | 数值 |")
    lines.append(f"|------|------|")
    lines.append(f"| 处理文件数 | {len(log_entries)} |")
    lines.append(f"| 生成主体/品牌文件 | {total_files} |")
    lines.append(f"| 有效数据总行数 | {total_rows} |")
    lines.append(f"| 未匹配总行数 | {total_unmatch} |")
    lines.append(f"")

    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"\n处理日志: {log_path.relative_to(log_dir.parent)}")
    return log_path


# ========================================
# 主函数
# ========================================

def main():
    print("=" * 60)
    print("按运营主体/城市/品牌拆表工具（独立版）")
    print("=" * 60)

    args = parse_args()
    force_config = args.config
    config = load_config()

    if not config or force_config:
        config = interactive_setup()
        if not config:
            return

    project_root = Path(config["项目根目录"])
    work_dir = Path(config["工作目录"])
    default_persons = config.get("默认对接人", DEFAULT_PERSONS)
    if isinstance(default_persons, str):
        default_persons = [default_persons]

    # 确保目录存在
    todo_dir = work_dir / "输入"
    done_dir = work_dir / "输出"
    archive_dir = work_dir / "原表存档"

    for dir_path in [todo_dir, done_dir, archive_dir]:
        dir_path.mkdir(parents=True, exist_ok=True)

    # 查找待拆文件
    todo_files = list(todo_dir.glob("*.xlsx")) + list(todo_dir.glob("*.xlsm"))
    if not todo_files:
        print(f"\n⚠️ 待拆目录中没有 Excel 文件: {todo_dir}")
        return

    # 加载公司库码表
    try:
        full_mapping = load_company_mabiao()
        if not full_mapping:
            print("\n❌ 无法从 lx_shujuku 加载公司库码表")
            return
    except Exception as e:
        print(f"❌ 加载公司库码表失败: {e}")
        return

    print(f"\n待拆文件数: {len(todo_files)}")
    print("码表来源: lx_shujuku / operator_brand")
    print(f"默认对接人: {default_persons}")
    print(f"所有对接人: {sorted(full_mapping['all_persons'])}")

    print("\n请选择拆分选项:")
    print("  1. 拆分维度: 运营主体 / 城市 / 品牌→运营主体 / 纯品牌")
    print("  2. 对接人范围: 默认对接人/全部")
    print("  3. 保留sheet: 指定不拆分的sheet（可选）")

    # 拆分维度
    if args.mode is not None:
        mode_map = {2: '城市', 3: '品牌', 4: '纯品牌'}
        split_mode = mode_map.get(args.mode, '运营主体')
        print(f"  → 使用命令行参数: 拆分维度={split_mode}")
    else:
        mode_input = input(f"\n拆分维度（1=运营主体，2=城市，3=品牌→运营主体，4=纯品牌，默认 运营主体）: ").strip()
        mode_map = {'2': '城市', '3': '品牌', '4': '纯品牌'}
        split_mode = mode_map.get(mode_input, '运营主体')

    # 对接人范围
    if args.person is not None:
        if args.person.lower() == 'all':
            target_persons = '全部'
        else:
            target_persons = [p.strip() for p in args.person.split(',') if p.strip()]
        print(f"  → 使用命令行参数: 对接人={target_persons}")
    else:
        person_input = input(f"\n对接人范围（1=默认 {','.join(default_persons)}，2=全部，回车默认 1）: ").strip()
        if person_input == '2':
            target_persons = '全部'
        elif person_input == '1' or person_input == '':
            target_persons = default_persons
        else:
            target_persons = person_input

    # 保留 sheet
    if args.keep_sheets is not None:
        override_keep_sheets = [s.strip() for s in args.keep_sheets.split(',') if s.strip()]
        print(f"  → 使用命令行参数: 保留sheet={override_keep_sheets}")
    else:
        override_keep_sheets = None

    # 按对接人筛选
    filtered_mapping, person_list = filter_by_person(full_mapping, target_persons)
    print(f"\n筛选对接人: {person_list}")
    print(f"筛选后运营主体: {len(filtered_mapping['all_zhuti'])} 个")

    # 处理每个文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_entries = []
    for file_path in todo_files:
        file_config = get_file_config(config, file_path.name)

        if override_keep_sheets:
            file_config["保留 sheet"] = override_keep_sheets
            print(f"\n  使用指定保留sheet: {override_keep_sheets}")

        col_info = detect_columns(file_path, file_config)
        result_files, stats, unmatch_data, header_info = process_file_optimized(
            file_path, filtered_mapping, file_config, split_mode, col_info,
        )

        zip_path = None
        archive_path = None
        if result_files:
            zip_path = pack_results(result_files, stats, filtered_mapping, file_path, done_dir, person_list, split_mode)

            archive_path = archive_dir / f"{timestamp}_{file_path.name}"
            shutil.move(str(file_path), str(archive_path))
            print(f"原表存档: {archive_path.name}")

        unmatch_count = sum(len(rows) for rows in unmatch_data.values()) if unmatch_data else 0
        if unmatch_data:
            output_unmatch_report(unmatch_data, file_path, done_dir, header_info)

        log_entries.append({
            'file_path': file_path,
            'zip_path': zip_path,
            'archive_path': archive_path,
            'stats': stats,
            'unmatch_count': unmatch_count,
        })

    # 写入处理日志
    log_dir = work_dir / "处理日志"
    write_processing_log(log_dir, log_entries, split_mode, person_list, default_persons)

    print("\n" + "=" * 60)
    print("拆分完成")
    print("=" * 60)


if __name__ == "__main__":
    main()
