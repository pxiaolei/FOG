"""
LX Skill 套件共享 Excel 工具

提供导入/拆表类 Skill 共享的 Excel 处理函数：
1. 公司库码表映射加载 (load_mabiao)
2. 列名模糊匹配 (find_column)
3. 样式复制 (copy_cell_style)
4. 表头检测 (detect_columns)
5. 拆分逻辑 (split_excel_by_mapping)

安全说明：
- 这些工具不涉及数据库操作
- 码表映射来自 lx_shujuku 查询 dataReporting.operator_brand，不读取本地 Excel 码表
"""

from copy import copy
import sys
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from lxx_share.utils import get_logger

logger = get_logger("lxx.excel_utils")

# ========================================
# 常量定义
# ========================================

# 城市字段列表
CITY_FIELDS = ["城市", "城市名称", "注册城市", "所属城市", "所属城市名称", "城市名", "city_name"]

# 品牌字段列表
BRAND_FIELDS = ["品牌", "品牌名称", "商家", "商家名称", "合作品牌", "合作商家", "运力品牌", "brand_name"]

# 对接人字段列表
PERSON_FIELDS = ["对接人", "对接人员", "负责人", "运营对接人"]

# 运营主体字段列表
OPERATOR_FIELDS = ["运营主体", "新 - 运营主体", "主体", "运营方"]

# 第二主体字段列表
SECOND_OPERATOR_FIELDS = ["第二主体", "第二运营主体", "备用主体"]


# ========================================
# 公共函数
# ========================================

def find_column(headers: List[str], target_fields: List[str]) -> Optional[int]:
    """
    模糊匹配列名

    Args:
        headers: 表头列表
        target_fields: 目标字段名列表（支持多个变体）

    Returns:
        列索引（从 1 开始），未找到返回 None
    """
    for idx, header in enumerate(headers, 1):
        if header:
            header_str = str(header).strip()
            for target in target_fields:
                if target in header_str or header_str in target:
                    return idx
    return None


def find_all_columns(headers: List[str], target_fields: List[str]) -> List[int]:
    """
    模糊匹配所有出现的列

    Args:
        headers: 表头列表
        target_fields: 目标字段名列表

    Returns:
        列索引列表（从 1 开始）
    """
    result = []
    for idx, header in enumerate(headers, 1):
        if header:
            header_str = str(header).strip()
            for target in target_fields:
                if target in header_str:
                    result.append(idx)
                    break
    return result


def copy_cell_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def detect_columns(file_path: str, file_config: Dict[str, Any]) -> Dict[str, Any]:
    """
    检测文件中有哪些关键列及其位置

    Args:
        file_path: Excel 文件路径
        file_config: 文件配置，包含城市字段、品牌字段等

    Returns:
        检测结果字典
    """
    wb = load_workbook(file_path, data_only=True)

    has_city = False
    has_brand = False
    city_col_idx = None
    brand_col_idx = None
    header_rows = 1

    city_fields = file_config.get("城市字段", CITY_FIELDS)
    brand_fields = file_config.get("品牌字段", BRAND_FIELDS)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state != 'visible':
            continue

        for row_idx in range(1, min(5, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    col_str = str(cell_value).strip()
                    if col_str in city_fields and city_col_idx is None:
                        has_city = True
                        city_col_idx = col_idx
                        header_rows = max(header_rows, row_idx)
                    if col_str in brand_fields and brand_col_idx is None:
                        has_brand = True
                        brand_col_idx = col_idx
                        header_rows = max(header_rows, row_idx)

    wb.close()

    return {
        'has_city': has_city,
        'has_brand': has_brand,
        'city_col_idx': city_col_idx,
        'brand_col_idx': brand_col_idx,
        'header_rows': header_rows
    }


def _ensure_lx_shujuku_path() -> None:
    """确保可 import lx_shujuku 包。"""
    for parent in Path(__file__).resolve().parents:
        scripts_dir = parent / "lx_shujuku" / "scripts"
        if scripts_dir.is_dir():
            if str(scripts_dir) not in sys.path:
                sys.path.insert(0, str(scripts_dir))
            return
    raise RuntimeError("未找到 lx_shujuku/scripts，无法加载公司库码表")


def load_mabiao(mabiao_path: Optional[str] = None, include_second_zhuti: bool = True) -> Optional[Dict[str, Any]]:
    """
    加载公司库码表，返回完整映射关系。

    保留 mabiao_path / include_second_zhuti 参数用于兼容旧调用；函数不再读取本地 Excel。
    """
    if mabiao_path:
        logger.warning("load_mabiao() 已改为读取 lx_shujuku.operator_brand，传入的本地码表路径会被忽略")
    if include_second_zhuti is False:
        logger.warning("公司库 operator_brand 当前不包含第二主体字段，include_second_zhuti=False 不改变返回结果")

    try:
        _ensure_lx_shujuku_path()
        from lx_shujuku import create_client

        logger.info("从 lx_shujuku 加载公司库 operator_brand 码表")
        return create_client().load_mabiao_mapping()
    except Exception as e:
        logger.error(f"加载公司库码表失败: {e}")
        return None


def load_city_operator_map(mabiao_path: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """兼容旧入口：返回公司库 operator_brand 构建的码表映射。"""
    return load_mabiao(mabiao_path)


def filter_by_person(mapping: Dict[str, Any], target_persons) -> Tuple[Dict[str, Any], List[str]]:
    """
    按对接人筛选映射

    Args:
        mapping: load_mabiao 返回的映射
        target_persons: 目标对接人列表或 "全部"

    Returns:
        (筛选后的映射, 对接人列表)
    """
    if target_persons == "全部" or target_persons == ["全部"]:
        return mapping, mapping['all_persons']

    if isinstance(target_persons, str):
        person_list = [p.strip() for p in target_persons.split(",")]
    else:
        person_list = target_persons

    # 筛选运营主体
    filtered_zhuti = set()
    for zhuti, persons in mapping['zhuti_to_person'].items():
        if any(p in person_list for p in persons):
            filtered_zhuti.add(zhuti)

    # 筛选城市
    filtered_cities = set()
    for city, zhuti_list in mapping['city_to_zhuti'].items():
        if any(z in filtered_zhuti for z in zhuti_list):
            filtered_cities.add(city)

    # 筛选品牌
    filtered_brands = set()
    for brand, zhuti_list in mapping['brand_to_zhuti'].items():
        if any(z in filtered_zhuti for z in zhuti_list):
            filtered_brands.add(brand)

    filtered_mapping = {
        'city_to_zhuti': {c: [z for z in zs if z in filtered_zhuti]
                          for c, zs in mapping['city_to_zhuti'].items() if c in filtered_cities},
        'brand_to_zhuti': {b: [z for z in zs if z in filtered_zhuti]
                           for b, zs in mapping['brand_to_zhuti'].items() if b in filtered_brands},
        'brand_city_to_zhuti': {k: [z for z in zs if z in filtered_zhuti]
                                for k, zs in mapping['brand_city_to_zhuti'].items()
                                if k[0] in filtered_brands and k[1] in filtered_cities},
        'zhuti_to_person': {z: ps for z, ps in mapping['zhuti_to_person'].items() if z in filtered_zhuti},
        'all_zhuti': list(filtered_zhuti),
        'all_cities': list(filtered_cities),
        'all_brands': list(filtered_brands),
        'all_persons': person_list
    }

    return filtered_mapping, person_list


def get_split_mode_info(split_mode: str, col_info: Dict[str, Any], file_config: Dict[str, Any]) -> Dict[str, Any]:
    """
    获取拆分模式信息，用于日志输出

    Args:
        split_mode: 拆分模式 ('运营主体', '城市', '品牌', '纯品牌')
        col_info: detect_columns 返回的信息
        file_config: 文件配置

    Returns:
        信息字典
    """
    info = {
        'mode': split_mode,
        'has_city': col_info.get('has_city', False),
        'has_brand': col_info.get('has_brand', False),
    }

    if split_mode == '城市' and not info['has_city']:
        info['error'] = '缺少城市列，无法按城市拆分'
    elif split_mode == '品牌' and not info['has_brand']:
        info['error'] = '缺少品牌列，无法按品牌拆分'

    return info
